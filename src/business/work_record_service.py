# src/business/work_record_service.py - 작업 레코드 비즈니스 로직

from datetime import datetime, timedelta
from typing import List, Dict, Any
from ..database.db_manager import db
from ..database.models import WorkRecord
from .calculations import calculate_record_manpower, calculate_total_manpower
from ..utils.logger import logger


class WorkRecordService:
    """작업 레코드 비즈니스 로직 처리"""
    
    def __init__(self):
        self.db = db
    
    def _to_camel_case(self, snake_dict: Dict[str, Any]) -> Dict[str, Any]:
        """스네이크 케이스를 카멜 케이스로 변환"""
        camel_dict = {}
        for key, value in snake_dict.items():
            # 스네이크 케이스 -> 카멜 케이스
            parts = key.split('_')
            camel_key = parts[0] + ''.join(word.capitalize() for word in parts[1:])
            camel_dict[camel_key] = value
        return camel_dict
    
    def create_empty_records(self, count: int = 10) -> List[WorkRecord]:
        """빈 레코드 생성"""
        return [WorkRecord(record_number=i+1) for i in range(count)]
    
    def get_records_for_date(self, date: str) -> List[Dict[str, Any]]:
        """
        특정 날짜의 작업 레코드 조회
        인원 자동 계산 포함
        """
        records = self.db.load_work_records(date)
        
        # 레코드가 없으면 빈 레코드 10개 생성
        if not records:
            records = self.create_empty_records()
        
        # 인원 계산 및 딕셔너리 변환
        result = []
        for record in records:
            record_dict = record.to_dict()
            # 스네이크 케이스 -> 카멜 케이스 변환
            record_dict = self._to_camel_case(record_dict)
            # 인원 자동 계산
            record_dict['manpower'] = calculate_record_manpower(
                record.leader, 
                record.teammates
            )
            result.append(record_dict)
        
        return result
    
    def save_records_for_date(self, date: str, records_data: List[Dict[str, Any]], username: str) -> dict:
        """특정 날짜의 작업 레코드 저장.
        반환: {'success': True} 또는 {'success': False, 'message': str}
        """
        try:
            # 딕셔너리를 WorkRecord 객체로 변환
            records = []
            for i, data in enumerate(records_data):
                record = WorkRecord(
                    record_number=i + 1,
                    date=date,
                    contract_number=data.get('contractNumber', '').upper(),  # 자동 대문자
                    company=data.get('company', ''),
                    ship_name=data.get('shipName', ''),
                    engine_model=data.get('engineModel', '').upper(),  # 자동 대문자
                    work_content=data.get('workContent', ''),
                    location=data.get('location', ''),
                    leader=data.get('leader', ''),
                    teammates=data.get('teammates', '')
                )

                # 인원 자동 계산
                record.manpower = calculate_record_manpower(record.leader, record.teammates)
                records.append(record)

            # DB에 저장
            success = self.db.save_work_records(date, records, username)

            if success:
                logger.info(f"작업 레코드 저장 성공: {date}, 사용자: {username}")
                return {'success': True}
            else:
                return {'success': False, 'message': 'DB 저장 실패'}

        except Exception as e:
            logger.error(f"작업 레코드 저장 오류: {e}")
            return {'success': False, 'message': str(e)}
    
    def _get_last_working_day(self, current_date: str) -> str:
        """주어진 날짜 이전의 마지막 평일(주말·공휴일 제외) 반환.
        월요일이면 금요일, 연휴 다음날이면 연휴 시작 전 마지막 평일."""
        import json
        from pathlib import Path

        # 공휴일 목록 로드 (config/holidays.json)
        holidays: set = set()
        try:
            holidays_path = Path(__file__).parent.parent.parent / "config" / "holidays.json"
            with open(holidays_path, 'r', encoding='utf-8') as f:
                holidays = set(json.load(f).keys())
        except Exception as e:
            logger.warning(f"공휴일 데이터 로드 실패 (무시): {e}")

        date_obj = datetime.strptime(current_date, '%Y-%m-%d')
        candidate = date_obj - timedelta(days=1)

        # 최대 60일 전까지 탐색 (긴 연휴도 커버)
        for _ in range(60):
            candidate_str = candidate.strftime('%Y-%m-%d')
            # 평일(월~금, weekday 0~4)이고 공휴일 아니면 → 마지막 근무일
            if candidate.weekday() < 5 and candidate_str not in holidays:
                return candidate_str
            candidate -= timedelta(days=1)

        # fallback: 그냥 어제
        return (date_obj - timedelta(days=1)).strftime('%Y-%m-%d')

    def get_yesterday_records(self, current_date: str):
        """마지막 평일 작업 레코드 불러오기 (주말·공휴일 건너뜀).
        반환: (records: List[Dict], loaded_date: str)"""
        def _make_empty(n):
            return {
                'recordNumber': n, 'contractNumber': '', 'company': '',
                'shipName': '', 'engineModel': '', 'workContent': '',
                'location': '', 'leader': '', 'teammates': '', 'manpower': 0
            }

        try:
            target_date = self._get_last_working_day(current_date)
            target_records = self.get_records_for_date(target_date)

            # 빈 레코드 제외 (실제 데이터가 있는 것만)
            active_records = [
                r for r in target_records
                if (r.get('contractNumber') or r.get('company') or
                    r.get('shipName')        or r.get('engineModel') or
                    r.get('workContent')     or r.get('location') or
                    r.get('leader')          or r.get('teammates'))
            ]

            # record_number 재설정 (1부터 순차적으로)
            for i, record in enumerate(active_records):
                record['recordNumber'] = i + 1

            # 10개 미만이면 빈 레코드로 채우기
            while len(active_records) < 10:
                active_records.append(_make_empty(len(active_records) + 1))

            valid_count = len([r for r in active_records if r.get('company') or r.get('shipName')])
            logger.info(f"마지막 평일 작업 불러오기 성공: {target_date}, 유효 {valid_count}건")

            return active_records, target_date

        except Exception as e:
            logger.error(f"마지막 평일 작업 불러오기 오류: {e}")
            fallback_date = (
                datetime.strptime(current_date, '%Y-%m-%d') - timedelta(days=1)
            ).strftime('%Y-%m-%d')
            return [_make_empty(i + 1) for i in range(10)], fallback_date
    
    def generate_report(self, date: str, username: str) -> Dict[str, Any]:
        """작업 현황 보고서 생성"""
        records = self.get_records_for_date(date)
        
        # 활성 레코드만 필터링 (데이터가 있는 레코드)
        active_records = [
            r for r in records 
            if r.get('contractNumber') or r.get('company') or r.get('shipName')
        ]
        
        # 총 인원 계산
        total_manpower = sum(r.get('manpower', 0) for r in active_records)
        
        report = {
            'date': date,
            'username': username,
            'totalRecords': len(active_records),
            'totalManpower': total_manpower,
            'records': active_records
        }
        
        return report
    
    def get_date_list(self, start_date: str = None, end_date: str = None) -> List[str]:
        """레코드가 있는 날짜 목록"""
        return self.db.get_dates_with_records(start_date, end_date)
    
    def export_to_excel(self, date: str, output_path: str) -> bool:
        """Excel 파일로 내보내기"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            records = self.get_records_for_date(date)
            
            # 워크북 생성
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"작업현황_{date}"
            
            # 헤더
            headers = ['순번', '계약번호', '선사', '선명', '엔진모델', 
                      '작업내용', '장소', '작업자', '인원', '동반자']
            
            # 헤더 스타일
            header_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            header_font = Font(bold=True)
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # 데이터
            for row_idx, record in enumerate(records, 2):
                # camelCase 변환 후이므로 'recordNumber' 키 사용 (snake_case 폴백 포함)
                ws.cell(row=row_idx, column=1, value=record.get('recordNumber') or record.get('record_number'))
                ws.cell(row=row_idx, column=2, value=record.get('contractNumber'))
                ws.cell(row=row_idx, column=3, value=record.get('company'))
                ws.cell(row=row_idx, column=4, value=record.get('shipName'))
                ws.cell(row=row_idx, column=5, value=record.get('engineModel'))
                ws.cell(row=row_idx, column=6, value=record.get('workContent'))
                ws.cell(row=row_idx, column=7, value=record.get('location'))
                ws.cell(row=row_idx, column=8, value=record.get('leader'))
                ws.cell(row=row_idx, column=9, value=record.get('manpower'))
                ws.cell(row=row_idx, column=10, value=record.get('teammates'))
            
            # 총합 행
            total_row = len(records) + 2
            ws.cell(row=total_row, column=8, value="총 인원")
            ws.cell(row=total_row, column=9, value=sum(r.get('manpower', 0) for r in records))
            
            # 저장
            wb.save(output_path)
            logger.info(f"Excel 내보내기 성공: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Excel 내보내기 실패: {e}")
            return False


# 싱글톤 인스턴스
work_record_service = WorkRecordService()
