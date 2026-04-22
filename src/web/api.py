# src/web/api.py - 웹 API (Python ↔ JavaScript)

import difflib
import json
import eel
import re
import threading
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
from ..business.work_record_service import work_record_service
from ..business.calculations import separate_workers, split_manpower_by_type
from ..database.db_manager import db
from ..database.auth_manager import auth_manager
from ..sync.cloud_sync import cloud_sync
from ..utils.logger import logger
from ..utils.config import config
from ..utils.settings_manager import settings_manager
from ..utils.path_manager import path_manager
from ..utils.update_manager import update_manager
from ..utils.telegram_notifier import telegram_notifier
from ..utils.erp_macro import erp_macro


# ============================================================================
# 백그라운드 스레드 추적 (앱 종료 전 graceful wait)
# ============================================================================
_pending_threads: list = []
_pending_threads_lock = threading.Lock()

# ERP 팝업 컨텍스트 캐시 (open_erp_input_window → get_erp_popup_context 공유)
_erp_popup_context: dict = {}
_erp_popup_proc: Optional[object] = None   # 현재 실행 중인 Chrome 팝업 프로세스


def _start_tracked_thread(target, args=(), daemon: bool = True) -> threading.Thread:
    """스레드 시작 + 전역 리스트에 등록 (완료된 스레드는 자동 정리)"""
    t = threading.Thread(target=target, args=args, daemon=daemon)
    with _pending_threads_lock:
        _pending_threads[:] = [x for x in _pending_threads if x.is_alive()]
        _pending_threads.append(t)
    t.start()
    return t


def wait_pending_threads(timeout: float = 5.0) -> None:
    """앱 종료 전 최대 timeout초 대기 (main.py finally 블록에서 호출)"""
    with _pending_threads_lock:
        threads = list(_pending_threads)
    for t in threads:
        t.join(timeout=timeout)
    logger.info(f"백그라운드 스레드 대기 완료 ({len(threads)}건)")


def _date_to_md(date_str: str) -> str:
    """'YYYY-MM-DD' 문자열을 'M/D' 형식으로 변환. 형식이 잘못된 경우 빈 문자열 반환."""
    if not date_str:
        return ''
    parts = date_str.split('-')
    if len(parts) != 3:
        return ''
    try:
        return f"{int(parts[1])}/{int(parts[2])}"
    except (ValueError, IndexError):
        return ''


# ============================================================================
# 연결 확인 (스플래시 화면용)
# ============================================================================

@eel.expose
def ping() -> bool:
    """Python 백엔드 연결 확인용 (스플래시 → 로그인 전환 트리거)"""
    return True


@eel.expose
def open_external_url(url: str) -> bool:
    """시스템 기본 브라우저/앱으로 URL 열기 (Eel 앱 내 target=_blank 대체)"""
    try:
        import os as _os
        _os.startfile(url)
        return True
    except Exception:
        try:
            import subprocess as _subprocess
            _subprocess.Popen(['cmd', '/c', 'start', '', url], shell=False)
            return True
        except Exception as e:
            logger.error(f"외부 URL 열기 실패: {url} - {e}")
            return False


# ============================================================================
# 인증 관리
# ============================================================================

@eel.expose
def authenticate(user_id: str, password: str) -> Dict[str, Any]:
    """사용자 인증"""
    try:
        # #11 — 입력 길이 제한
        if not user_id or not password:
            return {'success': False, 'message': '아이디와 비밀번호를 입력해주세요.'}
        if len(user_id) > 30 or len(password) > 200:
            return {'success': False, 'message': '입력값이 너무 깁니다.'}
        result = auth_manager.authenticate(user_id, password)
        
        if result is None:
            return {'success': False, 'message': '아이디 또는 비밀번호가 일치하지 않습니다.'}

        if 'error' in result:
            err = result['error']
            if err == 'USER_NOT_FOUND':
                return {'success': False, 'message': '없는 사용자입니다.'}
            elif err == 'WRONG_PASSWORD':
                return {'success': False, 'message': '비밀번호가 일치하지 않습니다.'}
            else:
                return {'success': False, 'message': err}
        
        # 로그인 성공
        return {
            'success': True,
            'user': {
                'user_id': result['user_id'],
                'full_name': result['full_name'],
                'role': result['role'],
                'client_version': result.get('client_version', ''),
                'tray_mode': bool(result.get('tray_mode', 0)),
                'leave_report_edit': bool(result.get('leave_report_edit', 0)),
                'can_write': bool(result.get('can_write', 0)),
                'erp_input': bool(result.get('erp_input', 0))
            }
        }
    except Exception as e:
        logger.error(f"인증 오류: {e}")
        return {'success': False, 'message': '인증 중 오류가 발생했습니다.'}


@eel.expose
def create_remember_token(user_id: str) -> Dict[str, Any]:
    """자동 로그인 토큰 생성"""
    try:
        token = auth_manager.create_remember_token(user_id)
        if token:
            return {'success': True, 'token': token}
        return {'success': False, 'message': '토큰 생성 실패'}
    except Exception as e:
        logger.error(f"Remember token 생성 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def auto_login(token: str) -> Dict[str, Any]:
    """세션 토큰으로 자동 로그인"""
    try:
        user = auth_manager.validate_remember_token(token)
        if user:
            days_remaining = auth_manager.get_token_days_remaining(token)
            return {'success': True, 'user': user, 'days_remaining': days_remaining}
        return {'success': False, 'message': '토큰이 만료되었거나 유효하지 않습니다.'}
    except Exception as e:
        logger.error(f"자동 로그인 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def clear_remember_token(token: str) -> Dict[str, Any]:
    """자동 로그인 토큰 삭제 (로그아웃 시)"""
    try:
        auth_manager.clear_remember_token(token)
        return {'success': True}
    except Exception as e:
        return {'success': False}


@eel.expose
def register_user(user_id: str, password: str, full_name: str) -> Dict[str, Any]:
    """사용자 등록 요청"""
    try:
        # #11 — 입력 길이 제한
        if len(user_id) > 30 or len(password) > 200 or len(full_name) > 50:
            return {'success': False, 'message': '입력값이 허용 길이를 초과했습니다.'}
        if len(password) < 4:
            return {'success': False, 'message': '비밀번호는 4자 이상이어야 합니다.'}
        success = auth_manager.register_request(user_id, password, full_name)
        
        if success:
            return {
                'success': True,
                'message': '등록 요청이 전송되었습니다. 관리자 승인 후 이용 가능합니다.'
            }
        else:
            return {
                'success': False,
                'message': '이미 존재하는 아이디입니다.'
            }
    except Exception as e:
        logger.error(f"등록 요청 오류: {e}")
        return {'success': False, 'message': '등록 요청 중 오류가 발생했습니다.'}


# ============================================================================
# 관리자 기능
# ============================================================================

@eel.expose
def admin_get_all_users(admin_id: str = '') -> List[Dict[str, Any]]:
    """모든 사용자 조회 (관리자)"""
    try:
        user = auth_manager.get_user(admin_id) if admin_id else None
        if not user or user.get('role') != 'admin':
            return []
        users = auth_manager.get_all_users()
        return users
    except Exception as e:
        logger.error(f"사용자 목록 조회 오류: {e}")
        return []


@eel.expose
def admin_get_pending_requests(admin_id: str = '') -> List[Dict[str, Any]]:
    """승인 대기 요청 조회 (관리자)"""
    try:
        user = auth_manager.get_user(admin_id) if admin_id else None
        if not user or user.get('role') != 'admin':
            return []
        requests = auth_manager.get_pending_requests()
        return requests
    except Exception as e:
        logger.error(f"대기 요청 조회 오류: {e}")
        return []


@eel.expose
def admin_approve_user(user_id: str, admin_id: str) -> Dict[str, Any]:
    """사용자 승인 (관리자)"""
    try:
        success = auth_manager.approve_user(user_id, admin_id)
        return {
            'success': success,
            'message': '사용자가 승인되었습니다.' if success else '승인 실패'
        }
    except Exception as e:
        logger.error(f"사용자 승인 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def admin_reject_user(user_id: str, admin_id: str, note: str = '') -> Dict[str, Any]:
    """사용자 거부 (관리자)"""
    try:
        success = auth_manager.reject_user(user_id, admin_id, note)
        return {
            'success': success,
            'message': '사용자가 거부되었습니다.' if success else '거부 실패'
        }
    except Exception as e:
        logger.error(f"사용자 거부 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def admin_delete_user(user_id: str, admin_id: str) -> Dict[str, Any]:
    """사용자 퇴사 처리 (관리자) — soft delete"""
    try:
        success = auth_manager.delete_user(user_id, admin_id)
        return {
            'success': success,
            'message': '퇴사 처리되었습니다.' if success else '퇴사 처리 실패'
        }
    except Exception as e:
        logger.error(f"사용자 퇴사 처리 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def admin_update_user_status(user_id: str, status: str, admin_id: str) -> Dict[str, Any]:
    """사용자 상태 변경 (관리자)"""
    try:
        success = auth_manager.update_user_status(user_id, status, admin_id)
        return {
            'success': success,
            'message': '사용자 상태가 변경되었습니다.' if success else '상태 변경 실패'
        }
    except Exception as e:
        logger.error(f"사용자 상태 변경 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def select_folder_path() -> Dict[str, Any]:
    """폴더 선택 다이얼로그 (Windows Shell API - Embedded Python 호환)"""
    try:
        import ctypes
        import ctypes.wintypes

        # Embedded Python에는 Tkinter가 없으므로 Windows Shell32 API 직접 사용
        BIF_RETURNONLYFSDIRS = 0x0001
        BIF_NEWDIALOGSTYLE   = 0x0040  # 새 스타일 폴더 선택창

        class BROWSEINFOW(ctypes.Structure):
            _fields_ = [
                ('hwndOwner',      ctypes.wintypes.HWND),
                ('pidlRoot',       ctypes.c_void_p),
                ('pszDisplayName', ctypes.c_wchar_p),
                ('lpszTitle',      ctypes.c_wchar_p),
                ('ulFlags',        ctypes.wintypes.UINT),
                ('lpfn',           ctypes.c_void_p),
                ('lParam',         ctypes.c_long),
                ('iImage',         ctypes.c_int),
            ]

        shell32 = ctypes.windll.shell32
        ole32   = ctypes.windll.ole32
        ole32.CoInitialize(None)

        display_buf = ctypes.create_unicode_buffer(260)
        bi = BROWSEINFOW()
        bi.hwndOwner      = None
        bi.pidlRoot       = None
        bi.pszDisplayName = display_buf
        bi.lpszTitle      = '폴더 선택'
        bi.ulFlags        = BIF_RETURNONLYFSDIRS | BIF_NEWDIALOGSTYLE
        bi.lpfn           = None
        bi.lParam         = 0

        pidl = shell32.SHBrowseForFolderW(ctypes.byref(bi))
        if pidl:
            path_buf = ctypes.create_unicode_buffer(260)
            shell32.SHGetPathFromIDListW(pidl, path_buf)
            ole32.CoTaskMemFree(pidl)
            return {'success': True, 'path': path_buf.value}
        else:
            return {'success': False, 'path': None}

    except Exception as e:
        logger.error(f"폴더 선택 오류: {e}")
        return {'success': False, 'error': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def admin_get_paths(admin_id: str = '') -> Dict[str, Any]:
    """경로 조회 (관리자)"""
    try:
        user = auth_manager.get_user(admin_id) if admin_id else None
        if not user or user.get('role') != 'admin':
            return {'success': False, 'paths': {}}
        settings = settings_manager.get_current_settings()
        return {
            'success': True,
            'paths': {
                'local_db_path': settings.get('local_db_path', ''),
                'cloud_sync_path': settings.get('cloud_sync_path', ''),
                'backup_path': settings.get('backup_path', '')
            }
        }
    except Exception as e:
        logger.error(f"경로 조회 오류: {e}")
        return {'success': False, 'paths': {}}


@eel.expose
def admin_get_settings(admin_id: str = '') -> Dict[str, Any]:
    """설정 조회 (관리자)"""
    try:
        user = auth_manager.get_user(admin_id) if admin_id else None
        if not user or user.get('role') != 'admin':
            return {}
        return settings_manager.get_current_settings()
    except Exception as e:
        logger.error(f"설정 조회 오류: {e}")
        return {}


def _get_admin_user(admin_id: str = '') -> Optional[Dict[str, Any]]:
    user = auth_manager.get_user(admin_id) if admin_id else None
    if not user or user.get('role') != 'admin':
        return None
    return user


@eel.expose
def admin_get_owner_company_catalog(admin_id: str = '') -> Dict[str, Any]:
    """선사(owner_company) 목록과 선박/장비 요약 조회"""
    try:
        if not _get_admin_user(admin_id):
            return {
                'success': False,
                'message': '관리자 권한이 필요합니다.',
                'owners': [],
                'ownerSuggestions': [],
            }

        owner_map: Dict[str, Dict[str, Any]] = {}

        def _ensure_owner(owner_name: str) -> Dict[str, Any]:
            return owner_map.setdefault(owner_name, {
                'name': owner_name,
                'workRecordCount': 0,
                'holidayCount': 0,
                'projectCount': 0,
                'ships': {},
            })

        def _ensure_ship(owner: Dict[str, Any], ship_name: str) -> Dict[str, Any]:
            return owner['ships'].setdefault(ship_name, {
                'shipName': ship_name,
                'workRecordCount': 0,
                'holidayCount': 0,
                'projectCount': 0,
                'engineModels': set(),
            })

        with db.get_connection() as conn:
            work_rows = conn.execute('''
                SELECT company, ship_name, engine_model
                FROM work_records
                WHERE company IS NOT NULL AND company != ''
            ''').fetchall()
            board_rows = conn.execute('''
                SELECT company, ship_name, engine_model
                FROM board_projects
                WHERE company IS NOT NULL AND company != ''
            ''').fetchall()
            holiday_rows = conn.execute('''
                SELECT owner_company, ship_name
                FROM holiday_work_entries
                WHERE owner_company IS NOT NULL AND owner_company != ''
            ''').fetchall()

        for row in work_rows:
            owner_name = str(row['company'] or '').strip()
            if not owner_name:
                continue
            owner = _ensure_owner(owner_name)
            ship_name = _display_ship_label(row['ship_name'])
            ship = _ensure_ship(owner, ship_name)
            engine_model = str(row['engine_model'] or '').strip()
            if engine_model:
                ship['engineModels'].add(engine_model)
            ship['workRecordCount'] += 1
            owner['workRecordCount'] += 1

        for row in board_rows:
            owner_name = str(row['company'] or '').strip()
            if not owner_name:
                continue
            owner = _ensure_owner(owner_name)
            ship_name = _display_ship_label(row['ship_name'])
            ship = _ensure_ship(owner, ship_name)
            engine_model = str(row['engine_model'] or '').strip()
            if engine_model:
                ship['engineModels'].add(engine_model)
            ship['projectCount'] += 1
            owner['projectCount'] += 1

        for row in holiday_rows:
            owner_name = str(row['owner_company'] or '').strip()
            if not owner_name:
                continue
            owner = _ensure_owner(owner_name)
            ship_name = _display_ship_label(row['ship_name'])
            ship = _ensure_ship(owner, ship_name)
            ship['holidayCount'] += 1
            owner['holidayCount'] += 1

        owners = []
        for owner_name, owner in owner_map.items():
            ships = []
            for ship_name, ship in owner['ships'].items():
                ships.append({
                    'shipName': ship_name,
                    'workRecordCount': ship['workRecordCount'],
                    'holidayCount': ship['holidayCount'],
                    'projectCount': ship['projectCount'],
                    'totalCount': ship['workRecordCount'] + ship['holidayCount'] + ship['projectCount'],
                    'engineModels': sorted(list(ship['engineModels']), key=_mixed_locale_sort_key),
                })
            ships.sort(key=lambda item: _mixed_locale_sort_key(item['shipName']))
            owners.append({
                'name': owner_name,
                'workRecordCount': owner['workRecordCount'],
                'holidayCount': owner['holidayCount'],
                'projectCount': owner['projectCount'],
                'totalCount': owner['workRecordCount'] + owner['holidayCount'] + owner['projectCount'],
                'shipCount': len(ships),
                'ships': ships,
                'shipSuggestions': _build_merge_suggestions([item['shipName'] for item in ships]),
            })

        owners.sort(key=lambda item: _mixed_locale_sort_key(item['name']))
        return {
            'success': True,
            'owners': owners,
            'ownerSuggestions': _build_merge_suggestions([item['name'] for item in owners]),
        }
    except Exception as e:
        logger.error(f"선사 목록 조회 오류: {e}")
        return {
            'success': False,
            'message': '선사 목록 조회 중 오류가 발생했습니다.',
            'owners': [],
            'ownerSuggestions': [],
        }


@eel.expose
def admin_get_vendor_company_catalog(admin_id: str = '') -> Dict[str, Any]:
    """외주 업체 목록과 소속 인원 집계 조회"""
    try:
        if not _get_admin_user(admin_id):
            return {'success': False, 'message': '관리자 권한이 필요합니다.', 'vendors': []}

        vendor_map: Dict[str, Dict[str, Any]] = {}
        with db.get_connection() as conn:
            work_rows = conn.execute('''
                SELECT teammates
                FROM work_records
                WHERE teammates IS NOT NULL AND teammates != ''
            ''').fetchall()
            holiday_rows = conn.execute('''
                SELECT name, vendor_company, company
                FROM holiday_work_entries
                WHERE name IS NOT NULL AND name != ''
            ''').fetchall()

        for row in work_rows:
            vendor_workers = _extract_vendor_workers_from_teammates(row['teammates'] or '')
            for vendor_name, worker_names in vendor_workers.items():
                vendor = vendor_map.setdefault(vendor_name, {
                    'name': vendor_name,
                    'workRecordCount': 0,
                    'holidayCount': 0,
                    'workers': {},
                })
                vendor['workRecordCount'] += 1
                for worker_name in worker_names:
                    worker = vendor['workers'].setdefault(worker_name, {
                        'name': worker_name,
                        'workCount': 0,
                        'holidayCount': 0,
                    })
                    worker['workCount'] += 1

        for row in holiday_rows:
            vendor_name = str(row['vendor_company'] or row['company'] or '').strip()
            worker_name = _normalize_holiday_worker_name(row['name'] or '')
            if not vendor_name or not worker_name:
                continue
            vendor = vendor_map.setdefault(vendor_name, {
                'name': vendor_name,
                'workRecordCount': 0,
                'holidayCount': 0,
                'workers': {},
            })
            vendor['holidayCount'] += 1
            worker = vendor['workers'].setdefault(worker_name, {
                'name': worker_name,
                'workCount': 0,
                'holidayCount': 0,
            })
            worker['holidayCount'] += 1

        vendors = []
        for vendor_name, vendor in vendor_map.items():
            workers = []
            for worker_name, worker in vendor['workers'].items():
                workers.append({
                    'name': worker_name,
                    'workCount': worker['workCount'],
                    'holidayCount': worker['holidayCount'],
                    'totalCount': worker['workCount'] + worker['holidayCount'],
                })
            workers.sort(key=lambda item: _mixed_locale_sort_key(item['name']))
            vendors.append({
                'name': vendor_name,
                'workRecordCount': vendor['workRecordCount'],
                'holidayCount': vendor['holidayCount'],
                'workerCount': len(workers),
                'workers': workers,
                'workerSuggestions': _build_merge_suggestions([item['name'] for item in workers]),
            })

        vendors.sort(key=lambda item: _mixed_locale_sort_key(item['name']))
        return {'success': True, 'vendors': vendors}
    except Exception as e:
        logger.error(f"외주 업체 목록 조회 오류: {e}")
        return {'success': False, 'message': '외주 업체 목록 조회 중 오류가 발생했습니다.', 'vendors': []}


def _plan_merge_vendor_workers(vendor_company: str, source_names: List[str],
                               target_name: str) -> Dict[str, Any]:
    vendor_display = str(vendor_company or '').strip()
    normalized_sources = sorted({
        _normalize_holiday_worker_name(name)
        for name in (source_names or [])
        if _normalize_holiday_worker_name(name)
    })
    target_display = str(target_name or '').strip() or (normalized_sources[0] if normalized_sources else '')
    if not vendor_display:
        return {'success': False, 'message': '외주 업체명을 선택하세요.'}
    if len(normalized_sources) < 2:
        return {'success': False, 'message': '병합할 직원명을 2개 이상 선택하세요.'}
    if not target_display:
        return {'success': False, 'message': '대표 이름을 입력하세요.'}

    updates = []
    work_updates = 0
    holiday_updates = 0
    vendor_key = _normalize_company_label(vendor_display)

    with db.get_connection() as conn:
        cursor = conn.cursor()
        work_rows = cursor.execute('''
            SELECT id, teammates
            FROM work_records
            WHERE teammates IS NOT NULL AND teammates != ''
        ''').fetchall()
        for row in work_rows:
            new_teammates, changed = _replace_vendor_worker_names_in_teammates(
                row['teammates'] or '',
                vendor_display,
                normalized_sources,
                target_display,
            )
            if not changed:
                continue
            updates.append({
                'table': 'work_records',
                'id': row['id'],
                'old_fields': {'teammates': row['teammates'] or ''},
                'new_fields': {'teammates': new_teammates},
            })
            work_updates += 1

        holiday_rows = cursor.execute('''
            SELECT id, name, vendor_company, company
            FROM holiday_work_entries
            WHERE name IS NOT NULL AND name != ''
        ''').fetchall()
        for row in holiday_rows:
            row_vendor = str(row['vendor_company'] or row['company'] or '').strip()
            if _normalize_company_label(row_vendor) != vendor_key:
                continue
            if _normalize_holiday_worker_name(row['name'] or '') not in normalized_sources:
                continue
            updates.append({
                'table': 'holiday_work_entries',
                'id': row['id'],
                'old_fields': {
                    'name': row['name'] or '',
                    'vendor_company': row['vendor_company'] or '',
                    'company': row['company'] or '',
                },
                'new_fields': {
                    'name': target_display,
                    'vendor_company': vendor_display,
                    'company': vendor_display,
                },
            })
            holiday_updates += 1

    details = f"{', '.join(normalized_sources)} -> {target_display} (work={work_updates}, holiday={holiday_updates})"
    return {
        'success': True,
        'targetDisplay': target_display,
        'normalizedSources': normalized_sources,
        'updates': updates,
        'workUpdates': work_updates,
        'holidayUpdates': holiday_updates,
        'message': f'미리보기: 작업 {work_updates}건, 휴일 {holiday_updates}건 변경',
        'details': details,
    }


def _plan_merge_owner_companies(source_names: List[str], target_name: str) -> Dict[str, Any]:
    normalized_sources = sorted({
        _normalize_holiday_worker_name(name)
        for name in (source_names or [])
        if _normalize_holiday_worker_name(name)
    })
    target_display = str(target_name or '').strip() or (normalized_sources[0] if normalized_sources else '')
    source_keys = {_normalize_company_label(name) for name in normalized_sources}

    if len(normalized_sources) < 2:
        return {'success': False, 'message': '병합할 선사를 2개 이상 선택하세요.'}
    if not target_display:
        return {'success': False, 'message': '대표 선사명을 입력하세요.'}

    updates = []
    work_updates = 0
    project_updates = 0
    holiday_updates = 0

    with db.get_connection() as conn:
        cursor = conn.cursor()
        work_rows = cursor.execute('''
            SELECT id, company
            FROM work_records
            WHERE company IS NOT NULL AND company != ''
        ''').fetchall()
        for row in work_rows:
            if _normalize_company_label(row['company'] or '') not in source_keys:
                continue
            updates.append({
                'table': 'work_records',
                'id': row['id'],
                'old_fields': {'company': row['company'] or ''},
                'new_fields': {'company': target_display},
            })
            work_updates += 1

        board_rows = cursor.execute('''
            SELECT id, company
            FROM board_projects
            WHERE company IS NOT NULL AND company != ''
        ''').fetchall()
        for row in board_rows:
            if _normalize_company_label(row['company'] or '') not in source_keys:
                continue
            updates.append({
                'table': 'board_projects',
                'id': row['id'],
                'old_fields': {'company': row['company'] or ''},
                'new_fields': {'company': target_display},
            })
            project_updates += 1

        holiday_rows = cursor.execute('''
            SELECT id, owner_company
            FROM holiday_work_entries
            WHERE owner_company IS NOT NULL AND owner_company != ''
        ''').fetchall()
        for row in holiday_rows:
            if _normalize_company_label(row['owner_company'] or '') not in source_keys:
                continue
            updates.append({
                'table': 'holiday_work_entries',
                'id': row['id'],
                'old_fields': {'owner_company': row['owner_company'] or ''},
                'new_fields': {'owner_company': target_display},
            })
            holiday_updates += 1

    details = f"{', '.join(normalized_sources)} -> {target_display} (work={work_updates}, project={project_updates}, holiday={holiday_updates})"
    return {
        'success': True,
        'targetDisplay': target_display,
        'normalizedSources': normalized_sources,
        'updates': updates,
        'workUpdates': work_updates,
        'projectUpdates': project_updates,
        'holidayUpdates': holiday_updates,
        'message': f'미리보기: 작업 {work_updates}건, 등록 {project_updates}건, 휴일 {holiday_updates}건 변경',
        'details': details,
    }


def _plan_merge_owner_ships(owner_name: str, source_names: List[str], target_name: str) -> Dict[str, Any]:
    owner_display = str(owner_name or '').strip()
    owner_key = _normalize_company_label(owner_display)
    normalized_sources = sorted({
        _display_ship_label(name)
        for name in (source_names or [])
        if _display_ship_label(name)
    })
    target_display = _display_ship_label(target_name or (normalized_sources[0] if normalized_sources else ''))
    target_storage = _storage_ship_label(target_display)
    source_keys = {_normalize_ship_label(name) for name in normalized_sources}

    if not owner_key:
        return {'success': False, 'message': '선사를 먼저 선택하세요.'}
    if len(normalized_sources) < 2:
        return {'success': False, 'message': '병합할 선박을 2개 이상 선택하세요.'}
    if not target_display:
        return {'success': False, 'message': '대표 선박명을 입력하세요.'}

    updates = []
    work_updates = 0
    project_updates = 0
    holiday_updates = 0

    with db.get_connection() as conn:
        cursor = conn.cursor()
        work_rows = cursor.execute('''
            SELECT id, company, ship_name
            FROM work_records
            WHERE company IS NOT NULL AND company != ''
        ''').fetchall()
        for row in work_rows:
            if _normalize_company_label(row['company'] or '') != owner_key:
                continue
            if _normalize_ship_label(row['ship_name'] or '') not in source_keys:
                continue
            updates.append({
                'table': 'work_records',
                'id': row['id'],
                'old_fields': {'ship_name': row['ship_name'] or ''},
                'new_fields': {'ship_name': target_storage},
            })
            work_updates += 1

        board_rows = cursor.execute('''
            SELECT id, company, ship_name
            FROM board_projects
            WHERE company IS NOT NULL AND company != ''
        ''').fetchall()
        for row in board_rows:
            if _normalize_company_label(row['company'] or '') != owner_key:
                continue
            if _normalize_ship_label(row['ship_name'] or '') not in source_keys:
                continue
            updates.append({
                'table': 'board_projects',
                'id': row['id'],
                'old_fields': {'ship_name': row['ship_name'] or ''},
                'new_fields': {'ship_name': target_storage},
            })
            project_updates += 1

        holiday_rows = cursor.execute('''
            SELECT id, owner_company, ship_name
            FROM holiday_work_entries
            WHERE owner_company IS NOT NULL AND owner_company != ''
        ''').fetchall()
        for row in holiday_rows:
            if _normalize_company_label(row['owner_company'] or '') != owner_key:
                continue
            if _normalize_ship_label(row['ship_name'] or '') not in source_keys:
                continue
            updates.append({
                'table': 'holiday_work_entries',
                'id': row['id'],
                'old_fields': {'ship_name': row['ship_name'] or ''},
                'new_fields': {'ship_name': target_storage},
            })
            holiday_updates += 1

    details = f"{owner_display}: {', '.join(normalized_sources)} -> {target_display} (work={work_updates}, project={project_updates}, holiday={holiday_updates})"
    return {
        'success': True,
        'targetDisplay': target_display,
        'normalizedSources': normalized_sources,
        'updates': updates,
        'workUpdates': work_updates,
        'projectUpdates': project_updates,
        'holidayUpdates': holiday_updates,
        'message': f'미리보기: 작업 {work_updates}건, 등록 {project_updates}건, 휴일 {holiday_updates}건 변경',
        'details': details,
    }


@eel.expose
def admin_preview_merge_vendor_workers(vendor_company: str, source_names: List[str],
                                       target_name: str = '', admin_id: str = '') -> Dict[str, Any]:
    if not _get_admin_user(admin_id):
        return {'success': False, 'message': '관리자 권한이 필요합니다.'}
    return _plan_merge_vendor_workers(vendor_company, source_names, target_name)


@eel.expose
def admin_preview_merge_owner_companies(source_names: List[str], target_name: str = '',
                                        admin_id: str = '') -> Dict[str, Any]:
    if not _get_admin_user(admin_id):
        return {'success': False, 'message': '관리자 권한이 필요합니다.'}
    return _plan_merge_owner_companies(source_names, target_name)


@eel.expose
def admin_preview_merge_owner_ships(owner_name: str, source_names: List[str], target_name: str = '',
                                    admin_id: str = '') -> Dict[str, Any]:
    if not _get_admin_user(admin_id):
        return {'success': False, 'message': '관리자 권한이 필요합니다.'}
    return _plan_merge_owner_ships(owner_name, source_names, target_name)


@eel.expose
def admin_get_last_merge_undo(admin_id: str = '') -> Dict[str, Any]:
    if not _get_admin_user(admin_id):
        return {'success': False, 'message': '관리자 권한이 필요합니다.', 'available': False}
    snapshot = _load_last_merge_undo()
    if not snapshot or not snapshot.get('updates'):
        return {'success': True, 'available': False, 'summary': ''}
    return {
        'success': True,
        'available': True,
        'summary': str(snapshot.get('details') or snapshot.get('action') or '').strip(),
        'action': snapshot.get('action', ''),
        'target': snapshot.get('target', ''),
    }


@eel.expose
def admin_undo_last_merge(admin_id: str = '') -> Dict[str, Any]:
    if not _get_admin_user(admin_id):
        return {'success': False, 'message': '관리자 권한이 필요합니다.'}
    snapshot = _load_last_merge_undo()
    if not snapshot or not snapshot.get('updates'):
        return {'success': False, 'message': '되돌릴 최근 병합 내역이 없습니다.'}

    try:
        with db.get_connection() as conn:
            cursor = conn.cursor()
            restored = _apply_merge_updates(cursor, snapshot.get('updates', []), use_old_values=True)
        db.add_activity_log(
            admin_id,
            'undo_merge',
            str(snapshot.get('target') or ''),
            str(snapshot.get('details') or '')
        )
        _clear_last_merge_undo()
        return {
            'success': True,
            'message': f'최근 병합 되돌리기 완료: {restored}건 복원',
            'restoredCount': restored,
        }
    except Exception as e:
        logger.error(f"병합 되돌리기 오류: {e}")
        return {'success': False, 'message': '병합 되돌리기 중 오류가 발생했습니다.'}


@eel.expose
def admin_merge_vendor_workers(vendor_company: str, source_names: List[str],
                               target_name: str, admin_id: str = '') -> Dict[str, Any]:
    """특정 외주 업체 소속 직원명을 하나의 대표 이름으로 병합"""
    try:
        if not _get_admin_user(admin_id):
            return {'success': False, 'message': '관리자 권한이 필요합니다.'}
        plan = _plan_merge_vendor_workers(vendor_company, source_names, target_name)
        if not plan.get('success'):
            return plan

        with db.get_connection() as conn:
            cursor = conn.cursor()
            _apply_merge_updates(cursor, plan.get('updates', []), use_old_values=False)

        vendor_display = str(vendor_company or '').strip()
        _save_last_merge_undo(_build_undo_snapshot(
            'merge_vendor_workers',
            vendor_display,
            plan.get('details', ''),
            plan.get('updates', []),
        ))
        db.add_activity_log(admin_id, 'merge_vendor_workers', vendor_display, plan.get('details', ''))
        return {
            'success': True,
            'message': f"병합 완료: 작업 {plan.get('workUpdates', 0)}건, 휴일 {plan.get('holidayUpdates', 0)}건 수정",
            'workUpdates': plan.get('workUpdates', 0),
            'holidayUpdates': plan.get('holidayUpdates', 0),
        }
    except Exception as e:
        logger.error(f"외주 직원 병합 오류: {e}")
        return {'success': False, 'message': '외주 직원 병합 중 오류가 발생했습니다.'}


@eel.expose
def admin_merge_owner_companies(source_names: List[str], target_name: str,
                                admin_id: str = '') -> Dict[str, Any]:
    """중복 선사명을 하나의 대표 이름으로 병합"""
    try:
        if not _get_admin_user(admin_id):
            return {'success': False, 'message': '관리자 권한이 필요합니다.'}
        plan = _plan_merge_owner_companies(source_names, target_name)
        if not plan.get('success'):
            return plan

        with db.get_connection() as conn:
            cursor = conn.cursor()
            _apply_merge_updates(cursor, plan.get('updates', []), use_old_values=False)

        target_display = str(plan.get('targetDisplay') or target_name or '').strip()
        _save_last_merge_undo(_build_undo_snapshot(
            'merge_owner_companies',
            target_display,
            plan.get('details', ''),
            plan.get('updates', []),
        ))
        db.add_activity_log(admin_id, 'merge_owner_companies', target_display, plan.get('details', ''))
        return {
            'success': True,
            'message': f"병합 완료: 작업 {plan.get('workUpdates', 0)}건, 등록 {plan.get('projectUpdates', 0)}건, 휴일 {plan.get('holidayUpdates', 0)}건 수정",
            'workUpdates': plan.get('workUpdates', 0),
            'projectUpdates': plan.get('projectUpdates', 0),
            'holidayUpdates': plan.get('holidayUpdates', 0),
        }
    except Exception as e:
        logger.error(f"선사 병합 오류: {e}")
        return {'success': False, 'message': '선사 병합 중 오류가 발생했습니다.'}


@eel.expose
def admin_merge_owner_ships(owner_name: str, source_names: List[str], target_name: str,
                            admin_id: str = '') -> Dict[str, Any]:
    """선사 내부의 중복 선박명을 하나의 대표 이름으로 병합"""
    try:
        if not _get_admin_user(admin_id):
            return {'success': False, 'message': '관리자 권한이 필요합니다.'}
        plan = _plan_merge_owner_ships(owner_name, source_names, target_name)
        if not plan.get('success'):
            return plan

        with db.get_connection() as conn:
            cursor = conn.cursor()
            _apply_merge_updates(cursor, plan.get('updates', []), use_old_values=False)

        owner_display = str(owner_name or '').strip()
        _save_last_merge_undo(_build_undo_snapshot(
            'merge_owner_ships',
            owner_display,
            plan.get('details', ''),
            plan.get('updates', []),
        ))
        db.add_activity_log(admin_id, 'merge_owner_ships', owner_display, plan.get('details', ''))
        return {
            'success': True,
            'message': f"병합 완료: 작업 {plan.get('workUpdates', 0)}건, 등록 {plan.get('projectUpdates', 0)}건, 휴일 {plan.get('holidayUpdates', 0)}건 수정",
            'workUpdates': plan.get('workUpdates', 0),
            'projectUpdates': plan.get('projectUpdates', 0),
            'holidayUpdates': plan.get('holidayUpdates', 0),
        }
    except Exception as e:
        logger.error(f"선박 병합 오류: {e}")
        return {'success': False, 'message': '선박 병합 중 오류가 발생했습니다.'}


@eel.expose
def admin_update_local_db_path(new_path: str, admin_id: str) -> Dict[str, Any]:
    """로컬 DB 경로 변경 (관리자 전용)"""
    try:
        user = auth_manager.get_user(admin_id)
        if not user or user.get('role') != 'admin':
            return {'success': False, 'message': '관리자 권한이 필요합니다.'}
        logger.info(f"로컬 DB 경로 변경 요청: {new_path} by {admin_id}")
        return settings_manager.update_local_db_path(new_path)
    except Exception as e:
        logger.error(f"경로 변경 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def admin_update_cloud_path(new_path: str, admin_id: str) -> Dict[str, Any]:
    """클라우드 경로 변경 (관리자)"""
    try:
        user = auth_manager.get_user(admin_id)
        if not user or user.get('role') != 'admin':
            return {'success': False, 'message': '관리자 권한이 필요합니다.'}
        logger.info(f"클라우드 경로 변경 요청: {new_path} by {admin_id}")
        result = settings_manager.update_cloud_path(new_path)
        if result.get('success'):
            # settings_manager는 JSON 파일만 업데이트 → 메모리 config 직접 갱신
            config.set('database.cloud_path', new_path)
            config.set('database.cloud_sync_enabled', True)
            # 회사 PC 모드 자동 설정
            config.set('database.sync_mode', 'company')
            config.save()
            # 싱글톤 cloud_sync 인스턴스 즉시 재초기화 (재시작 불필요)
            cloud_sync.cloud_folder = cloud_sync._get_cloud_folder()
            cloud_sync.enabled = config.cloud_sync_enabled
            if cloud_sync.cloud_folder:
                logger.info(f"클라우드 폴더 재초기화 완료 [company]: {cloud_sync.cloud_folder}")
                result['message'] = f'클라우드 경로가 설정되었습니다 (회사 PC 모드): {cloud_sync.cloud_folder}'
            else:
                result['message'] = '경로가 저장되었으나 폴더를 찾을 수 없습니다. 경로를 확인하세요.'
        return result
    except Exception as e:
        logger.error(f"경로 변경 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def admin_update_backup_path(new_path: str, admin_id: str) -> Dict[str, Any]:
    """백업 경로 변경 (관리자)"""
    try:
        user = auth_manager.get_user(admin_id)
        if not user or user.get('role') != 'admin':
            return {'success': False, 'message': '관리자 권한이 필요합니다.'}
        logger.info(f"백업 경로 변경 요청: {new_path} by {admin_id}")
        return settings_manager.update_backup_path(new_path)
    except Exception as e:
        logger.error(f"경로 변경 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def admin_create_backup(admin_id: str) -> Dict[str, Any]:
    """수동 백업 생성 (관리자)"""
    try:
        user = auth_manager.get_user(admin_id)
        if not user or user.get('role') != 'admin':
            return {'success': False, 'message': '관리자 권한이 필요합니다.'}
        logger.info(f"수동 백업 생성 요청 by {admin_id}")
        return settings_manager.create_backup()
    except Exception as e:
        logger.error(f"백업 생성 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


# ============================================================================
# 사용자 관리 (기존 호환성 유지)
# ============================================================================

@eel.expose
def login_user(username: str) -> bool:
    """사용자 로그인 (레거시)"""
    try:
        success = db.add_or_update_user(username)
        if success:
            logger.info(f"사용자 로그인: {username}")
            db.add_activity_log(username, 'login', '', '로그인')
        return success
    except Exception as e:
        logger.error(f"로그인 오류: {e}")
        return False


@eel.expose
def get_recent_users() -> List[Dict[str, Any]]:
    """최근 사용자 목록"""
    try:
        users = db.get_all_users()
        return [u.to_dict() for u in users[:10]]
    except Exception as e:
        logger.error(f"사용자 목록 조회 오류: {e}")
        return []


# ============================================================================
# 작업 레코드 관리
# ============================================================================

@eel.expose
def load_work_records(date: str, work_type: str = 'day') -> List[Dict[str, Any]]:
    """작업 레코드 로드 (work_type: 'day'|'night')"""
    try:
        logger.info(f"작업 레코드 로드 요청: {date} [{work_type}]")
        if not date or len(date) != 10:  # #11/#15 — 날짜 형식 기본 검증
            logger.warning(f"잘못된 날짜 형식: {date!r}")
            return []
        try:
            datetime.strptime(date, '%Y-%m-%d')
        except ValueError:
            logger.warning(f"유효하지 않은 날짜: {date!r}")
            return []
        wt = work_type if work_type in ('day', 'night') else 'day'
        records = work_record_service.get_records_for_date(date, wt)
        return records
    except Exception as e:
        logger.error(f"작업 레코드 로드 오류: {e}")
        return []  # #15 — JS 호환 유지 (빈 배열=오류/데이터없음 모두 동일 처리)


@eel.expose
def save_work_records(date: str, records: List[Dict[str, Any]],
                      username: str, work_type: str = 'day') -> Dict[str, Any]:
    """작업 레코드 저장 (work_type: 'day'|'night')"""
    try:
        # ── 쓰기 권한 검증 (admin 역할 또는 can_write=1 인 사용자만 저장 가능)
        if not username:
            return {'success': False, 'message': '로그인 정보가 없습니다.'}
        if not auth_manager.get_can_write_by_fullname(username):
            logger.warning(f"쓰기 권한 없는 저장 시도: {username}")
            return {'success': False, 'message': '쓰기 권한이 없습니다. 관리자에게 문의하세요.'}

        logger.info(f"작업 레코드 저장 요청: {date} [{work_type}], 사용자: {username}")
        if not date or len(date) != 10:
            return {'success': False, 'message': '잘못된 날짜 형식'}
        try:
            datetime.strptime(date, '%Y-%m-%d')
        except ValueError:
            return {'success': False, 'message': '유효하지 않은 날짜'}
        wt = work_type if work_type in ('day', 'night') else 'day'
        save_result = work_record_service.save_records_for_date(date, records, username, wt)
        success = save_result.get('success', False)

        # 저장 성공 시 클라우드 동기화 (백그라운드 스레드 - UI 블로킹 방지)
        if success and cloud_sync.enabled:
            # external 모드: push + 알림 생성 / company 모드: push만
            _sync_fn = (cloud_sync.sync_to_cloud_notify
                        if cloud_sync.sync_mode == 'external'
                        else cloud_sync.sync_to_cloud)
            _start_tracked_thread(target=_sync_fn)

        return {'success': success, 'message': '저장되었습니다.' if success else save_result.get('message', '저장 실패')}
    except Exception as e:
        logger.error(f"작업 레코드 저장 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def get_date_save_info(date: str, work_type: str = 'day') -> Dict[str, Any]:
    """날짜별 마지막 저장 정보 조회 (JS 충돌 감지용)"""
    try:
        wt = work_type if work_type in ('day', 'night') else 'day'
        with db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(
                'SELECT updated_at, updated_by FROM work_records '
                'WHERE date = ? AND work_type = ? ORDER BY updated_at DESC LIMIT 1',
                (date, wt)
            )
            row = cursor.fetchone()
            if row:
                return {
                    'has_records': True,
                    'updated_at': row['updated_at'] or '',
                    'updated_by': row['updated_by'] or ''
                }
            return {'has_records': False, 'updated_at': '', 'updated_by': ''}
    except Exception as e:
        logger.error(f"날짜 저장 정보 조회 실패: {e}")
        return {'has_records': False, 'updated_at': '', 'updated_by': ''}


@eel.expose
def load_holiday_work_entries(period_key: str) -> List[Dict[str, Any]]:
    """휴일 작업 인원 명단 로드 (period_key = 해당 주 금요일 날짜 YYYY-MM-DD)"""
    try:
        rows = db.load_holiday_work_entries(period_key)
        # snake_case → camelCase 변환
        result = []
        for row in rows:
            vendor_company = row.get('vendor_company', '') or row.get('company', '')
            result.append({
                'id': row.get('id'),
                'periodKey': row.get('period_key', ''),
                'seq': row.get('seq', 0),
                'department': row.get('department', ''),
                'rank': row.get('rank', ''),
                'name': row.get('name', ''),
                'friWork': row.get('fri_work', '-'),
                'satWork': row.get('sat_work', '-'),
                'sunWork': row.get('sun_work', '-'),
                'workContent': row.get('work_content', ''),
                'contractNumber': row.get('contract_number', ''),
                'company': vendor_company,
                'ownerCompany': row.get('owner_company', ''),
                'vendorCompany': vendor_company,
                'shipName': row.get('ship_name', ''),
            })
        return result
    except Exception as e:
        logger.error(f"휴일 작업 명단 로드 오류: {e}")
        return []


@eel.expose
def save_holiday_work_entries(period_key: str, entries: List[Dict[str, Any]],
                              username: str) -> Dict[str, Any]:
    """휴일 작업 인원 명단 저장"""
    try:
        if not username:
            return {'success': False, 'message': '로그인 정보가 없습니다.'}
        if not auth_manager.get_can_write_by_fullname(username):
            return {'success': False, 'message': '쓰기 권한이 없습니다.'}
        if not period_key or len(period_key) != 10:
            return {'success': False, 'message': '잘못된 기간 키'}
        success = db.save_holiday_work_entries(period_key, entries, username)
        return {'success': success, 'message': '저장되었습니다.' if success else '저장 실패'}
    except Exception as e:
        logger.error(f"휴일 작업 명단 저장 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def get_latest_friday(date_str: str = '') -> str:
    """주어진 날짜(또는 오늘)로부터 가장 가까운 이전 금요일 반환 (YYYY-MM-DD)"""
    try:
        if date_str and len(date_str) == 10:
            base = datetime.strptime(date_str, '%Y-%m-%d')
        else:
            base = datetime.now()
        # isoweekday: 월=1 … 금=5 … 일=7
        days_since_fri = (base.isoweekday() - 5) % 7
        friday = base - timedelta(days=days_since_fri)
        return friday.strftime('%Y-%m-%d')
    except Exception as e:
        logger.error(f"금요일 계산 오류: {e}")
        return ''


@eel.expose
def get_holiday_period_dates(period_key: str) -> Dict[str, str]:
    """period_key(금요일) 기준 금/토/일 날짜+라벨 dict 반환"""
    try:
        fri = datetime.strptime(period_key, '%Y-%m-%d')
        sat = fri + timedelta(days=1)
        sun = fri + timedelta(days=2)
        weekday_labels = {0: '월', 1: '화', 2: '수', 3: '목', 4: '금', 5: '토', 6: '일'}
        return {
            'fri': fri.strftime('%Y-%m-%d'),
            'sat': sat.strftime('%Y-%m-%d'),
            'sun': sun.strftime('%Y-%m-%d'),
            'friLabel': f"금({fri.day})",
            'satLabel': f"토({sat.day})",
            'sunLabel': f"일({sun.day})",
        }
    except Exception as e:
        logger.error(f"휴일 기간 날짜 계산 오류: {e}")
        return {'fri': '', 'sat': '', 'sun': '',
                'friLabel': '금', 'satLabel': '토', 'sunLabel': '일'}


@eel.expose
def clear_all_records(admin_id: str = '') -> Dict[str, Any]:
    """전체 작업 레코드 삭제 (관리자 전용)"""
    try:
        # 관리자 권한 확인
        if not admin_id:
            logger.warning("clear_all_records: admin_id 없음 - 권한 거부")
            return {'success': False, 'message': '관리자 권한이 필요합니다.'}

        user = auth_manager.get_user(admin_id)
        if not user or user.get('role') != 'admin':
            logger.warning(f"clear_all_records: 권한 없음 - user_id={admin_id}")
            return {'success': False, 'message': '관리자만 실행할 수 있습니다.'}

        success = db.clear_all_work_records()
        if success:
            db.add_activity_log(admin_id, 'clear_all', '', '전체 작업 레코드 삭제')
            logger.info(f"전체 작업 레코드 삭제 완료 by {admin_id}")
            return {'success': True, 'message': '전체 작업 레코드가 삭제되었습니다.'}
        else:
            return {'success': False, 'message': '삭제 실패'}
    except Exception as e:
        logger.error(f"전체 레코드 삭제 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def load_yesterday_records(current_date: str) -> Dict[str, Any]:
    """마지막 평일 작업 불러오기 (주말·공휴일 자동 건너뜀)"""
    try:
        logger.info(f"마지막 평일 작업 불러오기 요청: {current_date}")
        records, loaded_date = work_record_service.get_yesterday_records(current_date)
        return {'records': records, 'date': loaded_date}
    except Exception as e:
        logger.error(f"어제 작업 불러오기 오류: {e}")
        return {'records': [], 'date': ''}


@eel.expose
def get_date_list(start_date: str = None, end_date: str = None) -> List[str]:
    """레코드가 있는 날짜 목록"""
    try:
        dates = work_record_service.get_date_list(start_date, end_date)
        return dates
    except Exception as e:
        logger.error(f"날짜 목록 조회 오류: {e}")
        return []


# ============================================================================
# 보고서
# ============================================================================

@eel.expose
def generate_report(date: str, username: str) -> Dict[str, Any]:
    """보고서 생성"""
    try:
        report = work_record_service.generate_report(date, username)
        return report
    except Exception as e:
        logger.error(f"보고서 생성 오류: {e}")
        return {}


@eel.expose
def get_project_start_date_by_contract(contract_number: str) -> str:
    """계약번호 기준 공사 시작일 조회"""
    try:
        if not contract_number or contract_number.strip() == '':
            return ''
        
        # 해당 계약번호의 최초 작업일 조회
        query = '''
            SELECT MIN(date) as start_date, COUNT(*) as count
            FROM work_records
            WHERE contract_number = ?
            AND contract_number IS NOT NULL
            AND contract_number != ''
        '''
        
        logger.info(f"계약번호 조회: {contract_number}")
        
        result = db.execute_query(query, (contract_number,))
        logger.info(f"쿼리 결과: {result}")
        
        if result and result[0][0]:
            from datetime import datetime
            start_date_str = result[0][0]  # YYYY-MM-DD 형식
            count = result[0][1] if len(result[0]) > 1 else 0
            logger.info(f"찾은 데이터: 시작일={start_date_str}, 건수={count}")
            try:  # #6 — strptime 예외 처리
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
                result_str = f"{start_date.month}/{start_date.day}"
                logger.info(f"반환값: {result_str}")
                return result_str
            except ValueError:
                logger.warning(f"날짜 형식 오류 (계약번호 기준): {start_date_str!r}")
                return ''
        
        logger.warning(f"계약번호 {contract_number}에 대한 데이터를 찾을 수 없음")
        return ''
    except Exception as e:
        logger.error(f"계약번호 기준 시작일 조회 오류: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return ''


@eel.expose
def get_project_start_date(ship_name: str) -> str:
    """선박별 공사 시작일 조회 (최초 작업일)"""
    try:
        if not ship_name or ship_name.strip() == '':
            return ''
        
        # 해당 선박의 모든 작업 기록 조회
        query = '''
            SELECT MIN(date) as start_date
            FROM work_records
            WHERE ship_name = ?
            AND ship_name != ''
        '''
        
        result = db.execute_query(query, (ship_name,))
        if result and result[0][0]:
            from datetime import datetime
            start_date_str = result[0][0]  # YYYY-MM-DD 형식
            try:  # #6 — strptime 예외 처리
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
                return f"{start_date.month}/{start_date.day}"
            except ValueError:
                logger.warning(f"날짜 형식 오류 (선박 기준): {start_date_str!r}")
                return ''
        return ''
    except Exception as e:
        logger.error(f"공사 시작일 조회 오류: {e}")
        return ''


@eel.expose
def get_project_start_dates_batch(contract_numbers: list, ship_names: list) -> dict:
    """여러 계약번호/선박명의 공사 시작일을 한 번에 조회 (일일보고 N+1 방지)"""
    result = {}
    try:
        contract_numbers = contract_numbers[:500] if contract_numbers else []
        ship_names = ship_names[:500] if ship_names else []
        if contract_numbers:
            placeholders = ','.join('?' * len(contract_numbers))
            rows = db.execute_query(
                f"SELECT contract_number, MIN(date) FROM work_records "
                f"WHERE contract_number IN ({placeholders}) "
                "AND contract_number != '' GROUP BY contract_number",
                contract_numbers
            )
            for cn, min_date in (rows or []):
                if min_date:
                    try:  # #6 — strptime 예외 처리
                        d = datetime.strptime(min_date, '%Y-%m-%d')
                        result[cn] = f"{d.month}/{d.day}"
                    except ValueError:
                        logger.warning(f"날짜 형식 오류 (배치-계약): {min_date!r}")

        if ship_names:
            placeholders = ','.join('?' * len(ship_names))
            rows = db.execute_query(
                f"SELECT ship_name, MIN(date) FROM work_records "
                f"WHERE ship_name IN ({placeholders}) "
                "AND (contract_number = '' OR contract_number IS NULL) "
                "GROUP BY ship_name",
                ship_names
            )
            for sn, min_date in (rows or []):
                if min_date and sn not in result:
                    try:  # #6 — strptime 예외 처리
                        d = datetime.strptime(min_date, '%Y-%m-%d')
                        result[sn] = f"{d.month}/{d.day}"
                    except ValueError:
                        logger.warning(f"날짜 형식 오류 (배치-선박): {min_date!r}")
    except Exception as e:
        logger.error(f"배치 시작일 조회 오류: {e}")
    return result


def _parse_search_ot(end_time: str) -> float:
    """조회 탭 OT 계산용 헬퍼 (17:00 이후 시간, 석식 공제 없음)"""
    try:
        parts = str(end_time or '').strip().split(':')
        h = int(parts[0])
        m = int(parts[1]) if len(parts) > 1 else 0
        total_min = h * 60 + m
        ot_min = total_min - 17 * 60
        if ot_min <= 0:
            return 0.0
        return round(ot_min / 60.0, 2)
    except Exception:
        return 0.0


def _build_search_record(row) -> Dict[str, Any]:
    work_type = row[10] or 'day'
    end_time = row[11] or ''
    ot = _parse_search_ot(end_time) if work_type == 'night' else 0.0
    return {
        'date': row[0] or '',
        'recordNumber': int(row[1]) if row[1] else 0,
        'contractNumber': row[2] or '',
        'contract_number': row[2] or '',
        'company': row[3] or '',
        'shipName': row[4] or '',
        'ship_name': row[4] or '',
        'engineModel': row[5] or '',
        'engine_model': row[5] or '',
        'workContent': row[6] or '',
        'work_content': row[6] or '',
        'leader': row[7] or '',
        'manpower': float(row[8]) if row[8] else 0.0,
        'teammates': row[9] or '',
        'workType': work_type,
        'endTime': end_time,
        'ot': ot,
        'otSource': 'night' if ot > 0 else '',
        'isSynthetic': False,
    }


def _query_search_work_records(search_type: str, query_text: str) -> List[Dict[str, Any]]:
    base_sql = '''
        SELECT date, record_number, contract_number, company, ship_name, engine_model,
               work_content, leader, manpower, teammates,
               COALESCE(work_type, 'day') AS work_type, COALESCE(end_time, '') AS end_time
        FROM work_records
    '''
    if search_type == 'contract':
        sql = base_sql + '''
            WHERE contract_number = ?
              AND contract_number IS NOT NULL
              AND contract_number != ''
            ORDER BY date, record_number, work_type
        '''
        rows = db.execute_query(sql, (query_text.strip().upper(),))
    elif search_type == 'ship':
        sql = base_sql + '''
            WHERE ship_name LIKE ?
              AND ship_name IS NOT NULL
              AND ship_name != ''
            ORDER BY date, record_number, work_type
        '''
        rows = db.execute_query(sql, (f'%{query_text.strip().upper()}%',))
    elif search_type == 'company':
        sql = base_sql + '''
            WHERE (teammates LIKE ? OR teammates LIKE ?)
            ORDER BY date, record_number, work_type
        '''
        name = query_text.strip()
        rows = db.execute_query(sql, (f'%{name}(%', f'%{name}[%'))
    else:
        rows = []
    return [_build_search_record(row) for row in (rows or [])]


def _holiday_value_to_ot(work_value: Any) -> float:
    text = str(work_value or '').strip()
    return 8.0 if text and text != '-' else 0.0


EMPTY_SHIP_LABEL = '(선박 미입력)'
LAST_MERGE_UNDO_KEY = 'admin.last_merge_undo'


def _normalize_holiday_worker_name(name: Any) -> str:
    text = str(name or '')
    text = re.sub(r'<[^>]+>', '', text)
    text = text.replace('*', '').strip()
    return text


def _normalize_company_label(name: Any) -> str:
    return _normalize_holiday_worker_name(name).lower()


def _normalize_ship_label(name: Any) -> str:
    text = _normalize_holiday_worker_name(name)
    if not text or text == EMPTY_SHIP_LABEL:
        return '__empty_ship__'
    return text.lower()


def _display_ship_label(name: Any) -> str:
    text = _normalize_holiday_worker_name(name)
    return text or EMPTY_SHIP_LABEL


def _storage_ship_label(name: Any) -> str:
    text = _normalize_holiday_worker_name(name)
    return '' if not text or text == EMPTY_SHIP_LABEL else text


def _normalize_merge_suggestion_key(name: Any) -> str:
    text = _normalize_holiday_worker_name(name)
    if not text or text == EMPTY_SHIP_LABEL:
        return ''
    return ''.join(ch.casefold() for ch in text if ch.isalnum())


def _levenshtein_distance_limit_one(left: str, right: str) -> int:
    if left == right:
        return 0
    if abs(len(left) - len(right)) > 1:
        return 2

    if len(left) > len(right):
        left, right = right, left

    if len(left) == len(right):
        mismatches = sum(1 for idx in range(len(left)) if left[idx] != right[idx])
        return mismatches if mismatches <= 1 else 2

    idx = 0
    jdx = 0
    mismatches = 0
    while idx < len(left) and jdx < len(right):
        if left[idx] == right[jdx]:
            idx += 1
            jdx += 1
            continue
        mismatches += 1
        if mismatches > 1:
            return 2
        jdx += 1

    if jdx < len(right) or idx < len(left):
        mismatches += 1
    return mismatches if mismatches <= 1 else 2


def _is_likely_merge_suggestion_pair(left_key: str, right_key: str) -> bool:
    if not left_key or not right_key or left_key == right_key:
        return False
    if _levenshtein_distance_limit_one(left_key, right_key) <= 1:
        return True
    if left_key[0] != right_key[0]:
        return False
    if abs(len(left_key) - len(right_key)) > 3:
        return False
    ratio = difflib.SequenceMatcher(None, left_key, right_key).ratio()
    return ratio >= 0.84


def _build_merge_suggestions(names: List[str], limit: int = 6) -> List[Dict[str, Any]]:
    unique_names: List[str] = []
    seen_names = set()
    for raw_name in names or []:
        name = str(raw_name or '').strip()
        if not name or name in seen_names or name == EMPTY_SHIP_LABEL:
            continue
        seen_names.add(name)
        unique_names.append(name)

    suggestions: List[Dict[str, Any]] = []
    seen_groups = set()
    key_groups: Dict[str, List[str]] = {}
    for name in unique_names:
        key = _normalize_merge_suggestion_key(name)
        if len(key) < 3:
            continue
        key_groups.setdefault(key, []).append(name)

    for group in key_groups.values():
        if len(group) < 2:
            continue
        names_key = tuple(sorted(group))
        if names_key in seen_groups:
            continue
        seen_groups.add(names_key)
        suggestions.append({
            'names': list(group),
            'label': ' / '.join(group),
            'reason': '형식 차이',
        })
        if len(suggestions) >= limit:
            return suggestions

    for idx, left_name in enumerate(unique_names):
        left_key = _normalize_merge_suggestion_key(left_name)
        if len(left_key) < 3:
            continue
        for right_name in unique_names[idx + 1:]:
            right_key = _normalize_merge_suggestion_key(right_name)
            if len(right_key) < 3:
                continue
            if left_key == right_key:
                continue
            if not _is_likely_merge_suggestion_pair(left_key, right_key):
                continue
            names_key = tuple(sorted([left_name, right_name]))
            if names_key in seen_groups:
                continue
            seen_groups.add(names_key)
            suggestions.append({
                'names': [left_name, right_name],
                'label': f'{left_name} / {right_name}',
                'reason': '유사 이름',
            })
            if len(suggestions) >= limit:
                return suggestions

    return suggestions


def _build_undo_snapshot(action: str, target: str, details: str,
                         updates: List[Dict[str, Any]]) -> Dict[str, Any]:
    return {
        'action': action,
        'target': target,
        'details': details,
        'createdAt': datetime.now().isoformat(),
        'updates': updates,
    }


def _save_last_merge_undo(snapshot: Dict[str, Any]) -> None:
    db.set_setting(LAST_MERGE_UNDO_KEY, json.dumps(snapshot, ensure_ascii=False))


def _load_last_merge_undo() -> Optional[Dict[str, Any]]:
    raw_value = db.get_setting(LAST_MERGE_UNDO_KEY, '')
    if not raw_value:
        return None
    try:
        return json.loads(raw_value)
    except Exception:
        return None


def _clear_last_merge_undo() -> None:
    db.set_setting(LAST_MERGE_UNDO_KEY, '')


def _apply_merge_updates(cursor, updates: List[Dict[str, Any]], use_old_values: bool = False) -> int:
    table_fields = {
        'work_records': {'company', 'ship_name', 'teammates'},
        'board_projects': {'company', 'ship_name'},
        'holiday_work_entries': {'owner_company', 'vendor_company', 'company', 'ship_name', 'name'},
    }
    now = datetime.now().isoformat()
    applied = 0

    for update in updates or []:
        table = str(update.get('table') or '')
        row_id = update.get('id')
        values = update.get('old_fields' if use_old_values else 'new_fields') or {}
        if table not in table_fields or not values or row_id in (None, ''):
            continue
        field_names = [field for field in values.keys() if field in table_fields[table]]
        if not field_names:
            continue
        assignments = ', '.join([f'{field} = ?' for field in field_names] + ['updated_at = ?'])
        params = [values[field] for field in field_names] + [now, row_id]
        cursor.execute(f'UPDATE {table} SET {assignments} WHERE id = ?', params)
        applied += 1

    return applied


def _mixed_locale_sort_key(name: Any):
    text = str(name or '').strip()
    if not text:
        return (2, '', '')
    first = text[0]
    code = ord(first)
    if 0xAC00 <= code <= 0xD7A3:
        group = 0  # 한글 우선
    elif ('A' <= first <= 'Z') or ('a' <= first <= 'z'):
        group = 1  # 영문 다음
    else:
        group = 2
    return (group, text.lower(), text)


def _iter_vendor_segments(teammates: str):
    text = str(teammates or '')
    patterns = [
        ('contract', re.compile(r'([^,\[\]()\n]+?)\(([^)]+)\)'), '(', ')'),
        ('daily', re.compile(r'([^,\[\]()\n]+?)\[([^\]]+)\]'), '[', ']'),
    ]
    for _kind, pattern, open_ch, close_ch in patterns:
        for match in pattern.finditer(text):
            yield {
                'company': _normalize_holiday_worker_name(match.group(1)),
                'raw_company': match.group(1),
                'names_blob': match.group(2),
                'open_ch': open_ch,
                'close_ch': close_ch,
                'full_match': match.group(0),
            }


def _preserve_worker_style(original_name: str, target_name: str) -> str:
    raw = str(original_name or '').strip()
    if raw.lower().startswith('<i>') and raw.lower().endswith('</i>'):
        return f"<i>{target_name}</i>"
    if raw.startswith('*') and raw.endswith('*') and len(raw) >= 2:
        return f"*{target_name}*"
    return target_name


def _replace_vendor_worker_names_in_teammates(teammates: str, vendor_company: str,
                                              source_names: List[str], target_name: str):
    target_vendor_key = _normalize_company_label(vendor_company)
    source_keys = {_normalize_holiday_worker_name(name) for name in (source_names or []) if str(name or '').strip()}
    if not target_vendor_key or not source_keys:
        return teammates, False

    changed = False

    def _rewrite_names_blob(names_blob: str) -> str:
        nonlocal changed
        tokens = [token.strip() for token in str(names_blob or '').split(',')]
        new_tokens = []
        seen = set()
        local_changed = False
        for token in tokens:
            if not token:
                continue
            normalized = _normalize_holiday_worker_name(token)
            if normalized in source_keys:
                token = _preserve_worker_style(token, target_name)
                normalized = _normalize_holiday_worker_name(token)
                local_changed = True
            if not normalized or normalized in seen:
                continue
            seen.add(normalized)
            new_tokens.append(token)
        if local_changed:
            changed = True
        return ', '.join(new_tokens)

    def _replace_with_pattern(text: str, pattern, open_ch: str, close_ch: str) -> str:
        def _repl(match):
            company_raw = match.group(1)
            if _normalize_company_label(company_raw) != target_vendor_key:
                return match.group(0)
            rewritten_blob = _rewrite_names_blob(match.group(2))
            return f"{company_raw.strip()}{open_ch}{rewritten_blob}{close_ch}"
        return pattern.sub(_repl, text)

    updated = str(teammates or '')
    updated = _replace_with_pattern(updated, re.compile(r'([^,\[\]()\n]+?)\(([^)]+)\)'), '(', ')')
    updated = _replace_with_pattern(updated, re.compile(r'([^,\[\]()\n]+?)\[([^\]]+)\]'), '[', ']')
    return updated, changed and updated != str(teammates or '')


def _extract_vendor_workers_from_teammates(teammates: str) -> Dict[str, List[str]]:
    vendor_map: Dict[str, List[str]] = {}
    for segment in _iter_vendor_segments(teammates):
        company = segment['company']
        if not company:
            continue
        if company not in vendor_map:
            vendor_map[company] = []
        for raw_name in str(segment['names_blob'] or '').split(','):
            clean_name = _normalize_holiday_worker_name(raw_name)
            if clean_name:
                vendor_map[company].append(clean_name)
    return vendor_map


def _normalize_holiday_meta(meta: Dict[str, str]) -> Dict[str, str]:
    return {
        'contract_number': str(meta.get('contract_number', '') or '').strip(),
        'owner_company': str(meta.get('owner_company', '') or '').strip(),
        'vendor_company': str(meta.get('vendor_company', '') or '').strip(),
        'ship_name': str(meta.get('ship_name', '') or '').strip(),
        'work_content': str(meta.get('work_content', '') or '').strip(),
        'worker_type': str(meta.get('worker_type', '') or '').strip(),
    }


def _add_holiday_meta_candidate(meta_map: Dict[str, Dict[str, Dict[str, str]]], raw_name: Any,
                                meta: Dict[str, str]) -> None:
    clean_name = _normalize_holiday_worker_name(raw_name)
    if not clean_name:
        return
    normalized_meta = _normalize_holiday_meta(meta)
    signature = '|'.join([
        normalized_meta['contract_number'],
        normalized_meta['owner_company'],
        normalized_meta['vendor_company'],
        normalized_meta['ship_name'],
        normalized_meta['work_content'],
        normalized_meta['worker_type'],
    ])
    if clean_name not in meta_map:
        meta_map[clean_name] = {}
    meta_map[clean_name][signature] = normalized_meta


def _resolve_holiday_meta_candidate(meta_map: Dict[str, Dict[str, Dict[str, str]]],
                                    raw_name: Any) -> Optional[Dict[str, str]]:
    clean_name = _normalize_holiday_worker_name(raw_name)
    if not clean_name:
        return None
    candidates = meta_map.get(clean_name, {})
    if len(candidates) != 1:
        return None
    return next(iter(candidates.values()))


def _build_holiday_meta_map_for_period(period_key: str) -> Dict[str, Dict[str, Dict[str, str]]]:
    try:
        fri_dt = datetime.strptime(period_key, '%Y-%m-%d')
    except Exception:
        return {}

    date_keys = [(fri_dt + timedelta(days=offset)).strftime('%Y-%m-%d') for offset in range(3)]
    placeholders = ','.join('?' for _ in date_keys)
    sql = f'''
        SELECT contract_number, company, ship_name, work_content, leader, teammates
        FROM work_records
        WHERE date IN ({placeholders})
          AND COALESCE(work_type, 'day') = 'day'
    '''

    with db.get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(sql, tuple(date_keys))
        rows = cursor.fetchall()

    meta_map: Dict[str, Dict[str, Dict[str, str]]] = {}
    contract_pattern = re.compile(r'([^,\[\]()\n]+?)\(([^)]+)\)')
    daily_pattern = re.compile(r'([^,\[\]()\n]+?)\[([^\]]+)\]')

    for row in rows:
        contract_number = row['contract_number'] or ''
        ship_name = row['ship_name'] or ''
        owner_company = row['company'] or ''
        work_content = row['work_content'] or ''
        leader = row['leader'] or ''
        teammates = row['teammates'] or ''
        in_house_meta = {
            'contract_number': contract_number,
            'owner_company': owner_company,
            'vendor_company': '',
            'ship_name': ship_name,
            'work_content': work_content,
            'worker_type': 'inhouse',
        }

        _add_holiday_meta_candidate(meta_map, leader, in_house_meta)

        remaining = teammates
        for match in contract_pattern.finditer(teammates):
            vendor_company = _normalize_holiday_worker_name(match.group(1))
            for worker_name in match.group(2).split(','):
                _add_holiday_meta_candidate(meta_map, worker_name, {
                    'contract_number': contract_number,
                    'owner_company': owner_company,
                    'vendor_company': vendor_company,
                    'ship_name': ship_name,
                    'work_content': work_content,
                    'worker_type': 'vendor',
                })
            remaining = remaining.replace(match.group(0), '')

        for match in daily_pattern.finditer(remaining):
            vendor_company = _normalize_holiday_worker_name(match.group(1))
            for worker_name in match.group(2).split(','):
                _add_holiday_meta_candidate(meta_map, worker_name, {
                    'contract_number': contract_number,
                    'owner_company': owner_company,
                    'vendor_company': vendor_company,
                    'ship_name': ship_name,
                    'work_content': work_content,
                    'worker_type': 'vendor',
                })
            remaining = remaining.replace(match.group(0), '')

        for worker_name in remaining.split(','):
            _add_holiday_meta_candidate(meta_map, worker_name, in_house_meta)

    return meta_map


def _enrich_holiday_entry_metadata(row: Dict[str, Any],
                                   period_meta_cache: Dict[str, Dict[str, Dict[str, Dict[str, str]]]]) -> Dict[str, Any]:
    enriched = dict(row)
    enriched['owner_company'] = str(enriched.get('owner_company', '') or '').strip()
    enriched['vendor_company'] = str(
        enriched.get('vendor_company', '') or enriched.get('company', '') or ''
    ).strip()
    period_key = enriched.get('period_key', '')
    if not period_key:
        return enriched

    meta_map = period_meta_cache.get(period_key)
    if meta_map is None:
        meta_map = _build_holiday_meta_map_for_period(period_key)
        period_meta_cache[period_key] = meta_map

    meta = _resolve_holiday_meta_candidate(meta_map, enriched.get('name', ''))
    if meta:
        existing_vendor_company = str(enriched.get('vendor_company') or '').strip()
        owner_company = str(meta.get('owner_company', '') or '').strip()
        worker_type = str(meta.get('worker_type', '') or '').strip()
        if not enriched.get('owner_company'):
            enriched['owner_company'] = owner_company
        if worker_type == 'inhouse':
            if not existing_vendor_company or (owner_company and existing_vendor_company == owner_company):
                enriched['vendor_company'] = ''
        elif (not existing_vendor_company) or (owner_company and existing_vendor_company == owner_company):
            enriched['vendor_company'] = meta.get('vendor_company', '')

        enriched['contract_number'] = enriched.get('contract_number') or meta.get('contract_number', '')
        enriched['ship_name'] = enriched.get('ship_name') or meta.get('ship_name', '')
        enriched['work_content'] = enriched.get('work_content') or meta.get('work_content', '')

    enriched['company'] = enriched.get('vendor_company', '')
    if enriched.get('contract_number') and enriched.get('ship_name'):
        return enriched

    if not meta:
        return enriched
    return enriched


def _build_holiday_ot_records(rows) -> List[Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    for row in rows or []:
        period_key = row['period_key']
        try:
            fri_dt = datetime.strptime(period_key, '%Y-%m-%d')
        except Exception:
            continue
        day_defs = [
            ('fri_work', fri_dt, '금'),
            ('sat_work', fri_dt + timedelta(days=1), '토'),
            ('sun_work', fri_dt + timedelta(days=2), '일'),
        ]
        for key, dt_obj, label in day_defs:
            ot = _holiday_value_to_ot(row[key])
            if ot <= 0:
                continue
            work_content = row['work_content'] or ''
            records.append({
                'date': dt_obj.strftime('%Y-%m-%d'),
                'recordNumber': int(row['seq'] or 0),
                'contractNumber': row['contract_number'] or '',
                'contract_number': row['contract_number'] or '',
                'company': row['owner_company'] or '',
                'ownerCompany': row['owner_company'] or '',
                'vendorCompany': row['vendor_company'] or row['company'] or '',
                'shipName': row['ship_name'] or '',
                'ship_name': row['ship_name'] or '',
                'engineModel': '',
                'engine_model': '',
                'workContent': work_content,
                'work_content': work_content,
                'leader': row['name'] or '',
                'manpower': 0.0,
                'teammates': row['vendor_company'] or row['company'] or '',
                'workType': 'holiday',
                'endTime': '',
                'ot': ot,
                'otSource': 'holiday',
                'isSynthetic': True,
                'holidayLabel': label,
            })
    return records


def _query_holiday_ot_records(search_type: str, query_text: str) -> List[Dict[str, Any]]:
    if search_type not in ('contract', 'ship', 'company'):
        return []
    sql = '''
        SELECT period_key, seq, name, fri_work, sat_work, sun_work, work_content,
               contract_number, company, owner_company, vendor_company, ship_name
        FROM holiday_work_entries
        ORDER BY period_key, seq
    '''
    with db.get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(sql)
        rows = cursor.fetchall()
    period_meta_cache: Dict[str, Dict[str, Dict[str, Dict[str, str]]]] = {}
    normalized_query = str(query_text or '').strip()
    normalized_query_upper = normalized_query.upper()
    normalized_query_lower = normalized_query.lower()
    filtered_rows = []
    for raw_row in rows:
        row = _enrich_holiday_entry_metadata(dict(raw_row), period_meta_cache)
        contract_number = str(row.get('contract_number') or '').strip().upper()
        ship_name = str(row.get('ship_name') or '').strip().upper()
        vendor_company = str(row.get('vendor_company') or row.get('company') or '').strip().lower()
        if search_type == 'contract' and contract_number != normalized_query_upper:
            continue
        if search_type == 'ship' and normalized_query_upper not in ship_name:
            continue
        if search_type == 'company' and normalized_query_lower not in vendor_company:
            continue
        filtered_rows.append(row)
    return _build_holiday_ot_records(filtered_rows)


@eel.expose
def search_records_with_ot(search_type: str, query: str) -> Dict[str, Any]:
    """조회 탭용 통합 검색 + OT 집계"""
    try:
        st = (search_type or '').strip().lower()
        if st not in ('contract', 'ship', 'company'):
            return {'success': False, 'message': '지원하지 않는 조회 유형입니다.', 'records': [], 'summary': {}}
        if not query or not str(query).strip():
            return {'success': False, 'message': '조회어를 입력해주세요.', 'records': [], 'summary': {}}

        work_records = _query_search_work_records(st, str(query))
        holiday_records = _query_holiday_ot_records(st, str(query))
        records = sorted(
            work_records + holiday_records,
            key=lambda r: (r.get('date', ''), int(r.get('recordNumber', 0)), r.get('workType', '')),
        )

        night_ot = round(sum(float(r.get('ot', 0) or 0) for r in records if r.get('otSource') == 'night'), 1)
        holiday_ot = round(sum(float(r.get('ot', 0) or 0) for r in records if r.get('otSource') == 'holiday'), 1)
        total_ot = round(night_ot + holiday_ot, 1)
        total_manpower = round(sum(float(r.get('manpower', 0) or 0) for r in records), 1)

        return {
            'success': True,
            'records': records,
            'summary': {
                'totalRecords': len(records),
                'totalManpower': total_manpower,
                'totalOt': total_ot,
                'nightOt': night_ot,
                'holidayOt': holiday_ot,
            }
        }
    except Exception as e:
        logger.error(f"통합 검색 OT 집계 오류: {e}")
        return {'success': False, 'message': '조회 중 오류가 발생했습니다.', 'records': [], 'summary': {}}


@eel.expose
def search_records_by_contract(contract_number: str) -> List[Dict[str, Any]]:
    """계약번호로 작업 내역 조회"""
    try:
        result = search_records_with_ot('contract', contract_number)
        return result.get('records', []) if result.get('success') else []
    except Exception as e:
        logger.error(f"계약번호 검색 오류: {e}")
        return []


@eel.expose
def get_latest_record_by_contract(contract_number: str) -> Dict[str, Any]:
    """계약번호의 가장 최근 작업 내역 반환 (일일 작업 입력 자동완성용).
    반환 필드: found, company, shipName, engineModel, workContent, location"""
    try:
        if not contract_number or not contract_number.strip():
            return {'found': False}

        cn = contract_number.strip().upper()
        # 가장 최근 날짜·행번호 기준 1건
        query = '''
            SELECT company, ship_name, engine_model, work_content, location
            FROM work_records
            WHERE contract_number = ?
              AND contract_number != ''
            ORDER BY date DESC, record_number DESC
            LIMIT 1
        '''
        results = db.execute_query(query, (cn,))
        if results and results[0] and any(results[0]):
            row = results[0]
            return {
                'found':       True,
                'company':     row[0] or '',
                'shipName':    row[1] or '',
                'engineModel': row[2] or '',
                'workContent': row[3] or '',
                'location':    row[4] or '',
            }
        return {'found': False}
    except Exception as e:
        logger.error(f"계약번호 자동완성 조회 오류: {e}")
        return {'found': False}


@eel.expose
def get_latest_contract_number() -> str:
    """DB에서 가장 최근 날짜의 계약번호 반환"""
    try:
        results = db.execute_query(
            '''SELECT contract_number FROM work_records
               WHERE contract_number != '' AND contract_number IS NOT NULL
               ORDER BY date DESC, id DESC LIMIT 1'''
        )
        if results and results[0][0]:
            return results[0][0].strip().upper()
        return ''
    except Exception as e:
        logger.error(f"최신 계약번호 조회 오류: {e}")
        return ''


@eel.expose
def validate_contract_number(contract_number: str) -> Dict[str, Any]:
    """계약번호 형식 유효성 검사 (SH-YYYY-NNN-X 형식)"""
    import re
    if not contract_number or not contract_number.strip():
        return {'valid': True, 'message': ''}  # 빈 값은 허용

    cn = contract_number.strip().upper()
    # SH-YYYY-NNN-X (X는 영숫자 1자 이상)
    pattern = r'^SH-\d{4}-\d{3}-[A-Z0-9]+$'
    if re.match(pattern, cn):
        return {'valid': True, 'message': ''}
    return {
        'valid': False,
        'message': f'계약번호 형식이 올바르지 않습니다. (예: SH-2024-001-T)'
    }


@eel.expose
def load_vacation_data(date: str) -> Dict[str, str]:
    """날짜별 휴가자 현황 로드"""
    try:
        return db.load_vacation_records(date)
    except Exception as e:
        logger.error(f"휴가자 현황 로드 오류: {e}")
        return {'연차': '', '반차': '', '반반차': '', '공가': ''}


@eel.expose
def save_vacation_data(date: str, data: Dict, username: str) -> Dict[str, Any]:
    """날짜별 휴가자 현황 저장"""
    try:
        success = db.save_vacation_records(date, data, username)
        return {'success': success}
    except Exception as e:
        logger.error(f"휴가자 현황 저장 오류: {e}")
        return {'success': False}


# ============================================================================
# 직원 연차 관리 API
# ============================================================================

@eel.expose
def get_employee_leave_info(employee_name: str) -> Dict[str, Any]:
    """직원 연차 전체 정보 조회"""
    try:
        if not employee_name or not employee_name.strip():
            return {}
        return db.get_employee_leave_info(employee_name.strip())
    except Exception as e:
        logger.error(f"직원 연차 정보 조회 오류: {e}")
        return {}


@eel.expose
def save_employee_annual_config(employee_name: str, generation_month: int, note: str,
                                generation_day: int = 1) -> Dict[str, Any]:
    """직원 연차 설정 저장"""
    try:
        success = db.save_employee_annual_config(
            employee_name.strip(),
            int(generation_month),
            note or '',
            int(generation_day or 1)
        )
        return {'success': success}
    except Exception as e:
        logger.error(f"직원 연차 설정 저장 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def add_leave_grant(employee_name: str, grant_year: int, grant_month: int,
                    days: float, note: str) -> Dict[str, Any]:
    """연차 부여 이력 추가"""
    try:
        new_id = db.add_leave_grant(employee_name.strip(), int(grant_year), int(grant_month),
                                    float(days), note or '')
        if new_id >= 0:
            return {'success': True, 'id': new_id}
        return {'success': False, 'message': '추가 실패'}
    except Exception as e:
        logger.error(f"연차 부여 이력 추가 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def delete_leave_grant(grant_id: int) -> Dict[str, Any]:
    """연차 부여 이력 삭제"""
    try:
        success = db.delete_leave_grant(int(grant_id))
        return {'success': success}
    except Exception as e:
        logger.error(f"연차 부여 이력 삭제 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def add_leave_usage(employee_name: str, use_date: str, leave_type: str, note: str) -> Dict[str, Any]:
    """연차 사용 내역 추가"""
    try:
        if leave_type not in ('연차', '반차', '반반차', '공가'):
            return {'success': False, 'message': '유효하지 않은 휴가 종류입니다.'}
        new_id = db.add_leave_usage(employee_name.strip(), use_date, leave_type, note or '')
        if new_id >= 0:
            return {'success': True, 'id': new_id}
        return {'success': False, 'message': '추가 실패'}
    except Exception as e:
        logger.error(f"연차 사용 내역 추가 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def delete_leave_usage(usage_id: int) -> Dict[str, Any]:
    """연차 사용 내역 삭제"""
    try:
        success = db.delete_leave_usage(int(usage_id))
        return {'success': success}
    except Exception as e:
        logger.error(f"연차 사용 내역 삭제 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def get_employee_names_for_leave() -> List[str]:
    """연차 관리용 직원 이름 목록"""
    try:
        return db.get_employee_names_for_leave()
    except Exception as e:
        logger.error(f"직원 이름 목록 조회 오류: {e}")
        return []


@eel.expose
def get_employee_directory() -> Dict[str, Any]:
    """직원 명부 조회"""
    try:
        return {
            'success': True,
            'employees': db.get_employee_directory(),
            'externalHeaders': db.get_employee_directory_external_headers()
        }
    except Exception as e:
        logger.error(f"직원 명부 조회 오류: {e}")
        return {'success': False, 'message': '직원 명부를 불러오지 못했습니다.', 'employees': []}


@eel.expose
def save_employee_directory(rows_json: str) -> Dict[str, Any]:
    """직원 명부 저장"""
    try:
        payload = json.loads(rows_json or '[]')
        external_headers = None
        if isinstance(payload, dict):
            rows = payload.get('rows', [])
            external_headers = payload.get('externalHeaders', [])
        else:
            rows = payload
        if not isinstance(rows, list):
            return {'success': False, 'message': '직원 명부 형식이 올바르지 않습니다.'}
        success, message = db.save_employee_directory(rows, external_headers)
        return {'success': success, 'message': message}
    except Exception as e:
        logger.error(f"직원 명부 저장 오류: {e}")
        return {'success': False, 'message': '직원 명부 저장 중 오류가 발생했습니다.'}


@eel.expose
def save_work_hours_ot_override(name: str, date: str, start_time: str, end_time: str,
                                note: str = '') -> Dict[str, Any]:
    """근로시간관리 달력에서 수정한 연장근로 시작/종료 시간을 저장"""
    try:
        if not name or not name.strip():
            return {'success': False, 'message': '직원명을 입력하세요.'}
        if not date:
            return {'success': False, 'message': '날짜가 없습니다.'}
        success = db.save_work_hours_ot_override(
            name.strip(),
            date,
            (start_time or '').strip(),
            (end_time or '').strip(),
            note or ''
        )
        return {'success': success, 'message': '저장되었습니다.' if success else '저장 실패'}
    except Exception as e:
        logger.error(f"근로시간 OT 오버라이드 저장 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def get_work_hours_by_month(name: str, year: int, month: int,
                             meal_deduct: bool = True) -> Dict[str, Any]:
    """직원 월별 근로 시간 조회 (달력용)

    Args:
        name: 직원 이름
        year: 연도
        month: 월
        meal_deduct: 석식 공제 여부 (True=17시+근무 시 1시간 공제)

    Returns:
        {success, name, year, month, days: {YYYY-MM-DD: {regular, ot, leave_type, is_weekend}}}
    """
    try:
        import calendar as _cal

        if not name or not name.strip():
            return {'success': False, 'message': '직원명을 입력하세요.'}
        name = name.strip()
        year = int(year)
        month = int(month)
        if not (1 <= month <= 12) or year < 2000:
            return {'success': False, 'message': '유효하지 않은 연도/월입니다.'}

        import datetime as _dt
        today = _dt.date.today()
        days_in_month = _cal.monthrange(year, month)[1]
        first_day = _dt.date(year, month, 1)
        last_day = _dt.date(year, month, days_in_month)
        first_day_sunday_index = (first_day.weekday() + 1) % 7
        last_day_sunday_index = (last_day.weekday() + 1) % 7
        calendar_start = first_day - _dt.timedelta(days=first_day_sunday_index)
        calendar_end = last_day + _dt.timedelta(days=(6 - last_day_sunday_index))
        range_start = calendar_start.strftime('%Y-%m-%d')
        range_end = calendar_end.strftime('%Y-%m-%d')

        # ── 1. 달력에 표시할 주 범위 전체 휴가 사용 내역 ──
        leave_rows = db.execute_query(
            "SELECT use_date, leave_type, days FROM employee_leave_usage "
            "WHERE employee_name = ? AND use_date BETWEEN ? AND ? ORDER BY use_date",
            (name, range_start, range_end)
        ) or []
        # date → {leave_type, days}
        leave_map: Dict[str, Any] = {}
        for row in leave_rows:
            d, lt, dy = row[0], row[1], float(row[2]) if row[2] else 1.0
            if d not in leave_map:
                leave_map[d] = {'leave_type': lt, 'days': dy}

        # ── 2. 달력에 표시할 주 범위 전체 야간 근무 레코드 (work_type='night') ──
        night_rows = db.execute_query(
            "SELECT date, end_time FROM work_records "
            "WHERE work_type = 'night' AND date BETWEEN ? AND ? "
            "AND (leader LIKE ? OR teammates LIKE ?) "
            "ORDER BY date",
            (range_start, range_end, f'%{name}%', f'%{name}%')
        ) or []
        # date → max end_time (한 날에 여러 레코드 가능)
        night_map: Dict[str, str] = {}
        for row in night_rows:
            d, et = row[0], row[1] or ''
            if d not in night_map or et > night_map[d]:
                night_map[d] = et

        override_rows = db.execute_query(
            "SELECT work_date, start_time, end_time, note FROM work_hours_ot_overrides "
            "WHERE employee_name = ? AND work_date BETWEEN ? AND ?",
            (name, range_start, range_end)
        ) or []
        override_map: Dict[str, Any] = {}
        for row in override_rows:
            override_map[row[0]] = {
                'start_time': row[1] or '',
                'end_time': row[2] or '',
                'note': row[3] or ''
            }

        # ── 3. 달력에 표시할 주 범위 전체 휴일 근무 (holiday_work_entries) ──
        holiday_ot_map: Dict[str, float] = {}  # date → OT hours
        try:
            # 해당 주 범위에 걸친 금/토/일을 모두 포함하도록 period_key 후보를 넉넉히 수집
            check_start = calendar_start - _dt.timedelta(days=6)
            check_end = calendar_end + _dt.timedelta(days=6)
            candidate_fridays = []
            d = check_start
            while d <= check_end:
                if d.weekday() == 4:  # 금요일
                    candidate_fridays.append(d.strftime('%Y-%m-%d'))
                d += _dt.timedelta(days=1)

            for fri_key in candidate_fridays:
                hrows = db.execute_query(
                    "SELECT name, fri_work, sat_work, sun_work "
                    "FROM holiday_work_entries WHERE period_key = ?",
                    (fri_key,)
                ) or []
                for hr in hrows:
                    hname, fw, sw, sunw = hr[0], hr[1] or '-', hr[2] or '-', hr[3] or '-'
                    if hname != name:
                        continue
                    # 금요일 = fri_key, 토=+1, 일=+2
                    fri_dt = _dt.date.fromisoformat(fri_key)
                    for offset, val in [(0, fw), (1, sw), (2, sunw)]:
                        work_date = fri_dt + _dt.timedelta(days=offset)
                        if not (calendar_start <= work_date <= calendar_end):
                            continue
                        if val and val != '-':
                            ds = work_date.strftime('%Y-%m-%d')
                            holiday_ot_map[ds] = holiday_ot_map.get(ds, 0.0) + 8.0
        except Exception as he:
            logger.warning(f"휴일 근무 조회 오류 (무시): {he}")

        # ── OT 계산 헬퍼 ──
        def _parse_ot(end_time: str) -> float:
            """end_time 문자열(HH:MM)에서 17:00 기준 OT 시간 계산"""
            try:
                parts = end_time.strip().split(':')
                h, m = int(parts[0]), int(parts[1]) if len(parts) > 1 else 0
                total_min = h * 60 + m
                ot_min = total_min - 17 * 60  # 17:00 이후
                if ot_min <= 0:
                    return 0.0
                ot_hours = ot_min / 60.0
                if meal_deduct:
                    ot_hours = max(0.0, ot_hours - 1.0)  # 석식 1시간 공제
                return round(ot_hours, 2)
            except Exception:
                return 0.0

        def _parse_time_to_minutes(value: str) -> int | None:
            try:
                text = (value or '').strip()
                if not text:
                    return None
                if ':' in text:
                    h, m = text.split(':', 1)
                elif len(text) in (3, 4) and text.isdigit():
                    h, m = text[:-2], text[-2:]
                else:
                    return None
                hour = int(h)
                minute = int(m)
                if not (0 <= hour <= 23 and 0 <= minute <= 59):
                    return None
                return hour * 60 + minute
            except Exception:
                return None

        def _parse_override_ot(start_time: str, end_time: str) -> float:
            start_min = _parse_time_to_minutes(start_time)
            end_min = _parse_time_to_minutes(end_time)
            if start_min is None or end_min is None:
                return 0.0
            if end_min < start_min:
                end_min += 24 * 60
            return round(max(0, end_min - start_min) / 60.0, 2)

        # ── 일별 계산 ──
        result_days: Dict[str, Any] = {}
        current_day = calendar_start
        while current_day <= calendar_end:
            date_str = current_day.strftime('%Y-%m-%d')
            wd = current_day.weekday()  # 0=월, 6=일
            is_weekend = wd >= 5  # 토(5), 일(6)

            leave_info = leave_map.get(date_str)
            leave_type = leave_info['leave_type'] if leave_info else None
            is_future = current_day > today

            if is_future:
                regular = 0.0
                leave_type = None
            elif is_weekend:
                regular = 0.0
            elif leave_type in ('연차', '반차', '반반차'):
                regular = max(0.0, 8.0 - 8.0 * leave_info['days'])
            else:
                regular = 8.0

            ot = 0.0
            override_info = override_map.get(date_str)
            if not is_future:
                if override_info:
                    ot += _parse_override_ot(override_info.get('start_time'), override_info.get('end_time'))
                elif date_str in night_map:
                    ot += _parse_ot(night_map[date_str])
                if date_str in holiday_ot_map:
                    ot += holiday_ot_map[date_str]

            result_days[date_str] = {
                'regular': regular,
                'ot': round(ot, 2),
                'ot_start_time': override_info.get('start_time', '') if override_info else ('17:00' if date_str in night_map else ''),
                'ot_end_time': override_info.get('end_time', '') if override_info else night_map.get(date_str, ''),
                'ot_overridden': bool(override_info),
                'leave_type': leave_type,
                'is_weekend': is_weekend,
                'is_current_month': current_day.month == month,
                'is_future': is_future
            }
            current_day += _dt.timedelta(days=1)

        return {
            'success': True,
            'name': name,
            'year': year,
            'month': month,
            'meal_deduct': meal_deduct,
            'days': result_days
        }
    except Exception as e:
        logger.error(f"근로 시간 조회 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def search_records_by_ship(ship_name: str) -> List[Dict[str, Any]]:
    """선명으로 작업 내역 조회"""
    try:
        result = search_records_with_ot('ship', ship_name)
        return result.get('records', []) if result.get('success') else []
    except Exception as e:
        logger.error(f"선명 검색 오류: {e}")
        return []


@eel.expose
def search_records_by_company(company_name: str) -> List[Dict[str, Any]]:
    """외주 업체명으로 작업 내역 조회"""
    try:
        result = search_records_with_ot('company', company_name)
        return result.get('records', []) if result.get('success') else []
    except Exception as e:
        logger.error(f"업체명 검색 오류: {e}")
        return []


@eel.expose
def get_outsource_company_names() -> List[str]:
    """외주 업체명 목록 조회 (드롭다운 자동완성용)"""
    try:
        import re
        query = '''
            SELECT teammates FROM work_records
            WHERE teammates IS NOT NULL AND teammates != ''
        '''
        results = db.execute_query(query, ())
        companies = set()
        for row in results:
            teammates = row[0] or ''
            # 도급: 업체명(직원)
            for m in re.finditer(r'([^,\[\]()\n]+?)\(', teammates):
                co = m.group(1).replace('*', '').strip()
                if co:
                    companies.add(co)
            # 일당: 업체명[직원]
            for m in re.finditer(r'([^,\[\]()\n]+?)\[', teammates):
                co = m.group(1).replace('*', '').strip()
                if co:
                    companies.add(co)
        return sorted(list(companies))
    except Exception as e:
        logger.error(f"업체명 목록 조회 오류: {e}")
        return []


@eel.expose
def get_project_end_date(ship_name: str) -> str:
    """선박별 공사 마지막 작업일 조회"""
    try:
        if not ship_name or ship_name.strip() == '':
            return ''
        
        # 해당 선박의 마지막 작업일 조회
        query = '''
            SELECT MAX(date) as end_date
            FROM work_records
            WHERE ship_name = ?
            AND ship_name != ''
        '''
        
        result = db.execute_query(query, (ship_name,))
        if result and result[0][0]:
            from datetime import datetime
            end_date_str = result[0][0]  # YYYY-MM-DD 형식
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            # M/D 형식으로 반환
            return f"{end_date.month}/{end_date.day}"
        return ''
    except Exception as e:
        logger.error(f"공사 마지막일 조회 오류: {e}")
        return ''


@eel.expose
def debug_check_data(year: int, month: int) -> Dict[str, Any]:
    """디버깅용 데이터 확인"""
    try:
        # 전체 데이터 조회
        query1 = '''
            SELECT COUNT(*) as total
            FROM work_records
        '''
        total_result = db.execute_query(query1, ())
        total_count = total_result[0][0] if total_result else 0
        
        # 해당 월 데이터 조회
        year_str = str(year)
        month_str = str(month).zfill(2)
        
        query2 = '''
            SELECT COUNT(*) as month_total
            FROM work_records
            WHERE strftime('%Y', date) = ?
            AND strftime('%m', date) = ?
        '''
        month_result = db.execute_query(query2, (year_str, month_str))
        month_count = month_result[0][0] if month_result else 0
        
        # 모든 날짜 목록 조회
        query3 = '''
            SELECT DISTINCT date
            FROM work_records
            ORDER BY date DESC
            LIMIT 10
        '''
        dates_result = db.execute_query(query3, ())
        dates = [row[0] for row in dates_result] if dates_result else []
        
        return {
            'total_records': total_count,
            'month_records': month_count,
            'recent_dates': dates,
            'query_year': year_str,
            'query_month': month_str
        }
    except Exception as e:
        logger.error(f"디버깅 데이터 확인 오류: {e}")
        return {}


@eel.expose
def load_monthly_report_grouped(year: int, month: int) -> List[Dict[str, Any]]:
    """월간 보고서 - 선박별 그룹핑"""
    try:
        from datetime import datetime
        
        logger.info(f"월간 보고 요청: {year}년 {month}월")
        
        # 해당 월 작업 기록 조회
        # 계약번호가 다르면 별도 공사 → ship_name + contract_number 로 그룹핑
        # 시작/종료일: 계약번호 있으면 해당 계약 전체 기간, 없으면 선박명 기준
        query = '''
            SELECT
                wr.company,
                wr.ship_name,
                GROUP_CONCAT(DISTINCT wr.location) AS location,
                wr.engine_model,
                wr.work_content,
                wr.leader,
                wr.teammates,
                SUM(wr.manpower) as total_manpower,
                CASE
                    WHEN wr.contract_number != '' AND wr.contract_number IS NOT NULL THEN
                        (SELECT MIN(wr2.date) FROM work_records wr2
                         WHERE wr2.contract_number = wr.contract_number)
                    ELSE
                        (SELECT MIN(wr2.date) FROM work_records wr2
                         WHERE wr2.ship_name = wr.ship_name
                         AND (wr2.contract_number = '' OR wr2.contract_number IS NULL))
                END as true_start_date,
                CASE
                    WHEN wr.contract_number != '' AND wr.contract_number IS NOT NULL THEN
                        (SELECT MAX(wr2.date) FROM work_records wr2
                         WHERE wr2.contract_number = wr.contract_number)
                    ELSE
                        (SELECT MAX(wr2.date) FROM work_records wr2
                         WHERE wr2.ship_name = wr.ship_name
                         AND (wr2.contract_number = '' OR wr2.contract_number IS NULL))
                END as true_end_date,
                COALESCE(
                    (SELECT ps.status FROM project_status ps
                     WHERE ps.contract_number = wr.contract_number
                     AND wr.contract_number != ''
                     AND wr.contract_number IS NOT NULL
                     ORDER BY ps.updated_at DESC LIMIT 1),
                    (SELECT bp.status FROM board_projects bp
                     WHERE bp.contract_number = wr.contract_number
                     AND wr.contract_number != ''
                     AND wr.contract_number IS NOT NULL
                     ORDER BY bp.id DESC LIMIT 1),
                    (SELECT bp.status FROM board_projects bp
                     WHERE bp.ship_name = wr.ship_name ORDER BY bp.id DESC LIMIT 1)
                ) as board_status
            FROM work_records wr
            WHERE strftime('%Y', wr.date) = ?
            AND strftime('%m', wr.date) = ?
            AND wr.ship_name != ''
            GROUP BY wr.ship_name, wr.contract_number
            ORDER BY MIN(wr.date)
        '''

        year_str = str(year)
        month_str = str(month).zfill(2)
        logger.info(f"쿼리 파라미터: year={year_str}, month={month_str}")

        results = db.execute_query(query, (year_str, month_str))
        logger.info(f"조회 결과: {len(results) if results else 0}건")

        if not results:
            return []

        grouped_data = []
        for row in results:
            ship_name = row[1]

            # 공사기간: 전체 프로젝트 기간 기준, 보드 상태로 준공 여부 판단
            true_start_date = row[8] or ''
            true_end_date   = row[9] or ''
            board_status    = row[10] or ''

            if board_status == '준공':
                project_period = f"{true_start_date} ~ {true_end_date}"
            else:
                project_period = f"{true_start_date} ~ 진행중" if true_start_date else "진행중"
            
            # 작업내용 통합
            engine_model = row[3] or ''
            work_content = row[4] or ''
            
            if engine_model and work_content:
                full_work_content = f"{engine_model} {work_content}"
            elif engine_model:
                full_work_content = engine_model
            elif work_content:
                full_work_content = work_content
            else:
                full_work_content = '-'
            
            grouped_data.append({
                'company': row[0] or '',
                'ship_name': ship_name,
                'project_period': project_period,
                'location': row[2] or '',
                'work_content': full_work_content,
                'leader': row[5] or '',
                'teammates': row[6] or '',
                'total_manpower': float(row[7]) if row[7] else 0.0
            })
        
        return grouped_data
        
    except Exception as e:
        logger.error(f"월간 보고서 그룹핑 오류: {e}")
        return []


@eel.expose
def get_analytics_data(year: int) -> Dict[str, Any]:
    """E: 연간 통계 — 월별 공수 합계, 회사별 상위 10, 계약별 상위 10"""
    try:
        year_str = str(year)

        # 1) 월별 공수 합계
        monthly_rows = db.execute_query(
            "SELECT strftime('%m', date) as m, ROUND(SUM(manpower), 1) as total "
            "FROM work_records "
            "WHERE strftime('%Y', date) = ? "
            "GROUP BY m ORDER BY m",
            (year_str,)
        )
        monthly_map = {int(m): float(v) for m, v in (monthly_rows or [])}
        monthly_data = [monthly_map.get(i, 0.0) for i in range(1, 13)]

        # 2) 본공/외주 월별 분리 (Python 후처리)
        all_rows = db.execute_query(
            "SELECT strftime('%m', date) as m, leader, teammates "
            "FROM work_records WHERE strftime('%Y', date) = ?",
            (year_str,)
        )
        monthly_inhouse    = [0.0] * 12
        monthly_outsourced = [0.0] * 12
        for row in (all_rows or []):
            idx = int(row[0]) - 1
            ih, out = split_manpower_by_type(row[1] or '', row[2] or '')
            monthly_inhouse[idx]    += ih
            monthly_outsourced[idx] += out
        monthly_inhouse    = [round(v, 1) for v in monthly_inhouse]
        monthly_outsourced = [round(v, 1) for v in monthly_outsourced]

        # 3) 회사별 상위 10
        company_rows = db.execute_query(
            "SELECT company, ROUND(SUM(manpower), 1) as total "
            "FROM work_records "
            "WHERE strftime('%Y', date) = ? AND company != '' AND company IS NOT NULL "
            "GROUP BY company ORDER BY total DESC LIMIT 10",
            (year_str,)
        )

        # 4) 계약별 상위 10 (선명 포함)
        contract_rows = db.execute_query(
            "SELECT w1.contract_number, w1.company, "
            "(SELECT w2.ship_name FROM work_records w2 "
            " WHERE w2.contract_number = w1.contract_number "
            "   AND w2.ship_name != '' AND w2.ship_name IS NOT NULL LIMIT 1) AS ship_name, "
            "ROUND(SUM(w1.manpower), 1) AS total "
            "FROM work_records w1 "
            "WHERE strftime('%Y', w1.date) = ? AND w1.contract_number != '' AND w1.contract_number IS NOT NULL "
            "GROUP BY w1.contract_number ORDER BY total DESC LIMIT 10",
            (year_str,)
        )

        # 5) KPI 지표 추가
        # 5-1) 월별 고유 계약 건수
        monthly_cnt_rows = db.execute_query(
            "SELECT strftime('%m', date) as m, COUNT(DISTINCT contract_number) as cnt "
            "FROM work_records "
            "WHERE strftime('%Y', date) = ? AND contract_number != '' "
            "GROUP BY m ORDER BY m",
            (year_str,)
        )
        cnt_map = {int(m): int(c) for m, c in (monthly_cnt_rows or [])}
        monthly_project_count = [cnt_map.get(i, 0) for i in range(1, 13)]

        # 5-2) A/S 비율 (연간 전체 계약 건수 대비 A/S 발생 계약 건수)
        as_rows = db.execute_query(
            "SELECT "
            "  COUNT(DISTINCT CASE WHEN is_as = 1 THEN contract_number END), "
            "  COUNT(DISTINCT contract_number) "
            "FROM work_records "
            "WHERE strftime('%Y', date) = ? AND contract_number != ''",
            (year_str,)
        )
        as_cnt   = int(as_rows[0][0]) if as_rows else 0
        total_cn = int(as_rows[0][1]) if as_rows else 0
        as_rate  = round(as_cnt / total_cn * 100, 1) if total_cn > 0 else 0.0

        # 5-3) 외주 비율 & 연간 총 공수
        in_house_total   = round(sum(monthly_inhouse), 1)
        outsourced_total = round(sum(monthly_outsourced), 1)
        grand_total      = round(in_house_total + outsourced_total, 1)
        outsourced_rate  = round(outsourced_total / grand_total * 100, 1) if grand_total > 0 else 0.0

        return {
            'success':             True,
            'year':                year,
            'monthly':             monthly_data,
            'inHouseMonthly':      monthly_inhouse,
            'outsourcedMonthly':   monthly_outsourced,
            'inHouseTotal':        in_house_total,
            'outsourcedTotal':     outsourced_total,
            'companies': [{'name': r[0], 'total': float(r[1])} for r in (company_rows or [])],
            'contracts': [
                {'cn': r[0], 'company': r[1], 'ship': r[2] or r[0], 'total': float(r[3])}
                for r in (contract_rows or [])
            ],
            # KPI
            'monthlyProjectCount': monthly_project_count,
            'asRate':              as_rate,
            'asCnt':               as_cnt,
            'outsourcedRate':      outsourced_rate,
            'totalProjects':       total_cn,
        }
    except Exception as e:
        logger.error(f"통계 데이터 조회 오류: {e}")
        return {'success': False}


# =============================================================================
# v1.8.6 — 사용자 현황 / 오류 리포트
# =============================================================================

@eel.expose
def update_client_version(user_id: str, version: str) -> dict:
    """로그인 후 클라이언트가 자신의 버전을 서버에 등록"""
    try:
        auth_manager.update_user_version(user_id, version)
        return {'success': True}
    except Exception as e:
        logger.error(f"클라이언트 버전 등록 실패: {e}")
        return {'success': False}


@eel.expose
def report_error(user_id: str, error_type: str, error_message: str, stack_trace: str = '') -> dict:
    """JS/Python 오류를 DB에 저장"""
    try:
        user_name = ''
        try:
            user = auth_manager.get_user(user_id)
            if user:
                user_name = user.get('full_name', '')
        except Exception:
            pass
        app_version = config.get('update.current_version', '')
        db.add_error_report(user_id, user_name, app_version, error_type, error_message, stack_trace)
        return {'success': True}
    except Exception as e:
        logger.error(f"오류 리포트 저장 실패: {e}")
        return {'success': False}


@eel.expose
def admin_get_user_status(admin_id: str = '') -> list:
    """전체 사용자 + 버전 + 마지막 접속 (관리자 전용)"""
    try:
        user = auth_manager.get_user(admin_id) if admin_id else None
        if not user or user.get('role') != 'admin':
            return []
        rows = db.execute_query(
            "SELECT user_id, full_name, role, status, client_version, last_seen "
            "FROM auth_users ORDER BY CASE WHEN last_seen IS NULL THEN 1 ELSE 0 END, last_seen DESC",
            ()
        )
        if not rows:
            return []
        return [
            {
                'user_id': r[0],
                'full_name': r[1],
                'role': r[2],
                'status': r[3],
                'client_version': r[4],
                'last_seen': r[5],
            }
            for r in rows
        ]
    except Exception as e:
        logger.error(f"사용자 현황 조회 실패: {e}")
        return []


@eel.expose
def admin_get_error_reports(limit: int = 50, admin_id: str = '') -> list:
    """미해결 오류 리포트 목록 (관리자 전용)"""
    try:
        user = auth_manager.get_user(admin_id) if admin_id else None
        if not user or user.get('role') != 'admin':
            return []
        return db.get_error_reports(limit)
    except Exception as e:
        logger.error(f"오류 리포트 조회 실패: {e}")
        return []


@eel.expose
def admin_mark_error_read(error_id: int, admin_id: str = '') -> dict:
    """오류 리포트를 읽음 처리 (관리자 전용)"""
    try:
        user = auth_manager.get_user(admin_id) if admin_id else None
        if not user or user.get('role') != 'admin':
            return {'success': False}
        ok = db.mark_error_report_read(error_id)
        return {'success': ok}
    except Exception as e:
        logger.error(f"오류 리포트 읽음 처리 실패: {e}")
        return {'success': False}


@eel.expose
def admin_get_realtime_summary(admin_id: str = '') -> Dict[str, Any]:
    """관리자 실시간 현황 요약: 현재 접속, 오늘 작업률, 미결 오류 (관리자 전용)"""
    try:
        user = auth_manager.get_user(admin_id) if admin_id else None
        if not user or user.get('role') != 'admin':
            return {'success': False}

        from datetime import datetime, timedelta
        now = datetime.now()
        five_min_ago = (now - timedelta(minutes=5)).isoformat()
        today = now.strftime('%Y-%m-%d')

        # 현재 활성 사용자 (5분 내 last_seen)
        active_rows = db.execute_query(
            "SELECT COUNT(*) FROM auth_users WHERE last_seen >= ? AND status = 'active'",
            (five_min_ago,)
        )
        active_count = active_rows[0][0] if active_rows else 0

        # 오늘 작업 입력 건수 (비어있지 않은 레코드)
        today_rows = db.execute_query(
            "SELECT COUNT(*) FROM work_records WHERE date = ? AND (contract_number != '' OR work_content != '')",
            (today,)
        )
        today_filled = today_rows[0][0] if today_rows else 0

        # 미결 오류 리포트 수
        error_rows = db.execute_query(
            "SELECT COUNT(*) FROM error_reports WHERE is_read = 0",
            ()
        )
        error_count = error_rows[0][0] if error_rows else 0

        return {
            'success': True,
            'activeUsers': active_count,
            'todayFilledRecords': today_filled,
            'unresolvedErrors': error_count,
            'timestamp': now.strftime('%H:%M:%S')
        }
    except Exception as e:
        logger.error(f"실시간 요약 조회 오류: {e}")
        return {'success': False}


@eel.expose
def load_monthly_report(year: int, month: int) -> Dict[str, Any]:
    """월간 보고서 데이터 로드"""
    try:
        year_str = str(year)
        month_str = str(month).zfill(2)

        # 월 전체 데이터를 단일 쿼리로 조회 (N+1 제거)
        query = '''
            SELECT date, ship_name, work_content, company, manpower
            FROM work_records
            WHERE strftime('%Y', date) = ?
              AND strftime('%m', date) = ?
              AND (company != '' OR ship_name != '' OR work_content != '')
            ORDER BY date
        '''
        rows = db.execute_query(query, (year_str, month_str))

        # 날짜별로 그룹핑
        from collections import defaultdict
        day_groups: Dict[str, list] = defaultdict(list)
        for row in (rows or []):
            day_groups[row[0]].append(row)

        daily_data = []
        total_manpower = 0
        total_work_days = 0

        for date_str, day_rows in sorted(day_groups.items()):
            day = int(date_str.split('-')[2])
            day_manpower = sum(float(r[4] or 0) for r in day_rows)

            main_works = [r[1] for r in day_rows if r[1]]
            if not main_works:
                main_works = [r[2][:20] for r in day_rows if r[2]]

            daily_data.append({
                'date': date_str,
                'day': day,
                'work_count': len(day_rows),
                'manpower': day_manpower,
                'main_works': main_works[:3]
            })
            total_manpower += day_manpower
            total_work_days += 1

        avg_manpower = total_manpower / total_work_days if total_work_days > 0 else 0

        return {
            'total_work_days': total_work_days,
            'total_manpower': total_manpower,
            'avg_manpower': avg_manpower,
            'daily_data': daily_data
        }
    except Exception as e:
        logger.error(f"월간 보고서 로드 오류: {e}")
        return {
            'total_work_days': 0,
            'total_manpower': 0,
            'avg_manpower': 0,
            'daily_data': []
        }


@eel.expose
def export_to_excel(date: str) -> Dict[str, Any]:
    """Excel로 내보내기"""
    try:
        from pathlib import Path
        
        # 저장 경로 (바탕화면)
        desktop = Path.home() / "Desktop"
        filename = f"작업현황_{date}.xlsx"
        output_path = desktop / filename
        
        success = work_record_service.export_to_excel(date, str(output_path))
        
        return {
            'success': success,
            'path': str(output_path) if success else None,
            'message': f'Excel 파일이 저장되었습니다: {output_path}' if success else 'Excel 내보내기 실패'
        }
    except Exception as e:
        logger.error(f"Excel 내보내기 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def export_daily_report(date: str) -> Dict[str, Any]:
    """일일 보고서 Excel 내보내기"""
    try:
        from pathlib import Path
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from datetime import datetime
        
        # 데이터 로드
        records = work_record_service.get_records_for_date(date)
        valid_records = [r for r in records if r.get('company') or r.get('ship_name') or r.get('work_content')]
        
        # 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "일일보고"
        
        # 날짜 포맷
        date_obj = datetime.strptime(date, '%Y-%m-%d')
        weekdays = ['일', '월', '화', '수', '목', '금', '토']
        weekday = weekdays[date_obj.isoweekday() % 7]
        date_display = f"{date_obj.year}. {date_obj.month}월 {date_obj.day}일 ({weekday}) 현재"
        
        # 제목
        ws.merge_cells('A1:H1')
        ws['A1'] = '선진종합 선박별 작업 현황'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 날짜
        ws.merge_cells('A2:H2')
        ws['A2'] = date_display
        ws['A2'].font = Font(size=12)
        ws['A2'].alignment = Alignment(horizontal='right', vertical='center')
        
        # 헤더
        headers = ['No.', '선사', '선명', '공사기간', '장소', '작업내용', '담당공무', '협력업체']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')
        
        # 데이터
        # #1 — N+1 쿼리 방지: 선박명 배치 pre-fetch
        _batch_ships = list({r.get('ship_name', '') for r in valid_records if r.get('ship_name', '')})
        _project_dates = get_project_start_dates_batch([], _batch_ships) if _batch_ships else {}

        for row_idx, record in enumerate(valid_records, start=4):
            ship_name = record.get('ship_name', '')
            engine_model = record.get('engine_model', '')
            work_content = record.get('work_content', '')

            # 공사기간: 시작일 ~ 진행중
            start_date = _project_dates.get(ship_name, '') if ship_name else ''
            project_period = f"{start_date} ~ 진행중" if start_date else '진행중'
            
            # 작업내용: 엔진모델 + 작업내용
            if engine_model and work_content:
                full_work_content = f"{engine_model} {work_content}"
            elif engine_model:
                full_work_content = engine_model
            elif work_content:
                full_work_content = work_content
            else:
                full_work_content = ''
            
            # 본사/외주 분리
            leader = record.get('leader', '')
            teammates = record.get('teammates', '')
            in_house, outsourced = separate_workers(leader, teammates)
            
            ws.cell(row=row_idx, column=1, value=row_idx - 3)
            ws.cell(row=row_idx, column=2, value=record.get('company', ''))
            ws.cell(row=row_idx, column=3, value=ship_name)
            ws.cell(row=row_idx, column=4, value=project_period)
            ws.cell(row=row_idx, column=5, value=record.get('location', ''))
            ws.cell(row=row_idx, column=6, value=full_work_content)
            ws.cell(row=row_idx, column=7, value=in_house)
            ws.cell(row=row_idx, column=8, value=outsourced)
        
        # 셀 정렬 및 테두리
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=3, max_row=3+len(valid_records), min_col=1, max_col=8):
            for cell in row:
                cell.border = thin_border
                if cell.row > 3:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 20
        
        # 저장
        desktop = Path.home() / "Desktop"
        filename = f"일일보고_{date}.xlsx"
        output_path = desktop / filename
        wb.save(output_path)
        
        return {
            'success': True,
            'filename': str(output_path)
        }
    except Exception as e:
        logger.error(f"일일 보고서 내보내기 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def export_monthly_report(year: int, month: int) -> Dict[str, Any]:
    """월간 보고서 Excel 내보내기"""
    try:
        from pathlib import Path
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from calendar import monthrange
        
        # 데이터 로드
        report_data = load_monthly_report(year, month)
        
        # 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "월간보고"
        
        # 제목
        ws.merge_cells('A1:D1')
        ws['A1'] = f'{year}년 {month}월 작업 현황'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 통계
        ws.merge_cells('A2:D2')
        ws['A2'] = f"총 작업일수: {report_data['total_work_days']}일 | 총 투입인원: {report_data['total_manpower']:.1f}공 | 평균: {report_data['avg_manpower']:.1f}공/일"
        ws['A2'].font = Font(size=12)
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 헤더
        headers = ['날짜', '작업건수', '투입인원', '주요작업']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='D9B3FF', end_color='D9B3FF', fill_type='solid')
        
        # 데이터
        for row_idx, day_data in enumerate(report_data['daily_data'], start=4):
            date_obj = datetime.strptime(day_data['date'], '%Y-%m-%d')
            weekdays = ['일', '월', '화', '수', '목', '금', '토']
            weekday = weekdays[date_obj.isoweekday() % 7]
            
            ws.cell(row=row_idx, column=1, value=f"{month}/{day_data['day']} ({weekday})")
            ws.cell(row=row_idx, column=2, value=day_data['work_count'])
            ws.cell(row=row_idx, column=3, value=f"{day_data['manpower']:.1f}")
            ws.cell(row=row_idx, column=4, value=', '.join(day_data['main_works']))
        
        # 테두리
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=3, max_row=3+len(report_data['daily_data']), min_col=1, max_col=4):
            for cell in row:
                cell.border = thin_border
                if cell.row > 3:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 열 너비
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 50
        
        # 저장
        desktop = Path.home() / "Desktop"
        filename = f"월간보고_{year}{month:02d}.xlsx"
        output_path = desktop / filename
        wb.save(output_path)
        
        return {
            'success': True,
            'filename': str(output_path)
        }
    except Exception as e:
        logger.error(f"월간 보고서 내보내기 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


# ============================================================================
# 클라우드 동기화
# ============================================================================

@eel.expose
def sync_to_cloud() -> Dict[str, Any]:
    """클라우드로 동기화"""
    try:
        if not cloud_sync.enabled:
            return {'success': False, 'message': '클라우드 동기화가 비활성화되어 있습니다.'}
        
        success = cloud_sync.sync_to_cloud()
        return {
            'success': success,
            'message': '클라우드 동기화 완료' if success else '클라우드 동기화 실패'
        }
    except Exception as e:
        logger.error(f"클라우드 동기화 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def sync_from_cloud() -> Dict[str, Any]:
    """클라우드에서 동기화"""
    try:
        if not cloud_sync.enabled:
            return {'success': False, 'message': '클라우드 동기화가 비활성화되어 있습니다.'}
        
        success = cloud_sync.sync_from_cloud()
        return {
            'success': success,
            'message': '클라우드에서 동기화 완료' if success else '클라우드에서 동기화 실패'
        }
    except Exception as e:
        logger.error(f"클라우드 동기화 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def get_sync_status() -> Dict[str, Any]:
    """동기화 상태 조회"""
    try:
        return cloud_sync.get_sync_status()
    except Exception as e:
        logger.error(f"동기화 상태 조회 오류: {e}")
        return {}


@eel.expose
def get_cloud_sync_mode() -> str:
    """현재 sync_mode 반환: 'company' | 'external' | 'standalone'"""
    try:
        return cloud_sync.sync_mode
    except Exception as e:
        logger.error(f"sync_mode 조회 오류: {e}")
        return 'standalone'


@eel.expose
def connect_to_cloud_external(cloud_path: str) -> Dict[str, Any]:
    """
    외부 PC에서 클라우드에 연결 (관리자 전용)
    클라우드 경로 입력 → pull → lock 생성 → sync_mode=external
    """
    try:
        logger.info(f"외부 PC 클라우드 연결 요청: {cloud_path}")
        result = cloud_sync.connect_external(cloud_path)
        if result.get('success'):
            # db 싱글톤은 per-request 연결 방식이므로 재연결 불필요
            # (다음 쿼리부터 교체된 DB 파일을 자동으로 사용)
            logger.info("외부 PC 클라우드 연결 완료 — 다음 DB 접근부터 클라우드 DB 사용")
        return result
    except Exception as e:
        logger.error(f"외부 PC 클라우드 연결 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def disconnect_from_cloud() -> Dict[str, Any]:
    """
    외부 PC 클라우드 연결 해제 (관리자 전용)
    push + 알림 생성 + 잠금 삭제 + sync_mode=standalone
    """
    try:
        logger.info("외부 PC 클라우드 연결 해제 요청")
        return cloud_sync.disconnect_external()
    except Exception as e:
        logger.error(f"외부 PC 클라우드 연결 해제 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


# ============================================================================
# 앱 정보
# ============================================================================

@eel.expose
def get_app_info() -> Dict[str, Any]:
    """앱 정보 조회"""
    try:
        from src.utils.patch_system import patch_system
        patch_system._correct_version_from_applied_patches()
    except Exception:
        pass
    return {
        'name': config.app_name,
        'version': config.version,
        'db_path': str(config.db_path),
        'cloud_sync_enabled': cloud_sync.enabled,
        'cloud_folder': str(cloud_sync.cloud_folder) if cloud_sync.cloud_folder else None
    }


@eel.expose
def get_activity_logs(limit: int = 100, user_filter: str = '') -> List[Dict[str, Any]]:
    """활동 로그 조회 (관리자)"""
    try:
        logs = db.get_activity_logs(limit=limit, user=user_filter or None)
        return [log.to_dict() for log in logs]
    except Exception as e:
        logger.error(f"활동 로그 조회 오류: {e}")
        return []


# ============================================================================
# 대시보드 - 간트 차트
# ============================================================================

@eel.expose
def get_gantt_data(year: int, month: int) -> List[Dict[str, Any]]:
    """간트 차트용 프로젝트 데이터 조회 (해당 월과 겹치는 모든 프로젝트)"""
    try:
        import calendar
        last_day = calendar.monthrange(year, month)[1]
        month_start = f'{year}-{month:02d}-01'
        month_end = f'{year}-{month:02d}-{last_day:02d}'

        logger.info(f"간트 데이터 조회: {month_start} ~ {month_end}")

        # 해당 월과 기간이 겹치는 프로젝트 조회
        query = '''
            SELECT contract_number, company, ship_name, engine_model, work_content,
                   MIN(date) as start_date, MAX(date) as end_date,
                   COUNT(DISTINCT date) as work_days,
                   ROUND(SUM(manpower), 1) as total_manpower
            FROM work_records
            WHERE contract_number != '' AND contract_number IS NOT NULL
            GROUP BY contract_number
            HAVING end_date >= ? AND start_date <= ?
            ORDER BY start_date
        '''
        rows = db.execute_query(query, (month_start, month_end))

        # N+1 방지: 모든 계약번호의 해당 월 작업일을 단일 쿼리로 조회
        work_dates_map = {}
        if rows:
            all_cns = [row[0] for row in rows if row[0]][:500]
            if all_cns:
                placeholders = ','.join('?' * len(all_cns))
                date_rows_batch = db.execute_query(
                    f'SELECT contract_number, date FROM work_records '
                    f'WHERE contract_number IN ({placeholders}) AND date >= ? AND date <= ? '
                    f'ORDER BY contract_number, date',
                    all_cns + [month_start, month_end]
                )
                for cn, d in (date_rows_batch or []):
                    work_dates_map.setdefault(cn, []).append(d)

        projects = []
        for row in rows:
            contract_number = row[0] or ''
            work_dates = work_dates_map.get(contract_number, [])

            start_date = row[5] or ''
            end_date = row[6] or ''
            # M/D 형식으로 변환 (잘못된 날짜 형식 → 빈 문자열)
            start_md = _date_to_md(start_date)
            end_md = _date_to_md(end_date)

            projects.append({
                'contractNumber': contract_number,
                'company': row[1] or '',
                'shipName': row[2] or '',
                'engineModel': row[3] or '',
                'workContent': row[4] or '',
                'startDate': start_date,
                'endDate': end_date,
                'startMD': start_md,
                'endMD': end_md,
                'workDays': row[7] or 0,
                'totalManpower': row[8] or 0,
                'workDates': work_dates
            })

        logger.info(f"간트 데이터: {len(projects)}개 프로젝트")
        return projects

    except Exception as e:
        logger.error(f"간트 데이터 조회 오류: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return []


# ============================================================================
# 대시보드 - 칸반 보드
# ============================================================================

@eel.expose
def set_project_status(contract_number: str, status: str) -> Dict[str, Any]:
    """프로젝트 상태 수동 변경 (접수/착수/준공/auto)"""
    try:
        username = ''
        success = db.set_project_status(contract_number, status, username)
        if success:
            # 착공·준공 이벤트 텔레그램 알림 (비동기)
            if status in ('착수', '준공'):
                ship_name = telegram_notifier._get_ship_name(contract_number, None)
                _start_tracked_thread(
                    target=telegram_notifier.send_project_event,
                    args=(contract_number, status, ship_name)
                )
            return {'success': True, 'message': '상태가 변경되었습니다.'}
        return {'success': False, 'message': '상태 변경 실패'}
    except Exception as e:
        logger.error(f"프로젝트 상태 변경 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def create_board_project(data: Dict) -> Dict[str, Any]:
    """보드 프로젝트 생성 (접수 단계)"""
    try:
        username = data.get('username', '')
        project_id = db.create_board_project(data, username)
        if project_id:
            return {'success': True, 'id': project_id, 'message': '프로젝트가 등록되었습니다.'}
        return {'success': False, 'message': '프로젝트 등록 실패'}
    except Exception as e:
        logger.error(f"보드 프로젝트 생성 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def update_board_project(project_id: int, data: Dict) -> Dict[str, Any]:
    """보드 프로젝트 업데이트"""
    try:
        username = data.get('username', '')
        success = db.update_board_project(project_id, data, username)
        if success:
            return {'success': True, 'message': '프로젝트가 수정되었습니다.'}
        return {'success': False, 'message': '프로젝트 수정 실패'}
    except Exception as e:
        logger.error(f"보드 프로젝트 수정 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def update_project_milestones(project_id: int, target_start: str, target_end: str, actual_end: str) -> Dict[str, Any]:
    """프로젝트 마일스톤 날짜 업데이트"""
    try:
        data = {
            'target_start_date': target_start or '',
            'target_end_date':   target_end   or '',
            'actual_end_date':   actual_end   or '',
        }
        success = db.update_board_project(project_id, data)
        if success:
            return {'success': True, 'message': '마일스톤이 저장되었습니다.'}
        return {'success': False, 'message': '마일스톤 저장 실패'}
    except Exception as e:
        logger.error(f"마일스톤 업데이트 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def delete_board_project(project_id: int) -> Dict[str, Any]:
    """보드 프로젝트 삭제"""
    try:
        success = db.delete_board_project(project_id)
        if success:
            return {'success': True, 'message': '프로젝트가 삭제되었습니다.'}
        return {'success': False, 'message': '프로젝트 삭제 실패'}
    except Exception as e:
        logger.error(f"보드 프로젝트 삭제 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}



@eel.expose
def get_employee_profile(name: str, year: int = 0) -> Dict[str, Any]:
    """직원별 연간 공수·프로젝트·연차 현황 조회"""
    try:
        target_year = year if year else datetime.now().year
        rows = db.execute_query(
            "SELECT date, leader, teammates, manpower, contract_number, ship_name "
            "FROM work_records "
            "WHERE (leader LIKE ? OR teammates LIKE ?) AND date LIKE ? "
            "ORDER BY date",
            (f'%{name}%', f'%{name}%', f'{target_year}%')
        )
        monthly = [0.0] * 12
        projects: set = set()
        total_manpower = 0.0
        for row in (rows or []):
            date_str  = row[0] or ''
            leader    = row[1] or ''
            teammates = row[2] or ''
            contract  = row[4] or ''
            ship      = row[5] or ''
            try:
                month_idx = int(date_str.split('-')[1]) - 1
            except (IndexError, ValueError):
                month_idx = 0
            in_h, out = split_manpower_by_type(leader, teammates)
            manpower  = round(in_h + out, 2)
            monthly[max(0, min(11, month_idx))] += manpower
            total_manpower += manpower
            if contract:
                projects.add(contract)
            elif ship:
                projects.add(ship)

        # 연차 잔여
        leave_rows = db.execute_query(
            "SELECT total_days, used_annual, used_half "
            "FROM vacation_balances WHERE year=? AND employee_name=?",
            (target_year, name)
        )
        leave_balance = None
        if leave_rows:
            total_d  = float(leave_rows[0][0] or 0)
            used_a   = float(leave_rows[0][1] or 0)
            used_h   = float(leave_rows[0][2] or 0)
            remaining = total_d - used_a - used_h * 0.5
            leave_balance = {
                'total': total_d,
                'used': round(used_a + used_h * 0.5, 1),
                'remaining': round(remaining, 1)
            }
        return {
            'success': True,
            'name': name,
            'year': target_year,
            'totalManpower': round(total_manpower, 1),
            'monthlyManpower': [round(m, 1) for m in monthly],
            'projectCount': len(projects),
            'projects': sorted(projects)[:10],
            'leaveBalance': leave_balance,
        }
    except Exception as e:
        logger.error(f"직원 프로필 조회 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def estimate_completion(engine_model: str, work_content: str, target_start: str = '') -> Dict[str, Any]:
    """과거 유사 작업 기간 기반 완료일 예측"""
    try:
        from datetime import datetime as _dt, timedelta as _td
        rows = db.execute_query(
            "SELECT target_start_date, actual_end_date FROM board_projects "
            "WHERE actual_end_date != '' AND target_start_date != '' "
            "AND (engine_model = ? OR work_content LIKE ?) "
            "ORDER BY id DESC LIMIT 30",
            (engine_model or '', f'%{(work_content or "")[:10]}%')
        )
        durations = []
        for row in (rows or []):
            try:
                s = row[0]; e = row[1]
                d = (_dt.strptime(e, '%Y-%m-%d') - _dt.strptime(s, '%Y-%m-%d')).days
                if 1 <= d <= 365:
                    durations.append(d)
            except Exception:
                pass

        if not durations:
            return {'success': True, 'avgDays': None, 'sampleCount': 0, 'suggestionEndDate': ''}

        avg_days = round(sum(durations) / len(durations))
        suggestion_end = ''
        if target_start:
            try:
                end_dt = _dt.strptime(target_start, '%Y-%m-%d') + _td(days=avg_days)
                suggestion_end = end_dt.strftime('%Y-%m-%d')
            except Exception:
                pass

        return {
            'success': True,
            'avgDays': avg_days,
            'sampleCount': len(durations),
            'suggestionEndDate': suggestion_end,
        }
    except Exception as e:
        logger.error(f"완료 예측 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def get_kanban_data() -> Dict[str, Any]:
    """칸반 보드용 프로젝트 데이터 (접수/착수/준공/아카이브 4단계)"""
    try:
        from datetime import datetime, timedelta

        now = datetime.now()
        cutoff = (now - timedelta(days=7)).strftime('%Y-%m-%d')
        current_month_start = now.strftime('%Y-%m-01')

        # 1. 접수 단계 프로젝트 (board_projects 테이블에서)
        reception_projects = db.get_board_projects('접수')
        reception = []
        for bp in reception_projects:
            reception.append({
                'id': bp['id'],
                'contractNumber': bp.get('contract_number', ''),
                'company': bp.get('company', ''),
                'shipName': bp.get('ship_name', ''),
                'engineModel': bp.get('engine_model', ''),
                'workContent': bp.get('work_content', ''),
                'startDate': '',
                'endDate': '',
                'startMD': '',
                'endMD': '',
                'workDays': 0,
                'totalManpower': 0,
                'status': '접수',
                'source': 'board',
                'targetStartDate': bp.get('target_start_date', ''),
                'targetEndDate':   bp.get('target_end_date', ''),
                'actualEndDate':   bp.get('actual_end_date', ''),
            })

        # 2. 착수/준공 - work_records 기반 + board_projects 수동 상태
        query = '''
            SELECT contract_number, company, ship_name, engine_model, work_content,
                   MIN(date) as start_date, MAX(date) as end_date,
                   COUNT(DISTINCT date) as work_days,
                   ROUND(SUM(manpower), 1) as total_manpower
            FROM work_records
            WHERE contract_number != '' AND contract_number IS NOT NULL
            GROUP BY contract_number
            ORDER BY contract_number DESC
        '''
        rows = db.execute_query(query)

        # 수동 상태 조회 (project_status 테이블)
        manual_statuses = db.get_project_statuses()

        # board_projects에서 착수/준공 상태인 것도 조회
        board_started = {bp['contract_number']: bp for bp in db.get_board_projects('착수') if bp.get('contract_number')}
        board_done = {bp['contract_number']: bp for bp in db.get_board_projects('준공') if bp.get('contract_number')}

        started = []
        done = []
        archive = []

        for row in rows:
            contract_number = row[0] or ''
            end_date = row[6] or ''
            start_date = row[5] or ''

            # M/D 형식 (잘못된 날짜 형식 → 빈 문자열)
            start_md = _date_to_md(start_date)
            end_md = _date_to_md(end_date)

            # 상태 판단: 수동 > 자동
            manual_status = manual_statuses.get(contract_number, '')
            board_status = ''
            if contract_number in board_started:
                board_status = '착수'
            elif contract_number in board_done:
                board_status = '준공'

            # 우선순위: project_status > board_projects > 자동
            if manual_status and manual_status != 'auto':
                # 기존 project_status의 값 매핑
                if manual_status in ('inProgress', '착수'):
                    final_status = '착수'
                elif manual_status in ('completed', '준공'):
                    final_status = '준공'
                else:
                    final_status = manual_status
            elif board_status:
                final_status = board_status
            else:
                # 자동 판단
                if end_date >= cutoff:
                    final_status = '착수'
                else:
                    final_status = '준공'

            # 마일스톤 데이터 — board_projects에서 계약번호로 매칭
            bp_data = board_started.get(contract_number) or board_done.get(contract_number) or {}
            project = {
                'contractNumber': contract_number,
                'company': row[1] or '',
                'shipName': row[2] or '',
                'engineModel': row[3] or '',
                'workContent': row[4] or '',
                'startDate': start_date,
                'endDate': end_date,
                'startMD': start_md,
                'endMD': end_md,
                'workDays': row[7] or 0,
                'totalManpower': row[8] or 0,
                'status': final_status,
                'manualStatus': manual_status,
                'source': 'records',
                'boardProjectId': bp_data.get('id', None),
                'targetStartDate': bp_data.get('target_start_date', ''),
                'targetEndDate':   bp_data.get('target_end_date', ''),
                'actualEndDate':   bp_data.get('actual_end_date', ''),
            }

            if final_status == '착수':
                started.append(project)
            elif final_status == '준공':
                # 준공인데 해당 월 지남 → 아카이브
                if end_date and end_date < current_month_start:
                    project['status'] = '아카이브'
                    archive.append(project)
                else:
                    done.append(project)

        logger.info(f"칸반 데이터: 접수 {len(reception)}건, 착수 {len(started)}건, 준공 {len(done)}건, 아카이브 {len(archive)}건")
        return {
            'reception': reception,
            'started': started,
            'done': done,
            'archive': archive
        }

    except Exception as e:
        logger.error(f"칸반 데이터 조회 오류: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return {'reception': [], 'started': [], 'done': [], 'archive': []}


@eel.expose
def get_or_create_board_project(contract_number: str) -> Dict[str, Any]:
    """착수 직접 등록 프로젝트 — board_projects 항목이 없으면 자동 생성 후 ID 반환.
    마일스톤 편집 버튼 클릭 시 boardProjectId 없는 카드에서 호출."""
    try:
        cn = (contract_number or '').strip()
        if not cn:
            return {'success': False, 'message': '계약번호가 없습니다.'}
        with db.get_connection() as conn:
            # 이미 존재하는 항목 확인
            row = conn.execute(
                'SELECT id FROM board_projects WHERE contract_number = ?', (cn,)
            ).fetchone()
            if row:
                return {'success': True, 'projectId': row['id']}
            # work_records에서 기본 정보 조회
            rec = conn.execute(
                '''SELECT company, ship_name, engine_model, work_content
                   FROM work_records
                   WHERE contract_number = ? AND company != ''
                   ORDER BY date DESC LIMIT 1''',
                (cn,)
            ).fetchone()
            from datetime import datetime as _dt
            now = _dt.now().isoformat()
            company      = rec['company']      if rec else ''
            ship_name    = rec['ship_name']    if rec else ''
            engine_model = rec['engine_model'] if rec else ''
            work_content = rec['work_content'] if rec else ''
            cursor = conn.execute(
                '''INSERT INTO board_projects
                   (contract_number, company, ship_name, engine_model,
                    work_content, status, created_at, created_by, updated_at)
                   VALUES (?, ?, ?, ?, ?, '착수', ?, '', ?)''',
                (cn, company, ship_name, engine_model, work_content, now, now)
            )
            new_id = cursor.lastrowid
        logger.info(f"board_projects 자동 생성: {cn} → id={new_id}")
        return {'success': True, 'projectId': new_id}
    except Exception as e:
        logger.error(f"board_projects 생성 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


# ============================================================================
# 댓글 시스템
# ============================================================================

@eel.expose
def add_project_comment(contract_number: str, content: str, parent_id: int = None, board_project_id: int = None) -> Dict[str, Any]:
    """프로젝트 댓글 추가 (add_project_comment_with_user로 위임)"""
    return add_project_comment_with_user(contract_number, content, '', '', parent_id, board_project_id)


@eel.expose
def add_project_comment_with_user(contract_number, content, user_id, user_name, parent_id=None, board_project_id=None) -> Dict[str, Any]:
    """프로젝트 댓글 추가 (사용자 정보 포함)"""
    try:
        # Eel null/0 처리
        if not parent_id or parent_id == 'null':
            parent_id = None
        else:
            parent_id = int(parent_id)
        if not board_project_id or board_project_id == 'null':
            board_project_id = None
        else:
            board_project_id = int(board_project_id)
        if not contract_number or contract_number == 'null':
            contract_number = ''
        # user_id/user_name 방어 (NOT NULL 제약 대응)
        if not user_id or user_id == 'null':
            user_id = 'unknown'
        if not user_name or user_name == 'null':
            user_name = '익명'
        comment_id = db.add_comment(contract_number, user_id, user_name, content, parent_id, board_project_id)
        if comment_id:
            # 텔레그램 알림 발송 (실패해도 댓글은 정상 등록)
            try:
                ship_name = telegram_notifier._get_ship_name(contract_number, board_project_id)
                telegram_notifier.send_comment_notification(
                    contract_number=contract_number or '',
                    board_project_id=board_project_id,
                    commenter_user_id=user_id,
                    commenter_name=user_name,
                    comment_text=content,
                    ship_name=ship_name
                )
            except Exception as te:
                logger.error(f"텔레그램 알림 발송 실패: {te}")
            return {'success': True, 'id': comment_id, 'message': '댓글이 등록되었습니다.'}
        return {'success': False, 'message': '댓글 등록 실패'}
    except Exception as e:
        logger.error(f"댓글 추가 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def get_project_comments(contract_number=None, board_project_id=None) -> Dict[str, Any]:
    """프로젝트 댓글 조회"""
    try:
        # Eel에서 null/0 → None 변환 처리
        if not contract_number or contract_number == 'null':
            contract_number = None
        if not board_project_id or board_project_id == 'null':
            board_project_id = None
        else:
            board_project_id = int(board_project_id)
        comments = db.get_comments(contract_number, board_project_id)
        # camelCase 변환
        result = []
        for c in comments:
            result.append({
                'id': c['id'],
                'contractNumber': c.get('contract_number', ''),
                'boardProjectId': c.get('board_project_id'),
                'parentId': c.get('parent_id'),
                'userId': c.get('user_id', ''),
                'userName': c.get('user_name', ''),
                'content': c.get('content', ''),
                'createdAt': c.get('created_at', '')
            })
        return {'success': True, 'comments': result}
    except Exception as e:
        logger.error(f"댓글 조회 오류: {e}")
        return {'success': False, 'comments': []}


@eel.expose
def delete_project_comment(comment_id: int, user_id: str) -> Dict[str, Any]:
    """프로젝트 댓글 삭제"""
    try:
        success = db.delete_comment(comment_id, user_id)
        if success:
            return {'success': True, 'message': '댓글이 삭제되었습니다.'}
        return {'success': False, 'message': '삭제 권한이 없거나 댓글을 찾을 수 없습니다.'}
    except Exception as e:
        logger.error(f"댓글 삭제 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


# ============================================================================
# 엑셀 불러오기
# ============================================================================

@eel.expose
def import_excel_data(base64_data: str, username: str = 'admin') -> Dict[str, Any]:
    """엑셀 파일 데이터를 DB로 일괄 업로드"""
    try:
        user = auth_manager.get_user(username) if username else None
        if not user or user.get('role') != 'admin':
            return {'success': False, 'message': '관리자 권한 필요'}

        import base64
        import io
        from openpyxl import load_workbook
        from ..database.models import WorkRecord
        from ..business.calculations import calculate_record_manpower

        logger.info(f"엑셀 불러오기 시작 by {username}")

        # 셀 값 추출 헬퍼 (기울임체 감지 포함)
        def get_cell_value(cell, wrap_italic=False):
            """셀 값을 문자열로 추출. wrap_italic=True이면 기울임체 셀에 *이름* 형태로 반환"""
            if cell is None or cell.value is None:
                return ''
            val = str(cell.value).strip()
            if not val:
                return ''
            if wrap_italic and cell.font and cell.font.italic:
                return f'*{val}*'
            return val

        # base64 디코딩
        file_bytes = base64.b64decode(base64_data)

        # 파일 형식 감지: OLE2 magic bytes = .xls, ZIP header = .xlsx
        XLS_MAGIC = b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'
        is_xls = file_bytes[:8] == XLS_MAGIC

        if is_xls:
            # .xls (Excel 97-2003) → xlrd 사용 (이탤릭 감지 불가)
            try:
                import xlrd
            except ImportError:
                return {'success': False, 'message': 'xlrd 패키지가 필요합니다. 터미널에서 [pip install xlrd] 를 실행해 주세요.'}
            wb_xls = xlrd.open_workbook(file_contents=file_bytes)
            ws_xls = wb_xls.sheet_by_index(0)
            # 날짜 타입 셀(type 3)을 "날짜 : YYYY.MM.DD." 텍스트로 변환하여
            # 날짜 감지 정규식이 float 값을 놓치지 않도록 처리
            rows_raw = []
            for i in range(ws_xls.nrows):
                vals = list(ws_xls.row_values(i))
                types = ws_xls.row_types(i)
                for j, (v, t) in enumerate(zip(vals, types)):
                    if t == 3:  # XL_CELL_DATE
                        try:
                            dt = xlrd.xldate.xldate_as_datetime(v, wb_xls.datemode)
                            vals[j] = f"날짜 : {dt.strftime('%Y.%m.%d')}."
                        except Exception:
                            pass
                rows_raw.append(vals)
            logger.info(f"xls 파일 감지: {ws_xls.nrows}행")
        else:
            # .xlsx (Office Open XML) → openpyxl 사용
            wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
            ws = wb.active
            rows_raw = [[c.value for c in row] for row in ws.iter_rows()]
            wb.close()
            logger.info(f"xlsx 파일 감지: {len(rows_raw)}행")

        # 날짜별 레코드 그룹핑
        date_records = {}  # { 'YYYY-MM-DD': [WorkRecord, ...] }
        skipped = 0
        total_records = 0

        import re
        current_date = None

        for row_values in rows_raw:
            # row_values: list of cell values (str/float/None)
            row_text = ' '.join(str(v or '') for v in row_values)

            # 날짜 헤더 행 감지: "날짜 : YYYY.MM.DD" 형식 (구분자 유연화)
            date_match = re.search(
                r'날짜\s*[:\：]?\s*(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})',
                row_text
            )
            if not date_match:
                # "날짜 :" 없이 YYYY.MM.DD 단독 셀인 경우
                date_match = re.search(
                    r'^\s*(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})\s*$',
                    row_text.strip()
                )
            if date_match:
                y, m, d = date_match.groups()
                current_date = f'{y}-{int(m):02d}-{int(d):02d}'
                continue

            # 컬럼 헤더 행 스킵 (계약번호, 선사/선명 포함 행)
            if '계약번호' in row_text or ('선사' in row_text and '선명' in row_text):
                continue

            # 날짜가 아직 설정되지 않은 경우 건너뜀
            if current_date is None:
                continue

            # B열(인덱스 1) 이상이어야 데이터 행으로 처리
            if len(row_values) < 2:
                skipped += 1
                continue

            def _cell_str(val):
                """셀 값을 문자열로 변환 (None/숫자 안전)"""
                if val is None:
                    return ''
                # xlrd가 숫자형으로 반환하는 경우 정수 변환
                if isinstance(val, float) and val == int(val):
                    return str(int(val))
                return str(val).strip()

            # B~I열 매핑 (실제 엑셀 형식)
            # B(1)=계약번호, C(2)=선사, D(3)=선명, E(4)=엔진/모델
            # F(5)=작업내용/장소, G(6)=작업자, H(7)=인원(무시), I(8)=동반자
            contract_number = _cell_str(row_values[1]) if len(row_values) > 1 else ''
            company         = _cell_str(row_values[2]) if len(row_values) > 2 else ''
            ship_name       = _cell_str(row_values[3]).upper() if len(row_values) > 3 else ''
            engine_model    = _cell_str(row_values[4]).upper() if len(row_values) > 4 else ''
            work_content    = _cell_str(row_values[5]) if len(row_values) > 5 else ''
            leader_raw      = _cell_str(row_values[6]) if len(row_values) > 6 else ''
            leader          = leader_raw  # xls/xlsx 공통: 이탤릭 감지는 xlsx만 가능
            # H열(인덱스 7): 인원 수 — 자동 계산으로 대체, 건너뜀
            teammates       = _cell_str(row_values[8]) if len(row_values) > 8 else ''

            # 유효 데이터 확인 (빈 행 제외)
            if not any([contract_number, company, ship_name, work_content]):
                skipped += 1
                continue

            # 인원 계산
            manpower = calculate_record_manpower(leader, teammates)

            # 날짜별 그룹핑
            if current_date not in date_records:
                date_records[current_date] = []

            record = WorkRecord(
                date=current_date,
                record_number=len(date_records[current_date]) + 1,
                contract_number=contract_number,
                company=company,
                ship_name=ship_name,
                engine_model=engine_model,
                work_content=work_content,
                location='',
                leader=leader,
                teammates=teammates,
                manpower=manpower
            )
            date_records[current_date].append(record)
            total_records += 1

        # DB에 날짜별로 저장
        saved_dates = 0
        for date_str, records in date_records.items():
            success = db.save_work_records(date_str, records, username)
            if success:
                saved_dates += 1
            else:
                logger.error(f"날짜 {date_str} 저장 실패")

        logger.info(f"엑셀 불러오기 완료: {saved_dates}일, {total_records}건")

        return {
            'success': True,
            'total_dates': saved_dates,
            'total_records': total_records,
            'skipped': skipped
        }

    except Exception as e:
        logger.error(f"엑셀 불러오기 오류: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return {
            'success': False,
            'message': '엑셀 불러오기 실패 — 파일 형식을 확인하세요.'
        }


# ============================================================================
# 업데이트
# ============================================================================

@eel.expose
def get_startup_patch_result() -> Dict[str, Any]:
    """앱 시작 시 자동 적용된 패치 결과 반환 (로그인 후 JS에서 호출)"""
    try:
        from src.utils.patch_system import patch_system
        patch_system._correct_version_from_applied_patches()
        applied = getattr(patch_system, '_startup_patches_applied', 0)
        version = config.version

        # 자동 재시작 후 첫 실행 시 마커 파일에서 업데이트 정보 읽기
        import json as _json
        from pathlib import Path as _Path
        marker_path = _Path(__file__).parent.parent.parent / "data" / "just_updated.json"
        if marker_path.exists():
            try:
                data = _json.loads(marker_path.read_text(encoding='utf-8'))
                applied = data.get('applied_count', applied)
                version = data.get('version', version)
                marker_path.unlink()  # 1회 읽고 삭제
                logger.info(f"자동 재시작 마커 확인: v{version} 패치 {applied}개")
            except Exception as me:
                logger.error(f"재시작 마커 읽기 오류: {me}")

        return {
            'needs_restart': False,  # 이미 재시작 완료
            'applied_count': applied,
            'current_version': version
        }
    except Exception as e:
        return {'needs_restart': False, 'applied_count': 0, 'current_version': config.version}


@eel.expose
def check_for_updates(force: bool = False) -> Dict[str, Any]:
    """업데이트 확인"""
    try:
        result = update_manager.check_for_updates(force=force)
        return result
    except Exception as e:
        logger.error(f"업데이트 확인 오류: {e}")
        return {
            'update_available': False,
            'error': '요청 처리 중 오류가 발생했습니다.'
        }


@eel.expose
def download_and_apply_patches() -> Dict[str, Any]:
    """패치 ZIP 다운로드 및 적용"""
    try:
        result = update_manager.download_and_apply_patches()
        return result
    except Exception as e:
        logger.error(f"패치 적용 오류: {e}")
        return {
            'success': False,
            'message': '요청 처리 중 오류가 발생했습니다.'
        }


@eel.expose
def restart_app_after_update() -> Dict[str, Any]:
    """수동 패치 적용 후 앱을 재시작."""
    try:
        import os
        import subprocess
        import sys
        import time

        def _restart():
            time.sleep(0.8)
            subprocess.Popen([sys.executable] + sys.argv)
            os._exit(0)

        thread = threading.Thread(target=_restart, daemon=True)
        thread.start()
        return {'success': True, 'message': '프로그램을 재시작합니다.'}
    except Exception as e:
        logger.error(f"업데이트 후 재시작 오류: {e}")
        return {'success': False, 'message': '재시작 중 오류가 발생했습니다.'}


@eel.expose
def get_release_notes_for_version(version_tag: str) -> Dict[str, Any]:
    """특정 버전의 릴리즈 노트 조회"""
    try:
        notes = update_manager.get_release_notes(version_tag)
        if notes:
            return {
                'success': True,
                'release_notes': notes
            }
        else:
            return {
                'success': False,
                'message': '릴리즈 노트를 찾을 수 없습니다.'
            }
    except Exception as e:
        logger.error(f"릴리즈 노트 조회 오류: {e}")
        return {
            'success': False,
            'error': '요청 처리 중 오류가 발생했습니다.'
        }


def _normalize_patch_note_line(text: str) -> str:
    """릴리즈 노트 중복 제거용 정규화 문자열 생성."""
    if not text:
        return ''
    return ''.join(ch.casefold() for ch in str(text) if ch.isalnum())


def _extract_compact_patch_note_lines(body: str) -> List[str]:
    """릴리즈 본문에서 간단 패치 노트 줄만 추출."""
    if not body:
        return []

    summary_lines: List[str] = []
    seen: set[str] = set()

    for raw_line in re.split(r'[\r\n]+', str(body)):
        line = str(raw_line or '').strip()
        if not line:
            continue

        line = re.sub(r'^\s{0,3}(?:[-*+]|\d+\.)\s*', '', line)
        line = re.sub(r'^#{1,6}\s*', '', line)
        line = re.sub(r'`+', '', line).strip(' -:\t')
        if not line:
            continue
        if re.fullmatch(r'v?\d+(?:\.\d+){1,3}', line, re.IGNORECASE):
            continue

        normalized = _normalize_patch_note_line(line)
        if len(normalized) < 4 or normalized in seen:
            continue

        seen.add(normalized)
        summary_lines.append(line)

        if len(summary_lines) >= 8:
            break

    return summary_lines


@eel.expose
def get_compact_patch_notes(from_version: str, to_version: str = '') -> Dict[str, Any]:
    """버전 범위의 릴리즈 노트를 간략 요약으로 반환"""
    try:
        from packaging import version as pkg_version

        start_ver = str(from_version or '').strip().lstrip('v')
        end_ver = str(to_version or config.version or '').strip().lstrip('v')
        if not end_ver:
            return {'success': False, 'message': '현재 버전을 확인할 수 없습니다.'}

        try:
            end_parsed = pkg_version.parse(end_ver)
            start_parsed = pkg_version.parse(start_ver) if start_ver else None
        except Exception:
            return {'success': False, 'message': '버전 형식이 올바르지 않습니다.'}

        if start_parsed and start_parsed >= end_parsed:
            return {
                'success': False,
                'message': '패치 노트는 이전 버전보다 높은 버전으로 업데이트될 때만 표시됩니다.'
            }

        page = 1
        matched = []
        while page <= 5:
            releases = update_manager.get_all_releases(page=page, per_page=50) or []
            if not releases:
                break
            stop_scan = False
            for release in releases:
                tag_name = str(release.get('tag_name') or '').strip()
                if not tag_name:
                    continue
                try:
                    release_ver = pkg_version.parse(tag_name.lstrip('v'))
                except Exception:
                    continue
                if release_ver > end_parsed:
                    continue
                if start_parsed and release_ver <= start_parsed:
                    stop_scan = True
                    continue
                matched.append({
                    'tag': tag_name,
                    'version': str(release_ver),
                    'body': str(release.get('body') or '').strip(),
                })
            if stop_scan or len(releases) < 50:
                break
            page += 1

        matched.sort(key=lambda item: pkg_version.parse(item['version']))

        summary_lines = []
        summary_seen = set()
        for item in matched:
            for line in _extract_compact_patch_note_lines(item['body']):
                normalized = _normalize_patch_note_line(line)
                if not normalized or normalized in summary_seen:
                    continue
                summary_seen.add(normalized)
                summary_lines.append(line)
                if len(summary_lines) >= 8:
                    break
            if len(summary_lines) >= 8:
                break

        if not summary_lines:
            summary_lines = ['기능 개선 및 안정화 패치가 적용되었습니다.']

        return {
            'success': True,
            'fromVersion': start_ver,
            'toVersion': end_ver,
            'versions': [item['tag'] for item in matched],
            'notes': summary_lines,
        }
    except Exception as e:
        logger.error(f"패치 노트 요약 조회 오류: {e}")
        return {'success': False, 'message': '패치 노트를 불러오는 중 오류가 발생했습니다.'}


# ============================================================================
# 공휴일 데이터
# ============================================================================

@eel.expose
def get_holidays() -> Dict[str, str]:
    """공휴일 데이터 로드 (config/holidays.json)"""
    try:
        import json
        from pathlib import Path
        holidays_path = Path(__file__).parent.parent.parent / "config" / "holidays.json"
        with open(holidays_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        logger.error(f"공휴일 데이터 로드 오류: {e}")
        return {}


@eel.expose
def refresh_holidays(service_key: str, admin_id: str = '') -> Dict[str, Any]:
    """공공데이터포털 API로 공휴일 갱신 (관리자)"""
    try:
        user = auth_manager.get_user(admin_id) if admin_id else None
        if not user or user.get('role') != 'admin':
            return {'success': False, 'message': '관리자 권한 필요'}

        if not service_key or not service_key.strip():
            return {'success': False, 'message': '서비스키를 입력하세요.'}

        sk = service_key.strip()
        logger.info(f"공휴일 API 갱신 요청 by {admin_id or 'unknown'}")

        from ..utils.holiday_fetcher import update_holidays_file
        result = update_holidays_file(sk)

        # 서비스키 settings.json 저장
        config.set('holidays.data_go_kr_key', sk)
        config.save()

        counts = result.get('fetched_counts', {})
        summary = ', '.join(f"{y}년 {c}건" for y, c in sorted(counts.items()))
        msg = f"갱신 완료: {summary}"
        logger.info(msg)
        return {'success': True, 'message': msg}

    except Exception as e:
        logger.error(f"공휴일 갱신 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


# ============================================================================
# 텔레그램 알림 설정
# ============================================================================

@eel.expose
def generate_telegram_link_code(user_id: str) -> Dict[str, Any]:
    """텔레그램 연결 코드 생성"""
    try:
        code = auth_manager.generate_link_code(user_id)
        if code:
            bot_username = telegram_notifier.get_bot_username()
            # tg:// → PC Telegram Desktop 직접 실행 (os.startfile에 적합)
            deep_link = f'tg://resolve?domain={bot_username}&start={code}' if bot_username else ''
            # https:// → QR 코드(모바일 스캔)용
            web_link  = f'https://t.me/{bot_username}?start={code}' if bot_username else ''
            return {
                'success': True,
                'code': code,
                'botUsername': bot_username,
                'deepLink': deep_link,
                'webLink': web_link
            }
        return {'success': False, 'message': '코드 생성 실패'}
    except Exception as e:
        logger.error(f"텔레그램 코드 생성 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def unlink_telegram(user_id: str) -> Dict[str, Any]:
    """텔레그램 연결 해제"""
    try:
        success = auth_manager.unlink_telegram(user_id)
        return {'success': success}
    except Exception as e:
        logger.error(f"텔레그램 연결 해제 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def get_telegram_status(user_id: str) -> Dict[str, Any]:
    """텔레그램 연결 상태 조회"""
    try:
        return auth_manager.get_telegram_status(user_id)
    except Exception as e:
        logger.error(f"텔레그램 상태 조회 오류: {e}")
        return {'linked': False}


@eel.expose
def get_user_tray_mode(user_id: str) -> Dict[str, Any]:
    """사용자 트레이 모드 설정 조회"""
    try:
        return {'success': True, 'tray_mode': auth_manager.get_tray_mode(user_id)}
    except Exception as e:
        logger.error(f"트레이 설정 조회 오류: {e}")
        return {'success': False}


@eel.expose
def save_user_tray_mode(user_id: str, enabled: bool) -> Dict[str, Any]:
    """사용자 트레이 모드 설정 저장 + Python 전역 동기화"""
    try:
        auth_manager.update_tray_mode(user_id, bool(enabled))
        # Python main 모듈의 전역 변수 동기화
        try:
            import src.main as _main_mod
            _main_mod._tray_preference = bool(enabled)
        except Exception:
            pass
        return {'success': True}
    except Exception as e:
        logger.error(f"트레이 설정 저장 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def get_all_leave_monthly_report(year: int) -> Dict[str, Any]:
    """모든 직원의 연차 월별 현황 조회 (연차 월별 보고 탭)"""
    try:
        data = db.get_all_leave_monthly_report(int(year))
        return {'success': True, 'data': data}
    except Exception as e:
        logger.error(f"연차 월별 보고 조회 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def save_employee_leave_order(names_json: str) -> Dict[str, Any]:
    """직원 연차 보고 표시 순서 저장 (app_settings)"""
    try:
        with db.get_connection() as conn:
            conn.execute(
                "INSERT OR REPLACE INTO app_settings (key, value) VALUES ('employee_leave_order', ?)",
                (names_json,)
            )
        return {'success': True}
    except Exception as e:
        logger.error(f"직원 순서 저장 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def get_employee_leave_order() -> Dict[str, Any]:
    """저장된 직원 연차 보고 순서 반환"""
    try:
        with db.get_connection() as conn:
            row = conn.execute(
                "SELECT value FROM app_settings WHERE key = 'employee_leave_order'"
            ).fetchone()
        return {'success': True, 'order': row[0] if row else None}
    except Exception as e:
        logger.error(f"직원 순서 조회 오류: {e}")
        return {'success': True, 'order': None}


@eel.expose
def set_employee_leave_excluded(names_json: str) -> Dict[str, Any]:
    """직원 연차 보고 제외 목록 저장 (app_settings)"""
    try:
        with db.get_connection() as conn:
            conn.execute(
                "INSERT OR REPLACE INTO app_settings (key, value) VALUES ('employee_leave_excluded', ?)",
                (names_json,)
            )
        return {'success': True}
    except Exception as e:
        logger.error(f"직원 제외 목록 저장 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def get_employee_leave_excluded() -> Dict[str, Any]:
    """제외된 직원 이름 목록 반환 (JSON 배열 문자열)"""
    try:
        with db.get_connection() as conn:
            row = conn.execute(
                "SELECT value FROM app_settings WHERE key = 'employee_leave_excluded'"
            ).fetchone()
        return {'success': True, 'excluded': row[0] if row else '[]'}
    except Exception as e:
        logger.error(f"직원 제외 목록 조회 오류: {e}")
        return {'success': True, 'excluded': '[]'}


@eel.expose
def set_leave_report_edit(user_id: str, enabled: bool, admin_id: str) -> Dict[str, Any]:
    """연차 월별 보고 편집 권한 설정 (관리자 전용)"""
    try:
        with auth_manager.get_connection() as conn:
            row = conn.execute(
                'SELECT role FROM auth_users WHERE user_id = ?', (admin_id,)
            ).fetchone()
        if not row or row['role'] != 'admin':
            return {'success': False, 'message': '관리자 권한 필요'}
        ok = auth_manager.set_leave_report_edit(user_id, bool(enabled))
        if not ok:
            return {'success': False, 'message': 'DB 업데이트 실패 — 앱을 재시작하면 자동 수정됩니다.'}
        return {'success': True}
    except Exception as e:
        logger.error(f"leave_report_edit 설정 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def admin_set_write_permission(user_id: str, enabled: bool, admin_id: str) -> Dict[str, Any]:
    """일일 작업 쓰기 권한 부여/해제 (관리자 전용)"""
    try:
        with auth_manager.get_connection() as conn:
            row = conn.execute(
                'SELECT role FROM auth_users WHERE user_id = ?', (admin_id,)
            ).fetchone()
        if not row or row['role'] != 'admin':
            return {'success': False, 'message': '관리자 권한 필요'}
        ok = auth_manager.set_can_write(user_id, bool(enabled))
        if not ok:
            return {'success': False, 'message': 'DB 업데이트 실패 — 앱을 재시작하면 자동 수정됩니다.'}
        action = '쓰기 권한 부여' if enabled else '쓰기 권한 해제'
        db.add_activity_log(admin_id, 'set_write_permission', f'{user_id} → {action}')
        return {'success': True}
    except Exception as e:
        logger.error(f"쓰기 권한 설정 오류: {e}")
        return {'success': False, 'message': '쓰기 권한 설정 중 오류가 발생했습니다.'}


@eel.expose
def admin_set_erp_input(user_id: str, enabled: bool, admin_id: str) -> Dict[str, Any]:
    """ERP 입력 자동화 권한 부여/해제 (관리자 전용)"""
    try:
        with auth_manager.get_connection() as conn:
            row = conn.execute(
                'SELECT role FROM auth_users WHERE user_id = ?', (admin_id,)
            ).fetchone()
        if not row or row['role'] != 'admin':
            return {'success': False, 'message': '관리자 권한 필요'}
        ok = auth_manager.set_erp_input(user_id, bool(enabled))
        if not ok:
            return {'success': False, 'message': 'DB 업데이트 실패'}
        action = 'ERP입력 권한 부여' if enabled else 'ERP입력 권한 해제'
        db.add_activity_log(admin_id, 'set_erp_input', f'{user_id} → {action}')
        return {'success': True}
    except Exception as e:
        logger.error(f"ERP 입력 권한 설정 오류: {e}")
        return {'success': False, 'message': 'ERP 입력 권한 설정 중 오류가 발생했습니다.'}


def _check_erp_permission(user_id: str) -> Optional[str]:
    """ERP 권한 검증 헬퍼.
    통과하면 None 반환, 실패하면 오류 메시지 문자열 반환.
    erp_input 컬럼이 없는 구버전 DB도 안전하게 처리."""
    try:
        with auth_manager.get_connection() as conn:
            row = conn.execute(
                'SELECT * FROM auth_users WHERE user_id = ? AND status = ?',
                (user_id, 'active')
            ).fetchone()
        if not row:
            return '유효하지 않은 사용자입니다.'
        user = dict(row)
        if user.get('role') == 'admin':
            return None  # admin 무조건 허용
        if not bool(user.get('erp_input', 0)):
            return 'ERP 입력 권한이 없습니다. 관리자에게 권한 부여를 요청하세요.'
        return None
    except Exception as e:
        logger.error(f"ERP 권한 검증 오류: {e}")
        return '권한 검증 중 오류가 발생했습니다.'


@eel.expose
def get_records_for_erp(start_date: str, end_date: str, user_id: str) -> Dict[str, Any]:
    """날짜 범위 내 작업 레코드를 날짜별로 그룹화하여 반환 (ERP 입력용)"""
    try:
        err = _check_erp_permission(user_id)
        if err:
            return {'success': False, 'message': err}

        # 날짜 범위 검증
        if not start_date or not end_date or len(start_date) != 10 or len(end_date) != 10:
            return {'success': False, 'message': '날짜 형식이 올바르지 않습니다.'}

        # 작업 레코드 조회 (빈 레코드 제외)
        with db.get_connection() as conn:
            rows = conn.execute('''
                SELECT date, record_number, contract_number, work_content, leader, teammates
                FROM work_records
                WHERE date >= ? AND date <= ?
                  AND (contract_number != '' OR work_content != '')
                ORDER BY date ASC, record_number ASC
            ''', (start_date, end_date)).fetchall()

        # 날짜별 그룹핑
        from collections import defaultdict
        grouped: dict = defaultdict(list)
        for row in rows:
            grouped[row['date']].append({
                'recordNumber': row['record_number'],
                'contractNumber': row['contract_number'] or '',
                'workContent': row['work_content'] or '',
                'leader': row['leader'] or '',
                'teammates': row['teammates'] or '',
            })

        dates = [{'date': d, 'records': grouped[d]} for d in sorted(grouped.keys())]
        return {'success': True, 'dates': dates}
    except Exception as e:
        logger.error(f"ERP 레코드 조회 오류: {e}")
        return {'success': False, 'message': '레코드 조회 중 오류가 발생했습니다.'}


@eel.expose
def start_erp_macro(records_json: str, user_id: str) -> Dict[str, Any]:
    """백그라운드 스레드에서 ERP 매크로 시작"""
    try:
        import json as _json
        err = _check_erp_permission(user_id)
        if err:
            return {'success': False, 'message': err}

        if erp_macro.get_status()['running']:
            return {'success': False, 'message': '이미 실행 중입니다.'}

        dates_records = _json.loads(records_json)
        _start_tracked_thread(target=erp_macro.run, args=(dates_records,))
        return {'success': True, 'message': '매크로가 시작되었습니다.'}
    except Exception as e:
        logger.error(f"ERP 매크로 시작 오류: {e}")
        return {'success': False, 'message': 'ERP 매크로 시작 중 오류가 발생했습니다.'}


@eel.expose
def stop_erp_macro(user_id: str) -> Dict[str, Any]:
    """실행 중인 ERP 매크로 중단"""
    try:
        err = _check_erp_permission(user_id)
        if err:
            return {'success': False, 'message': err}

        erp_macro.stop()
        return {'success': True, 'message': '매크로 중단 요청이 전송되었습니다.'}
    except Exception as e:
        logger.error(f"ERP 매크로 중단 오류: {e}")
        return {'success': False, 'message': 'ERP 매크로 중단 중 오류가 발생했습니다.'}


@eel.expose
def install_erp_deps(user_id: str) -> Dict[str, Any]:
    """pyautogui, pywin32, pyperclip 자동 설치 (ERP 권한 필요)"""
    try:
        err = _check_erp_permission(user_id)
        if err:
            return {'success': False, 'message': err}
        import subprocess
        import sys
        result = subprocess.run(
            [sys.executable, '-m', 'pip', 'install', 'pyautogui', 'pywin32', 'pyperclip'],
            capture_output=True, encoding='utf-8', errors='replace', timeout=120
        )
        if result.returncode == 0:
            return {'success': True, 'message': '설치 완료. 입력 시작 버튼을 다시 눌러주세요.'}
        stderr_lower = (result.stderr or '').lower()
        if 'permission' in stderr_lower or 'access' in stderr_lower or 'winerror 5' in stderr_lower:
            msg = '⚠️ 권한 오류: 앱을 관리자 모드로 실행 후 다시 시도해주세요. (우클릭 → 관리자 권한으로 실행)'
        else:
            msg = '설치 실패. 관리자 모드로 실행 후 재시도하거나 관리자에게 문의하세요.'
        return {'success': False, 'message': msg}
    except Exception:
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


def _find_browser() -> str:
    """Chrome 또는 Edge 실행 파일 경로 반환"""
    import os as _os
    candidates = [
        r'C:\Program Files\Google\Chrome\Application\chrome.exe',
        r'C:\Program Files (x86)\Google\Chrome\Application\chrome.exe',
        r'C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe',
        r'C:\Program Files\Microsoft\Edge\Application\msedge.exe',
    ]
    for p in candidates:
        if _os.path.exists(p):
            return p
    return ''


@eel.expose
def open_erp_input_window(dates_json: str, user_id: str = '') -> Dict[str, Any]:
    """ERP 입력 팝업 창 열기 (Chrome --app 모드)"""
    global _erp_popup_context
    try:
        import json as _json, subprocess as _sp
        err = _check_erp_permission(user_id)
        if err:
            return {'success': False, 'message': err}

        dates_records = _json.loads(dates_json)
        if not dates_records:
            return {'success': False, 'message': '입력할 레코드가 없습니다.'}

        _erp_popup_context = {'dates_records': dates_records, 'user_id': user_id}

        # 이미 팝업이 실행 중이면 새 창 열지 않음 (중복 방지)
        global _erp_popup_proc
        if _erp_popup_proc is not None and _erp_popup_proc.poll() is None:
            return {'success': True, 'message': '이미 ERP 입력 창이 열려 있습니다.'}

        browser = _find_browser()
        if not browser:
            return {'success': False, 'message': 'Chrome 또는 Edge를 찾을 수 없습니다.'}

        _erp_popup_proc = _sp.Popen([
            browser,
            '--app=http://localhost:8686/erp_popup.html',
            '--window-size=440,620',
            '--window-position=80,120',
        ])
        return {'success': True, 'message': 'ERP 입력 창이 열립니다.'}
    except Exception as e:
        logger.error(f"ERP 팝업 창 실행 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def get_erp_popup_context() -> Dict[str, Any]:
    """팝업 창 로드 시 날짜/레코드 정보 반환"""
    return {'success': True, **_erp_popup_context}


@eel.expose
def start_erp_macro_inline(user_id: str = '') -> Dict[str, Any]:
    """팝업 [입력 시작] 클릭 → 캐시된 dates_records로 매크로 실행"""
    try:
        err = _check_erp_permission(user_id)
        if err:
            return {'success': False, 'message': err}
        records = _erp_popup_context.get('dates_records', [])
        if not records:
            return {'success': False, 'message': '입력할 레코드가 없습니다.'}
        if erp_macro.get_status()['running']:
            return {'success': False, 'message': '이미 실행 중입니다.'}
        erp_macro._stop_flag = False
        threading.Thread(
            target=erp_macro.run,
            args=(records,),
            kwargs={'save_daily': False},
            daemon=True
        ).start()
        return {'success': True, 'message': '매크로 시작'}
    except Exception as e:
        logger.error(f"ERP 매크로 인라인 시작 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def get_window_list() -> Dict[str, Any]:
    """현재 열린 가시 창 목록 반환 [{hwnd, title}]"""
    try:
        import win32gui as _wg
        windows = []
        def _cb(hwnd, _):
            if _wg.IsWindowVisible(hwnd):
                title = _wg.GetWindowText(hwnd)
                if title.strip():
                    windows.append({'hwnd': hwnd, 'title': title})
        _wg.EnumWindows(_cb, None)
        return {'success': True, 'windows': windows}
    except Exception as e:
        logger.error(f"창 목록 조회 오류: {e}")
        return {'success': False, 'windows': [], 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def set_erp_target_hwnd(hwnd: int, user_id: str = '') -> Dict[str, Any]:
    """사용자가 선택한 창 HWND를 ERPMacro에 지정"""
    try:
        err = _check_erp_permission(user_id)
        if err:
            return {'success': False, 'message': err}
        erp_macro.forced_hwnd = int(hwnd)
        return {'success': True}
    except Exception as e:
        logger.error(f"ERP HWND 지정 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def get_erp_macro_status(user_id: str) -> Dict[str, Any]:
    """매크로 진행 상태 조회"""
    try:
        err = _check_erp_permission(user_id)
        if err:
            return {'success': False, 'message': err}

        status = erp_macro.get_status()
        return {'success': True, **status}
    except Exception as e:
        logger.error(f"ERP 상태 조회 오류: {e}")
        return {'success': False, 'message': '상태 조회 중 오류가 발생했습니다.'}


@eel.expose
def diagnose_erp_controls(user_id: str = '') -> Dict[str, Any]:
    """ERP 창의 자식 컨트롤 목록 반환 (달력 컨트롤 탐색 진단용)"""
    try:
        err = _check_erp_permission(user_id)
        if err:
            return {'success': False, 'message': err}

        try:
            import win32gui
        except ImportError:
            return {'success': False, 'message': 'pywin32 미설치. pip install pywin32'}

        hwnd = erp_macro._find_erp_window(win32gui)
        if not hwnd:
            visible: list = []

            def _cb(h, _):
                t = win32gui.GetWindowText(h)
                if t and win32gui.IsWindowVisible(h):
                    visible.append(t[:60])

            win32gui.EnumWindows(_cb, None)
            return {
                'success': False,
                'message': 'ERP 창을 찾을 수 없습니다. 선진종합시스템을 먼저 실행하세요.',
                'open_windows': visible[:10],
            }

        controls = erp_macro._discover_controls(hwnd, win32gui)

        from src.utils.erp_macro import CALENDAR_CLASSES
        cal_set = {c.lower() for c in CALENDAR_CLASSES}
        calendar_candidates = [
            c for c in controls if c['cls'].lower() in cal_set
        ]

        safe_controls = [
            {'hwnd': c['hwnd'], 'cls': c['cls'],
             'text': c['text'][:50], 'rect': list(c['rect'])}
            for c in controls
        ]

        return {
            'success': True,
            'erp_hwnd': hwnd,
            'erp_title': win32gui.GetWindowText(hwnd),
            'controls': safe_controls,
            'calendar_candidates': [
                {'hwnd': c['hwnd'], 'cls': c['cls'],
                 'text': c['text'][:50], 'rect': list(c['rect'])}
                for c in calendar_candidates
            ],
        }
    except Exception as e:
        logger.error(f"ERP 컨트롤 진단 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def get_telegram_bot_enabled() -> Dict[str, Any]:
    """텔레그램 봇 활성화 상태 조회"""
    try:
        # config 우선, 없으면 DB 폴백 (PC2처럼 settings.json에 토큰 없는 환경 대응)
        bot_token = config.get('telegram.bot_token', '')
        if not bot_token:
            bot_token = db.get_setting('telegram.bot_token', '') or ''
            if bot_token:
                # DB에서 로드한 값을 config에도 반영 (이후 호출 빠르게)
                config.set('telegram.bot_token', bot_token)
                logger.info("get_telegram_bot_enabled: DB app_settings에서 토큰 로드")
        enabled = config.get('telegram.enabled', False)
        return {
            'success': True,
            'enabled': enabled,
            'hasToken': bool(bot_token),
            'botUsername': telegram_notifier.get_bot_username() if enabled else ''
        }
    except Exception as e:
        logger.error(f"텔레그램 봇 상태 조회 오류: {e}")
        return {'success': False, 'enabled': False, 'hasToken': False}


@eel.expose
def admin_save_telegram_settings(bot_token: str, enabled: bool, admin_id: str) -> Dict[str, Any]:
    """텔레그램 봇 설정 저장 (관리자)"""
    try:
        import requests as _requests
        # 토큰 유효성 검사 (비어있지 않을 경우)
        if bot_token and bot_token.strip():
            try:
                r = _requests.get(
                    f'https://api.telegram.org/bot{bot_token.strip()}/getMe',
                    timeout=5
                )
                data = r.json()
                if not r.ok or not data.get('ok'):
                    err_desc = data.get('description', '알 수 없는 오류')
                    logger.warning(f"텔레그램 토큰 유효성 검사 실패: {err_desc}")
                    return {
                        'success': False,
                        'message': f'유효하지 않은 봇 토큰입니다: {err_desc}\nBotFather에서 토큰을 확인하세요.'
                    }
                bot_name = data.get('result', {}).get('username', '')
                logger.info(f"텔레그램 토큰 유효: @{bot_name}")
            except _requests.exceptions.Timeout:
                logger.warning("텔레그램 토큰 검사 타임아웃 (네트워크 불안정) - 저장 계속 진행")
            except Exception as ve:
                logger.warning(f"텔레그램 토큰 검사 중 예외 (저장 계속): {ve}")

        config.set('telegram.bot_token', bot_token)
        config.set('telegram.enabled', enabled)
        config.save()
        # DB app_settings에도 동기화 (다른 PC에서 자동 로드용)
        db.set_setting('telegram.bot_token', bot_token)
        db.set_setting('telegram.enabled', 'true' if enabled else 'false')
        telegram_notifier.reconfigure(bot_token, enabled)
        logger.info(f"텔레그램 설정 변경: enabled={enabled}, by={admin_id}")
        return {'success': True, 'message': '텔레그램 설정이 저장되었습니다.'}
    except Exception as e:
        logger.error(f"텔레그램 설정 저장 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}


@eel.expose
def save_auto_capture_image(base64_data: str, date_str: str) -> Dict[str, Any]:
    """자동 캡처 이미지를 DB 폴더에 저장 (스케줄러 17:00 호출)"""
    try:
        import base64 as b64
        from datetime import date as _date
        db_folder = config.db_path.parent
        days_kr = ['월', '화', '수', '목', '금', '토', '일']
        d = _date.fromisoformat(date_str)
        day_name = days_kr[d.weekday()]
        filename = f"{d.year}.{d.month:02d}.{d.day:02d}.{day_name}(주간).png"
        img_data = b64.b64decode(base64_data)
        filepath = db_folder / filename
        with open(filepath, 'wb') as f:
            f.write(img_data)
        logger.info(f"자동 캡처 저장: {filepath}")
        return {'success': True, 'path': str(filepath)}
    except Exception as e:
        logger.error(f"자동 캡처 저장 오류: {e}")
        return {'success': False, 'message': '요청 처리 중 오류가 발생했습니다.'}
