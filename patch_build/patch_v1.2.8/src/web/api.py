# src/web/api.py - 웹 API (Python ↔ JavaScript)

import eel
import threading
from datetime import datetime
from typing import List, Dict, Any
from ..business.work_record_service import work_record_service
from ..business.calculations import separate_workers
from ..database.db_manager import db
from ..database.auth_manager import auth_manager
from ..sync.cloud_sync import cloud_sync
from ..utils.logger import logger
from ..utils.config import config
from ..utils.settings_manager import settings_manager
from ..utils.path_manager import path_manager
from ..utils.update_manager import update_manager
from ..utils.telegram_notifier import telegram_notifier


# ============================================================================
# 연결 확인 (스플래시 화면용)
# ============================================================================

@eel.expose
def ping() -> bool:
    """Python 백엔드 연결 확인용 (스플래시 → 로그인 전환 트리거)"""
    return True


# ============================================================================
# 인증 관리
# ============================================================================

@eel.expose
def authenticate(user_id: str, password: str) -> Dict[str, Any]:
    """사용자 인증"""
    try:
        result = auth_manager.authenticate(user_id, password)
        
        if result is None:
            return {'success': False, 'message': '아이디 또는 비밀번호가 일치하지 않습니다.'}

        if 'error' in result:
            err = result['error']
            if err == 'USER_NOT_FOUND':
                return {'success': False, 'message': '없는 사용자입니다.'}
            elif err == 'WRONG_PASSWORD':
                return {'success': False, 'message': '비밀번호가 일치하지 않습니다.'}
            else:
                return {'success': False, 'message': err}
        
        # 로그인 성공
        return {
            'success': True,
            'user': {
                'user_id': result['user_id'],
                'full_name': result['full_name'],
                'role': result['role']
            }
        }
    except Exception as e:
        logger.error(f"인증 오류: {e}")
        return {'success': False, 'message': f'인증 중 오류가 발생했습니다: {str(e)}'}


@eel.expose
def register_user(user_id: str, password: str, full_name: str) -> Dict[str, Any]:
    """사용자 등록 요청"""
    try:
        success = auth_manager.register_request(user_id, password, full_name)
        
        if success:
            return {
                'success': True,
                'message': '등록 요청이 전송되었습니다. 관리자 승인 후 이용 가능합니다.'
            }
        else:
            return {
                'success': False,
                'message': '이미 존재하는 아이디입니다.'
            }
    except Exception as e:
        logger.error(f"등록 요청 오류: {e}")
        return {'success': False, 'message': f'등록 요청 중 오류가 발생했습니다: {str(e)}'}


# ============================================================================
# 관리자 기능
# ============================================================================

@eel.expose
def admin_get_all_users() -> List[Dict[str, Any]]:
    """모든 사용자 조회 (관리자)"""
    try:
        users = auth_manager.get_all_users()
        return users
    except Exception as e:
        logger.error(f"사용자 목록 조회 오류: {e}")
        return []


@eel.expose
def admin_get_pending_requests() -> List[Dict[str, Any]]:
    """승인 대기 요청 조회 (관리자)"""
    try:
        requests = auth_manager.get_pending_requests()
        return requests
    except Exception as e:
        logger.error(f"대기 요청 조회 오류: {e}")
        return []


@eel.expose
def admin_approve_user(user_id: str, admin_id: str) -> Dict[str, Any]:
    """사용자 승인 (관리자)"""
    try:
        success = auth_manager.approve_user(user_id, admin_id)
        return {
            'success': success,
            'message': '사용자가 승인되었습니다.' if success else '승인 실패'
        }
    except Exception as e:
        logger.error(f"사용자 승인 오류: {e}")
        return {'success': False, 'message': f'오류: {str(e)}'}


@eel.expose
def admin_reject_user(user_id: str, admin_id: str, note: str = '') -> Dict[str, Any]:
    """사용자 거부 (관리자)"""
    try:
        success = auth_manager.reject_user(user_id, admin_id, note)
        return {
            'success': success,
            'message': '사용자가 거부되었습니다.' if success else '거부 실패'
        }
    except Exception as e:
        logger.error(f"사용자 거부 오류: {e}")
        return {'success': False, 'message': f'오류: {str(e)}'}


@eel.expose
def admin_delete_user(user_id: str, admin_id: str) -> Dict[str, Any]:
    """사용자 삭제 (관리자)"""
    try:
        success = auth_manager.delete_user(user_id, admin_id)
        return {
            'success': success,
            'message': '사용자가 삭제되었습니다.' if success else '삭제 실패'
        }
    except Exception as e:
        logger.error(f"사용자 삭제 오류: {e}")
        return {'success': False, 'message': f'오류: {str(e)}'}


@eel.expose
def admin_update_user_status(user_id: str, status: str, admin_id: str) -> Dict[str, Any]:
    """사용자 상태 변경 (관리자)"""
    try:
        success = auth_manager.update_user_status(user_id, status, admin_id)
        return {
            'success': success,
            'message': '사용자 상태가 변경되었습니다.' if success else '상태 변경 실패'
        }
    except Exception as e:
        logger.error(f"사용자 상태 변경 오류: {e}")
        return {'success': False, 'message': f'오류: {str(e)}'}


@eel.expose
def select_folder_path() -> Dict[str, Any]:
    """폴더 선택 다이얼로그 (Windows Shell API - Embedded Python 호환)"""
    try:
        import ctypes
        import ctypes.wintypes

        # Embedded Python에는 Tkinter가 없으므로 Windows Shell32 API 직접 사용
        BIF_RETURNONLYFSDIRS = 0x0001
        BIF_NEWDIALOGSTYLE   = 0x0040  # 새 스타일 폴더 선택창

        class BROWSEINFOW(ctypes.Structure):
            _fields_ = [
                ('hwndOwner',      ctypes.wintypes.HWND),
                ('pidlRoot',       ctypes.c_void_p),
                ('pszDisplayName', ctypes.c_wchar_p),
                ('lpszTitle',      ctypes.c_wchar_p),
                ('ulFlags',        ctypes.wintypes.UINT),
                ('lpfn',           ctypes.c_void_p),
                ('lParam',         ctypes.c_long),
                ('iImage',         ctypes.c_int),
            ]

        shell32 = ctypes.windll.shell32
        ole32   = ctypes.windll.ole32
        ole32.CoInitialize(None)

        display_buf = ctypes.create_unicode_buffer(260)
        bi = BROWSEINFOW()
        bi.hwndOwner      = None
        bi.pidlRoot       = None
        bi.pszDisplayName = display_buf
        bi.lpszTitle      = '폴더 선택'
        bi.ulFlags        = BIF_RETURNONLYFSDIRS | BIF_NEWDIALOGSTYLE
        bi.lpfn           = None
        bi.lParam         = 0

        pidl = shell32.SHBrowseForFolderW(ctypes.byref(bi))
        if pidl:
            path_buf = ctypes.create_unicode_buffer(260)
            shell32.SHGetPathFromIDListW(pidl, path_buf)
            ole32.CoTaskMemFree(pidl)
            return {'success': True, 'path': path_buf.value}
        else:
            return {'success': False, 'path': None}

    except Exception as e:
        logger.error(f"폴더 선택 오류: {e}")
        return {'success': False, 'error': str(e)}


@eel.expose
def admin_get_paths() -> Dict[str, Any]:
    """경로 조회 (관리자)"""
    try:
        settings = settings_manager.get_current_settings()
        return {
            'success': True,
            'paths': {
                'local_db_path': settings.get('local_db_path', ''),
                'cloud_sync_path': settings.get('cloud_sync_path', ''),
                'backup_path': settings.get('backup_path', '')
            }
        }
    except Exception as e:
        logger.error(f"경로 조회 오류: {e}")
        return {'success': False, 'paths': {}}


@eel.expose
def admin_get_settings() -> Dict[str, Any]:
    """설정 조회 (관리자)"""
    try:
        return settings_manager.get_current_settings()
    except Exception as e:
        logger.error(f"설정 조회 오류: {e}")
        return {}


@eel.expose
def admin_update_local_db_path(new_path: str, admin_id: str) -> Dict[str, Any]:
    """로컬 DB 경로 변경 (관리자)"""
    try:
        logger.info(f"로컬 DB 경로 변경 요청: {new_path} by {admin_id}")
        return settings_manager.update_local_db_path(new_path)
    except Exception as e:
        logger.error(f"경로 변경 오류: {e}")
        return {'success': False, 'message': f'오류: {str(e)}'}


@eel.expose
def admin_update_cloud_path(new_path: str, admin_id: str) -> Dict[str, Any]:
    """클라우드 경로 변경 (관리자)"""
    try:
        logger.info(f"클라우드 경로 변경 요청: {new_path} by {admin_id}")
        result = settings_manager.update_cloud_path(new_path)
        if result.get('success'):
            # settings_manager는 JSON 파일만 업데이트 → 메모리 config 직접 갱신
            config.set('database.cloud_path', new_path)
            config.set('database.cloud_sync_enabled', True)
            # 싱글톤 cloud_sync 인스턴스 즉시 재초기화 (재시작 불필요)
            cloud_sync.cloud_folder = cloud_sync._get_cloud_folder()
            cloud_sync.enabled = config.cloud_sync_enabled
            if cloud_sync.cloud_folder:
                logger.info(f"클라우드 폴더 재초기화 완료: {cloud_sync.cloud_folder}")
                result['message'] = f'클라우드 경로가 설정되었습니다: {cloud_sync.cloud_folder}'
            else:
                result['message'] = '경로가 저장되었으나 폴더를 찾을 수 없습니다. 경로를 확인하세요.'
        return result
    except Exception as e:
        logger.error(f"경로 변경 오류: {e}")
        return {'success': False, 'message': f'오류: {str(e)}'}


@eel.expose
def admin_update_backup_path(new_path: str, admin_id: str) -> Dict[str, Any]:
    """백업 경로 변경 (관리자)"""
    try:
        logger.info(f"백업 경로 변경 요청: {new_path} by {admin_id}")
        return settings_manager.update_backup_path(new_path)
    except Exception as e:
        logger.error(f"경로 변경 오류: {e}")
        return {'success': False, 'message': f'오류: {str(e)}'}


@eel.expose
def admin_create_backup(admin_id: str) -> Dict[str, Any]:
    """수동 백업 생성 (관리자)"""
    try:
        logger.info(f"수동 백업 생성 요청 by {admin_id}")
        return settings_manager.create_backup()
    except Exception as e:
        logger.error(f"백업 생성 오류: {e}")
        return {'success': False, 'message': f'오류: {str(e)}'}


# ============================================================================
# 사용자 관리 (기존 호환성 유지)
# ============================================================================

@eel.expose
def login_user(username: str) -> bool:
    """사용자 로그인 (레거시)"""
    try:
        success = db.add_or_update_user(username)
        if success:
            logger.info(f"사용자 로그인: {username}")
            db.add_activity_log(username, 'login', '', '로그인')
        return success
    except Exception as e:
        logger.error(f"로그인 오류: {e}")
        return False


@eel.expose
def get_recent_users() -> List[Dict[str, Any]]:
    """최근 사용자 목록"""
    try:
        users = db.get_all_users()
        return [u.to_dict() for u in users[:10]]
    except Exception as e:
        logger.error(f"사용자 목록 조회 오류: {e}")
        return []


# ============================================================================
# 작업 레코드 관리
# ============================================================================

@eel.expose
def load_work_records(date: str) -> List[Dict[str, Any]]:
    """작업 레코드 로드"""
    try:
        logger.info(f"작업 레코드 로드 요청: {date}")
        records = work_record_service.get_records_for_date(date)
        return records
    except Exception as e:
        logger.error(f"작업 레코드 로드 오류: {e}")
        return []


@eel.expose
def save_work_records(date: str, records: List[Dict[str, Any]], username: str) -> Dict[str, Any]:
    """작업 레코드 저장"""
    try:
        logger.info(f"작업 레코드 저장 요청: {date}, 사용자: {username}")
        success = work_record_service.save_records_for_date(date, records, username)
        
        # 저장 성공 시 클라우드 동기화 (백그라운드 스레드 - UI 블로킹 방지)
        if success and cloud_sync.enabled:
            t = threading.Thread(target=cloud_sync.sync_to_cloud, daemon=True)
            t.start()
        
        return {'success': success, 'message': '저장되었습니다.' if success else '저장 실패'}
    except Exception as e:
        logger.error(f"작업 레코드 저장 오류: {e}")
        return {'success': False, 'message': f'오류: {str(e)}'}


@eel.expose
def clear_all_records(admin_id: str = '') -> Dict[str, Any]:
    """전체 작업 레코드 삭제 (관리자 전용)"""
    try:
        # 관리자 권한 확인
        if not admin_id:
            logger.warning("clear_all_records: admin_id 없음 - 권한 거부")
            return {'success': False, 'message': '관리자 권한이 필요합니다.'}

        user = auth_manager.get_user(admin_id)
        if not user or user.get('role') != 'admin':
            logger.warning(f"clear_all_records: 권한 없음 - user_id={admin_id}")
            return {'success': False, 'message': '관리자만 실행할 수 있습니다.'}

        success = db.clear_all_work_records()
        if success:
            db.add_activity_log(admin_id, 'clear_all', '', '전체 작업 레코드 삭제')
            logger.info(f"전체 작업 레코드 삭제 완료 by {admin_id}")
            return {'success': True, 'message': '전체 작업 레코드가 삭제되었습니다.'}
        else:
            return {'success': False, 'message': '삭제 실패'}
    except Exception as e:
        logger.error(f"전체 레코드 삭제 오류: {e}")
        return {'success': False, 'message': f'오류: {str(e)}'}


@eel.expose
def load_yesterday_records(current_date: str) -> List[Dict[str, Any]]:
    """어제 작업 불러오기"""
    try:
        logger.info(f"어제 작업 불러오기: {current_date}")
        records = work_record_service.get_yesterday_records(current_date)
        return records
    except Exception as e:
        logger.error(f"어제 작업 불러오기 오류: {e}")
        return []


@eel.expose
def get_date_list(start_date: str = None, end_date: str = None) -> List[str]:
    """레코드가 있는 날짜 목록"""
    try:
        dates = work_record_service.get_date_list(start_date, end_date)
        return dates
    except Exception as e:
        logger.error(f"날짜 목록 조회 오류: {e}")
        return []


# ============================================================================
# 보고서
# ============================================================================

@eel.expose
def generate_report(date: str, username: str) -> Dict[str, Any]:
    """보고서 생성"""
    try:
        report = work_record_service.generate_report(date, username)
        return report
    except Exception as e:
        logger.error(f"보고서 생성 오류: {e}")
        return {}


@eel.expose
def get_project_start_date_by_contract(contract_number: str) -> str:
    """계약번호 기준 공사 시작일 조회"""
    try:
        if not contract_number or contract_number.strip() == '':
            return ''
        
        # 해당 계약번호의 최초 작업일 조회
        query = '''
            SELECT MIN(date) as start_date, COUNT(*) as count
            FROM work_records
            WHERE contract_number = ?
            AND contract_number IS NOT NULL
            AND contract_number != ''
        '''
        
        logger.info(f"계약번호 조회: {contract_number}")
        
        result = db.execute_query(query, (contract_number,))
        logger.info(f"쿼리 결과: {result}")
        
        if result and result[0][0]:
            from datetime import datetime
            start_date_str = result[0][0]  # YYYY-MM-DD 형식
            count = result[0][1] if len(result[0]) > 1 else 0
            logger.info(f"찾은 데이터: 시작일={start_date_str}, 건수={count}")
            
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            # M/D 형식으로 반환
            result_str = f"{start_date.month}/{start_date.day}"
            logger.info(f"반환값: {result_str}")
            return result_str
        
        logger.warning(f"계약번호 {contract_number}에 대한 데이터를 찾을 수 없음")
        return ''
    except Exception as e:
        logger.error(f"계약번호 기준 시작일 조회 오류: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return ''


@eel.expose
def get_project_start_date(ship_name: str) -> str:
    """선박별 공사 시작일 조회 (최초 작업일)"""
    try:
        if not ship_name or ship_name.strip() == '':
            return ''
        
        # 해당 선박의 모든 작업 기록 조회
        query = '''
            SELECT MIN(date) as start_date
            FROM work_records
            WHERE ship_name = ?
            AND ship_name != ''
        '''
        
        result = db.execute_query(query, (ship_name,))
        if result and result[0][0]:
            from datetime import datetime
            start_date_str = result[0][0]  # YYYY-MM-DD 형식
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            # M/D 형식으로 반환
            return f"{start_date.month}/{start_date.day}"
        return ''
    except Exception as e:
        logger.error(f"공사 시작일 조회 오류: {e}")
        return ''


@eel.expose
def search_records_by_contract(contract_number: str) -> List[Dict[str, Any]]:
    """계약번호로 작업 내역 조회"""
    try:
        if not contract_number or contract_number.strip() == '':
            return []

        query = '''
            SELECT date, company, ship_name, engine_model, work_content,
                   leader, manpower, teammates
            FROM work_records
            WHERE contract_number = ?
            AND contract_number IS NOT NULL
            AND contract_number != ''
            ORDER BY date, record_number
        '''

        logger.info(f"계약번호 검색: {contract_number}")
        results = db.execute_query(query, (contract_number.strip().upper(),))
        logger.info(f"검색 결과: {len(results) if results else 0}건")

        records = []
        for row in results:
            records.append({
                'date': row[0] or '',
                'company': row[1] or '',
                'ship_name': row[2] or '',
                'engine_model': row[3] or '',
                'work_content': row[4] or '',
                'leader': row[5] or '',
                'manpower': float(row[6]) if row[6] else 0.0,
                'teammates': row[7] or ''
            })

        return records
    except Exception as e:
        logger.error(f"계약번호 검색 오류: {e}")
        return []


@eel.expose
def get_latest_contract_number() -> str:
    """DB에서 가장 최근 날짜의 계약번호 반환"""
    try:
        results = db.execute_query(
            '''SELECT contract_number FROM work_records
               WHERE contract_number != '' AND contract_number IS NOT NULL
               ORDER BY date DESC, id DESC LIMIT 1'''
        )
        if results and results[0][0]:
            return results[0][0].strip().upper()
        return ''
    except Exception as e:
        logger.error(f"최신 계약번호 조회 오류: {e}")
        return ''


@eel.expose
def load_vacation_data(date: str) -> Dict[str, str]:
    """날짜별 휴가자 현황 로드"""
    try:
        return db.load_vacation_records(date)
    except Exception as e:
        logger.error(f"휴가자 현황 로드 오류: {e}")
        return {'연차': '', '반차': '', '공가': ''}


@eel.expose
def save_vacation_data(date: str, data: Dict, username: str) -> Dict[str, Any]:
    """날짜별 휴가자 현황 저장"""
    try:
        success = db.save_vacation_records(date, data, username)
        return {'success': success}
    except Exception as e:
        logger.error(f"휴가자 현황 저장 오류: {e}")
        return {'success': False}



@eel.expose
def search_records_by_ship(ship_name: str) -> List[Dict[str, Any]]:
    """선명으로 작업 내역 조회"""
    try:
        if not ship_name or ship_name.strip() == '':
            return []

        query = '''
            SELECT date, company, ship_name, engine_model, work_content,
                   leader, manpower, teammates
            FROM work_records
            WHERE ship_name = ?
            AND ship_name IS NOT NULL
            AND ship_name != ''
            ORDER BY date, record_number
        '''

        logger.info(f"선명 검색: {ship_name}")
        results = db.execute_query(query, (ship_name.strip().upper(),))
        logger.info(f"검색 결과: {len(results) if results else 0}건")

        records = []
        for row in results:
            records.append({
                'date': row[0] or '',
                'company': row[1] or '',
                'ship_name': row[2] or '',
                'engine_model': row[3] or '',
                'work_content': row[4] or '',
                'leader': row[5] or '',
                'manpower': float(row[6]) if row[6] else 0.0,
                'teammates': row[7] or ''
            })

        return records
    except Exception as e:
        logger.error(f"선명 검색 오류: {e}")
        return []


@eel.expose
def get_project_end_date(ship_name: str) -> str:
    """선박별 공사 마지막 작업일 조회"""
    try:
        if not ship_name or ship_name.strip() == '':
            return ''
        
        # 해당 선박의 마지막 작업일 조회
        query = '''
            SELECT MAX(date) as end_date
            FROM work_records
            WHERE ship_name = ?
            AND ship_name != ''
        '''
        
        result = db.execute_query(query, (ship_name,))
        if result and result[0][0]:
            from datetime import datetime
            end_date_str = result[0][0]  # YYYY-MM-DD 형식
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            # M/D 형식으로 반환
            return f"{end_date.month}/{end_date.day}"
        return ''
    except Exception as e:
        logger.error(f"공사 마지막일 조회 오류: {e}")
        return ''


@eel.expose
def debug_check_data(year: int, month: int) -> Dict[str, Any]:
    """디버깅용 데이터 확인"""
    try:
        # 전체 데이터 조회
        query1 = '''
            SELECT COUNT(*) as total
            FROM work_records
        '''
        total_result = db.execute_query(query1, ())
        total_count = total_result[0][0] if total_result else 0
        
        # 해당 월 데이터 조회
        year_str = str(year)
        month_str = str(month).zfill(2)
        
        query2 = '''
            SELECT COUNT(*) as month_total
            FROM work_records
            WHERE strftime('%Y', date) = ?
            AND strftime('%m', date) = ?
        '''
        month_result = db.execute_query(query2, (year_str, month_str))
        month_count = month_result[0][0] if month_result else 0
        
        # 모든 날짜 목록 조회
        query3 = '''
            SELECT DISTINCT date
            FROM work_records
            ORDER BY date DESC
            LIMIT 10
        '''
        dates_result = db.execute_query(query3, ())
        dates = [row[0] for row in dates_result] if dates_result else []
        
        return {
            'total_records': total_count,
            'month_records': month_count,
            'recent_dates': dates,
            'query_year': year_str,
            'query_month': month_str
        }
    except Exception as e:
        logger.error(f"디버깅 데이터 확인 오류: {e}")
        return {}


@eel.expose
def load_monthly_report_grouped(year: int, month: int) -> List[Dict[str, Any]]:
    """월간 보고서 - 선박별 그룹핑"""
    try:
        from datetime import datetime
        
        logger.info(f"월간 보고 요청: {year}년 {month}월")
        
        # 해당 월 작업 기록 조회 + 전체 기간(시작~종료) 및 보드 상태 서브쿼리
        query = '''
            SELECT
                wr.company,
                wr.ship_name,
                wr.location,
                wr.engine_model,
                wr.work_content,
                wr.leader,
                wr.teammates,
                SUM(wr.manpower) as total_manpower,
                (SELECT MIN(date) FROM work_records
                 WHERE ship_name = wr.ship_name AND ship_name != '') as true_start_date,
                (SELECT MAX(date) FROM work_records
                 WHERE ship_name = wr.ship_name AND ship_name != '') as true_end_date,
                COALESCE(
                    (SELECT ps.status FROM project_status ps
                     WHERE ps.contract_number IN (
                         SELECT DISTINCT wr2.contract_number FROM work_records wr2
                         WHERE wr2.ship_name = wr.ship_name AND wr2.contract_number != ''
                     )
                     ORDER BY ps.updated_at DESC LIMIT 1),
                    (SELECT bp.status FROM board_projects bp
                     WHERE bp.ship_name = wr.ship_name ORDER BY bp.id DESC LIMIT 1)
                ) as board_status
            FROM work_records wr
            WHERE strftime('%Y', wr.date) = ?
            AND strftime('%m', wr.date) = ?
            AND wr.ship_name != ''
            GROUP BY wr.ship_name
            ORDER BY MIN(wr.date)
        '''

        year_str = str(year)
        month_str = str(month).zfill(2)
        logger.info(f"쿼리 파라미터: year={year_str}, month={month_str}")

        results = db.execute_query(query, (year_str, month_str))
        logger.info(f"조회 결과: {len(results) if results else 0}건")

        if not results:
            return []

        grouped_data = []
        for row in results:
            ship_name = row[1]

            # 공사기간: 전체 프로젝트 기간 기준, 보드 상태로 준공 여부 판단
            true_start_date = row[8] or ''
            true_end_date   = row[9] or ''
            board_status    = row[10] or ''

            if board_status == '준공':
                project_period = f"{true_start_date} ~ {true_end_date}"
            else:
                project_period = f"{true_start_date} ~ 진행중" if true_start_date else "진행중"
            
            # 작업내용 통합
            engine_model = row[3] or ''
            work_content = row[4] or ''
            
            if engine_model and work_content:
                full_work_content = f"{engine_model} {work_content}"
            elif engine_model:
                full_work_content = engine_model
            elif work_content:
                full_work_content = work_content
            else:
                full_work_content = '-'
            
            grouped_data.append({
                'company': row[0] or '',
                'ship_name': ship_name,
                'project_period': project_period,
                'location': row[2] or '',
                'work_content': full_work_content,
                'leader': row[5] or '',
                'teammates': row[6] or '',
                'total_manpower': float(row[7]) if row[7] else 0.0
            })
        
        return grouped_data
        
    except Exception as e:
        logger.error(f"월간 보고서 그룹핑 오류: {e}")
        return []


@eel.expose
def load_monthly_report(year: int, month: int) -> Dict[str, Any]:
    """월간 보고서 데이터 로드"""
    try:
        year_str = str(year)
        month_str = str(month).zfill(2)

        # 월 전체 데이터를 단일 쿼리로 조회 (N+1 제거)
        query = '''
            SELECT date, ship_name, work_content, company, manpower
            FROM work_records
            WHERE strftime('%Y', date) = ?
              AND strftime('%m', date) = ?
              AND (company != '' OR ship_name != '' OR work_content != '')
            ORDER BY date
        '''
        rows = db.execute_query(query, (year_str, month_str))

        # 날짜별로 그룹핑
        from collections import defaultdict
        day_groups: Dict[str, list] = defaultdict(list)
        for row in (rows or []):
            day_groups[row[0]].append(row)

        daily_data = []
        total_manpower = 0
        total_work_days = 0

        for date_str, day_rows in sorted(day_groups.items()):
            day = int(date_str.split('-')[2])
            day_manpower = sum(float(r[4] or 0) for r in day_rows)

            main_works = [r[1] for r in day_rows if r[1]]
            if not main_works:
                main_works = [r[2][:20] for r in day_rows if r[2]]

            daily_data.append({
                'date': date_str,
                'day': day,
                'work_count': len(day_rows),
                'manpower': day_manpower,
                'main_works': main_works[:3]
            })
            total_manpower += day_manpower
            total_work_days += 1

        avg_manpower = total_manpower / total_work_days if total_work_days > 0 else 0

        return {
            'total_work_days': total_work_days,
            'total_manpower': total_manpower,
            'avg_manpower': avg_manpower,
            'daily_data': daily_data
        }
    except Exception as e:
        logger.error(f"월간 보고서 로드 오류: {e}")
        return {
            'total_work_days': 0,
            'total_manpower': 0,
            'avg_manpower': 0,
            'daily_data': []
        }


@eel.expose
def export_to_excel(date: str) -> Dict[str, Any]:
    """Excel로 내보내기"""
    try:
        from pathlib import Path
        
        # 저장 경로 (바탕화면)
        desktop = Path.home() / "Desktop"
        filename = f"작업현황_{date}.xlsx"
        output_path = desktop / filename
        
        success = work_record_service.export_to_excel(date, str(output_path))
        
        return {
            'success': success,
            'path': str(output_path) if success else None,
            'message': f'Excel 파일이 저장되었습니다: {output_path}' if success else 'Excel 내보내기 실패'
        }
    except Exception as e:
        logger.error(f"Excel 내보내기 오류: {e}")
        return {'success': False, 'message': f'오류: {str(e)}'}


@eel.expose
def export_daily_report(date: str) -> Dict[str, Any]:
    """일일 보고서 Excel 내보내기"""
    try:
        from pathlib import Path
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from datetime import datetime
        
        # 데이터 로드
        records = work_record_service.get_records_for_date(date)
        valid_records = [r for r in records if r.get('company') or r.get('ship_name') or r.get('work_content')]
        
        # 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "일일보고"
        
        # 날짜 포맷
        date_obj = datetime.strptime(date, '%Y-%m-%d')
        weekdays = ['일', '월', '화', '수', '목', '금', '토']
        weekday = weekdays[date_obj.isoweekday() % 7]
        date_display = f"{date_obj.year}. {date_obj.month}월 {date_obj.day}일 ({weekday}) 현재"
        
        # 제목
        ws.merge_cells('A1:H1')
        ws['A1'] = '선진종합 선박별 작업 현황'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 날짜
        ws.merge_cells('A2:H2')
        ws['A2'] = date_display
        ws['A2'].font = Font(size=12)
        ws['A2'].alignment = Alignment(horizontal='right', vertical='center')
        
        # 헤더
        headers = ['No.', '선사', '선명', '공사기간', '장소', '작업내용', '담당공무', '협력업체']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')
        
        # 데이터
        for row_idx, record in enumerate(valid_records, start=4):
            ship_name = record.get('ship_name', '')
            engine_model = record.get('engine_model', '')
            work_content = record.get('work_content', '')
            
            # 공사기간: 시작일 ~ 진행중
            start_date = get_project_start_date(ship_name) if ship_name else ''
            project_period = f"{start_date} ~ 진행중" if start_date else '진행중'
            
            # 작업내용: 엔진모델 + 작업내용
            if engine_model and work_content:
                full_work_content = f"{engine_model} {work_content}"
            elif engine_model:
                full_work_content = engine_model
            elif work_content:
                full_work_content = work_content
            else:
                full_work_content = ''
            
            # 본사/외주 분리
            leader = record.get('leader', '')
            teammates = record.get('teammates', '')
            in_house, outsourced = separate_workers(leader, teammates)
            
            ws.cell(row=row_idx, column=1, value=row_idx - 3)
            ws.cell(row=row_idx, column=2, value=record.get('company', ''))
            ws.cell(row=row_idx, column=3, value=ship_name)
            ws.cell(row=row_idx, column=4, value=project_period)
            ws.cell(row=row_idx, column=5, value=record.get('location', ''))
            ws.cell(row=row_idx, column=6, value=full_work_content)
            ws.cell(row=row_idx, column=7, value=in_house)
            ws.cell(row=row_idx, column=8, value=outsourced)
        
        # 셀 정렬 및 테두리
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=3, max_row=3+len(valid_records), min_col=1, max_col=8):
            for cell in row:
                cell.border = thin_border
                if cell.row > 3:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 20
        
        # 저장
        desktop = Path.home() / "Desktop"
        filename = f"일일보고_{date}.xlsx"
        output_path = desktop / filename
        wb.save(output_path)
        
        return {
            'success': True,
            'filename': str(output_path)
        }
    except Exception as e:
        logger.error(f"일일 보고서 내보내기 오류: {e}")
        return {'success': False, 'message': str(e)}


@eel.expose
def export_monthly_report(year: int, month: int) -> Dict[str, Any]:
    """월간 보고서 Excel 내보내기"""
    try:
        from pathlib import Path
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from calendar import monthrange
        
        # 데이터 로드
        report_data = load_monthly_report(year, month)
        
        # 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "월간보고"
        
        # 제목
        ws.merge_cells('A1:D1')
        ws['A1'] = f'{year}년 {month}월 작업 현황'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 통계
        ws.merge_cells('A2:D2')
        ws['A2'] = f"총 작업일수: {report_data['total_work_days']}일 | 총 투입인원: {report_data['total_manpower']:.1f}공 | 평균: {report_data['avg_manpower']:.1f}공/일"
        ws['A2'].font = Font(size=12)
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 헤더
        headers = ['날짜', '작업건수', '투입인원', '주요작업']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='D9B3FF', end_color='D9B3FF', fill_type='solid')
        
        # 데이터
        for row_idx, day_data in enumerate(report_data['daily_data'], start=4):
            date_obj = datetime.strptime(day_data['date'], '%Y-%m-%d')
            weekdays = ['일', '월', '화', '수', '목', '금', '토']
            weekday = weekdays[date_obj.isoweekday() % 7]
            
            ws.cell(row=row_idx, column=1, value=f"{month}/{day_data['day']} ({weekday})")
            ws.cell(row=row_idx, column=2, value=day_data['work_count'])
            ws.cell(row=row_idx, column=3, value=f"{day_data['manpower']:.1f}")
            ws.cell(row=row_idx, column=4, value=', '.join(day_data['main_works']))
        
        # 테두리
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=3, max_row=3+len(report_data['daily_data']), min_col=1, max_col=4):
            for cell in row:
                cell.border = thin_border
                if cell.row > 3:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 열 너비
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 50
        
        # 저장
        desktop = Path.home() / "Desktop"
        filename = f"월간보고_{year}{month:02d}.xlsx"
        output_path = desktop / filename
        wb.save(output_path)
        
        return {
            'success': True,
            'filename': str(output_path)
        }
    except Exception as e:
        logger.error(f"월간 보고서 내보내기 오류: {e}")
        return {'success': False, 'message': str(e)}


# ============================================================================
# 클라우드 동기화
# ============================================================================

@eel.expose
def sync_to_cloud() -> Dict[str, Any]:
    """클라우드로 동기화"""
    try:
        if not cloud_sync.enabled:
            return {'success': False, 'message': '클라우드 동기화가 비활성화되어 있습니다.'}
        
        success = cloud_sync.sync_to_cloud()
        return {
            'success': success,
            'message': '클라우드 동기화 완료' if success else '클라우드 동기화 실패'
        }
    except Exception as e:
        logger.error(f"클라우드 동기화 오류: {e}")
        return {'success': False, 'message': f'오류: {str(e)}'}


@eel.expose
def sync_from_cloud() -> Dict[str, Any]:
    """클라우드에서 동기화"""
    try:
        if not cloud_sync.enabled:
            return {'success': False, 'message': '클라우드 동기화가 비활성화되어 있습니다.'}
        
        success = cloud_sync.sync_from_cloud()
        return {
            'success': success,
            'message': '클라우드에서 동기화 완료' if success else '클라우드에서 동기화 실패'
        }
    except Exception as e:
        logger.error(f"클라우드 동기화 오류: {e}")
        return {'success': False, 'message': f'오류: {str(e)}'}


@eel.expose
def get_sync_status() -> Dict[str, Any]:
    """동기화 상태 조회"""
    try:
        return cloud_sync.get_sync_status()
    except Exception as e:
        logger.error(f"동기화 상태 조회 오류: {e}")
        return {}


# ============================================================================
# 앱 정보
# ============================================================================

@eel.expose
def get_app_info() -> Dict[str, Any]:
    """앱 정보 조회"""
    return {
        'name': config.app_name,
        'version': config.version,
        'db_path': str(config.db_path),
        'cloud_sync_enabled': cloud_sync.enabled,
        'cloud_folder': str(cloud_sync.cloud_folder) if cloud_sync.cloud_folder else None
    }


@eel.expose
def get_activity_logs(limit: int = 50) -> List[Dict[str, Any]]:
    """활동 로그 조회"""
    try:
        logs = db.get_activity_logs(limit=limit)
        return [log.to_dict() for log in logs]
    except Exception as e:
        logger.error(f"활동 로그 조회 오류: {e}")
        return []


# ============================================================================
# 대시보드 - 간트 차트
# ============================================================================

@eel.expose
def get_gantt_data(year: int, month: int) -> List[Dict[str, Any]]:
    """간트 차트용 프로젝트 데이터 조회 (해당 월과 겹치는 모든 프로젝트)"""
    try:
        import calendar
        last_day = calendar.monthrange(year, month)[1]
        month_start = f'{year}-{month:02d}-01'
        month_end = f'{year}-{month:02d}-{last_day:02d}'

        logger.info(f"간트 데이터 조회: {month_start} ~ {month_end}")

        # 해당 월과 기간이 겹치는 프로젝트 조회
        query = '''
            SELECT contract_number, company, ship_name, engine_model, work_content,
                   MIN(date) as start_date, MAX(date) as end_date,
                   COUNT(DISTINCT date) as work_days,
                   ROUND(SUM(manpower), 1) as total_manpower
            FROM work_records
            WHERE contract_number != '' AND contract_number IS NOT NULL
            GROUP BY contract_number
            HAVING end_date >= ? AND start_date <= ?
            ORDER BY start_date
        '''
        rows = db.execute_query(query, (month_start, month_end))

        projects = []
        for row in rows:
            contract_number = row[0] or ''

            # 해당 월 내 실제 작업일 조회
            dates_query = '''
                SELECT DISTINCT date FROM work_records
                WHERE contract_number = ? AND date >= ? AND date <= ?
                ORDER BY date
            '''
            date_rows = db.execute_query(dates_query, (contract_number, month_start, month_end))
            work_dates = [d[0] for d in date_rows] if date_rows else []

            start_date = row[5] or ''
            end_date = row[6] or ''
            # M/D 형식으로 변환
            start_md = ''
            end_md = ''
            if start_date:
                parts = start_date.split('-')
                start_md = f"{int(parts[1])}/{int(parts[2])}"
            if end_date:
                parts = end_date.split('-')
                end_md = f"{int(parts[1])}/{int(parts[2])}"

            projects.append({
                'contractNumber': contract_number,
                'company': row[1] or '',
                'shipName': row[2] or '',
                'engineModel': row[3] or '',
                'workContent': row[4] or '',
                'startDate': start_date,
                'endDate': end_date,
                'startMD': start_md,
                'endMD': end_md,
                'workDays': row[7] or 0,
                'totalManpower': row[8] or 0,
                'workDates': work_dates
            })

        logger.info(f"간트 데이터: {len(projects)}개 프로젝트")
        return projects

    except Exception as e:
        logger.error(f"간트 데이터 조회 오류: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return []


# ============================================================================
# 대시보드 - 칸반 보드
# ============================================================================

@eel.expose
def set_project_status(contract_number: str, status: str) -> Dict[str, Any]:
    """프로젝트 상태 수동 변경 (접수/착수/준공/auto)"""
    try:
        username = ''
        success = db.set_project_status(contract_number, status, username)
        if success:
            return {'success': True, 'message': f'상태가 변경되었습니다.'}
        return {'success': False, 'message': '상태 변경 실패'}
    except Exception as e:
        logger.error(f"프로젝트 상태 변경 오류: {e}")
        return {'success': False, 'message': str(e)}


@eel.expose
def create_board_project(data: Dict) -> Dict[str, Any]:
    """보드 프로젝트 생성 (접수 단계)"""
    try:
        username = data.get('username', '')
        project_id = db.create_board_project(data, username)
        if project_id:
            return {'success': True, 'id': project_id, 'message': '프로젝트가 등록되었습니다.'}
        return {'success': False, 'message': '프로젝트 등록 실패'}
    except Exception as e:
        logger.error(f"보드 프로젝트 생성 오류: {e}")
        return {'success': False, 'message': str(e)}


@eel.expose
def update_board_project(project_id: int, data: Dict) -> Dict[str, Any]:
    """보드 프로젝트 업데이트"""
    try:
        username = data.get('username', '')
        success = db.update_board_project(project_id, data, username)
        if success:
            return {'success': True, 'message': '프로젝트가 수정되었습니다.'}
        return {'success': False, 'message': '프로젝트 수정 실패'}
    except Exception as e:
        logger.error(f"보드 프로젝트 수정 오류: {e}")
        return {'success': False, 'message': str(e)}


@eel.expose
def delete_board_project(project_id: int) -> Dict[str, Any]:
    """보드 프로젝트 삭제"""
    try:
        success = db.delete_board_project(project_id)
        if success:
            return {'success': True, 'message': '프로젝트가 삭제되었습니다.'}
        return {'success': False, 'message': '프로젝트 삭제 실패'}
    except Exception as e:
        logger.error(f"보드 프로젝트 삭제 오류: {e}")
        return {'success': False, 'message': str(e)}


@eel.expose
def get_kanban_data() -> Dict[str, Any]:
    """칸반 보드용 프로젝트 데이터 (접수/착수/준공/아카이브 4단계)"""
    try:
        from datetime import datetime, timedelta

        now = datetime.now()
        cutoff = (now - timedelta(days=7)).strftime('%Y-%m-%d')
        current_month_start = now.strftime('%Y-%m-01')

        # 1. 접수 단계 프로젝트 (board_projects 테이블에서)
        reception_projects = db.get_board_projects('접수')
        reception = []
        for bp in reception_projects:
            reception.append({
                'id': bp['id'],
                'contractNumber': bp.get('contract_number', ''),
                'company': bp.get('company', ''),
                'shipName': bp.get('ship_name', ''),
                'engineModel': bp.get('engine_model', ''),
                'workContent': bp.get('work_content', ''),
                'startDate': '',
                'endDate': '',
                'startMD': '',
                'endMD': '',
                'workDays': 0,
                'totalManpower': 0,
                'status': '접수',
                'source': 'board'
            })

        # 2. 착수/준공 - work_records 기반 + board_projects 수동 상태
        query = '''
            SELECT contract_number, company, ship_name, engine_model, work_content,
                   MIN(date) as start_date, MAX(date) as end_date,
                   COUNT(DISTINCT date) as work_days,
                   ROUND(SUM(manpower), 1) as total_manpower
            FROM work_records
            WHERE contract_number != '' AND contract_number IS NOT NULL
            GROUP BY contract_number
            ORDER BY contract_number DESC
        '''
        rows = db.execute_query(query)

        # 수동 상태 조회 (project_status 테이블)
        manual_statuses = db.get_project_statuses()

        # board_projects에서 착수/준공 상태인 것도 조회
        board_started = {bp['contract_number']: bp for bp in db.get_board_projects('착수') if bp.get('contract_number')}
        board_done = {bp['contract_number']: bp for bp in db.get_board_projects('준공') if bp.get('contract_number')}

        started = []
        done = []
        archive = []

        for row in rows:
            contract_number = row[0] or ''
            end_date = row[6] or ''
            start_date = row[5] or ''

            # M/D 형식
            start_md = ''
            end_md = ''
            if start_date:
                parts = start_date.split('-')
                start_md = f"{int(parts[1])}/{int(parts[2])}"
            if end_date:
                parts = end_date.split('-')
                end_md = f"{int(parts[1])}/{int(parts[2])}"

            # 상태 판단: 수동 > 자동
            manual_status = manual_statuses.get(contract_number, '')
            board_status = ''
            if contract_number in board_started:
                board_status = '착수'
            elif contract_number in board_done:
                board_status = '준공'

            # 우선순위: project_status > board_projects > 자동
            if manual_status and manual_status != 'auto':
                # 기존 project_status의 값 매핑
                if manual_status in ('inProgress', '착수'):
                    final_status = '착수'
                elif manual_status in ('completed', '준공'):
                    final_status = '준공'
                else:
                    final_status = manual_status
            elif board_status:
                final_status = board_status
            else:
                # 자동 판단
                if end_date >= cutoff:
                    final_status = '착수'
                else:
                    final_status = '준공'

            project = {
                'contractNumber': contract_number,
                'company': row[1] or '',
                'shipName': row[2] or '',
                'engineModel': row[3] or '',
                'workContent': row[4] or '',
                'startDate': start_date,
                'endDate': end_date,
                'startMD': start_md,
                'endMD': end_md,
                'workDays': row[7] or 0,
                'totalManpower': row[8] or 0,
                'status': final_status,
                'source': 'records'
            }

            if final_status == '착수':
                started.append(project)
            elif final_status == '준공':
                # 준공인데 해당 월 지남 → 아카이브
                if end_date and end_date < current_month_start:
                    project['status'] = '아카이브'
                    archive.append(project)
                else:
                    done.append(project)

        logger.info(f"칸반 데이터: 접수 {len(reception)}건, 착수 {len(started)}건, 준공 {len(done)}건, 아카이브 {len(archive)}건")
        return {
            'reception': reception,
            'started': started,
            'done': done,
            'archive': archive
        }

    except Exception as e:
        logger.error(f"칸반 데이터 조회 오류: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return {'reception': [], 'started': [], 'done': [], 'archive': []}


# ============================================================================
# 댓글 시스템
# ============================================================================

@eel.expose
def add_project_comment(contract_number: str, content: str, parent_id: int = None, board_project_id: int = None) -> Dict[str, Any]:
    """프로젝트 댓글 추가 (add_project_comment_with_user로 위임)"""
    return add_project_comment_with_user(contract_number, content, '', '', parent_id, board_project_id)


@eel.expose
def add_project_comment_with_user(contract_number, content, user_id, user_name, parent_id=None, board_project_id=None) -> Dict[str, Any]:
    """프로젝트 댓글 추가 (사용자 정보 포함)"""
    try:
        # Eel null/0 처리
        if not parent_id or parent_id == 'null':
            parent_id = None
        else:
            parent_id = int(parent_id)
        if not board_project_id or board_project_id == 'null':
            board_project_id = None
        else:
            board_project_id = int(board_project_id)
        if not contract_number or contract_number == 'null':
            contract_number = ''
        # user_id/user_name 방어 (NOT NULL 제약 대응)
        if not user_id or user_id == 'null':
            user_id = 'unknown'
        if not user_name or user_name == 'null':
            user_name = '익명'
        comment_id = db.add_comment(contract_number, user_id, user_name, content, parent_id, board_project_id)
        if comment_id:
            # 텔레그램 알림 발송 (실패해도 댓글은 정상 등록)
            try:
                ship_name = telegram_notifier._get_ship_name(contract_number, board_project_id)
                telegram_notifier.send_comment_notification(
                    contract_number=contract_number or '',
                    board_project_id=board_project_id,
                    commenter_user_id=user_id,
                    commenter_name=user_name,
                    comment_text=content,
                    ship_name=ship_name
                )
            except Exception as te:
                logger.error(f"텔레그램 알림 발송 실패: {te}")
            return {'success': True, 'id': comment_id, 'message': '댓글이 등록되었습니다.'}
        return {'success': False, 'message': '댓글 등록 실패'}
    except Exception as e:
        logger.error(f"댓글 추가 오류: {e}")
        return {'success': False, 'message': str(e)}


@eel.expose
def get_project_comments(contract_number=None, board_project_id=None) -> Dict[str, Any]:
    """프로젝트 댓글 조회"""
    try:
        # Eel에서 null/0 → None 변환 처리
        if not contract_number or contract_number == 'null':
            contract_number = None
        if not board_project_id or board_project_id == 'null':
            board_project_id = None
        else:
            board_project_id = int(board_project_id)
        comments = db.get_comments(contract_number, board_project_id)
        # camelCase 변환
        result = []
        for c in comments:
            result.append({
                'id': c['id'],
                'contractNumber': c.get('contract_number', ''),
                'boardProjectId': c.get('board_project_id'),
                'parentId': c.get('parent_id'),
                'userId': c.get('user_id', ''),
                'userName': c.get('user_name', ''),
                'content': c.get('content', ''),
                'createdAt': c.get('created_at', '')
            })
        return {'success': True, 'comments': result}
    except Exception as e:
        logger.error(f"댓글 조회 오류: {e}")
        return {'success': False, 'comments': []}


@eel.expose
def delete_project_comment(comment_id: int, user_id: str) -> Dict[str, Any]:
    """프로젝트 댓글 삭제"""
    try:
        success = db.delete_comment(comment_id, user_id)
        if success:
            return {'success': True, 'message': '댓글이 삭제되었습니다.'}
        return {'success': False, 'message': '삭제 권한이 없거나 댓글을 찾을 수 없습니다.'}
    except Exception as e:
        logger.error(f"댓글 삭제 오류: {e}")
        return {'success': False, 'message': str(e)}


# ============================================================================
# 엑셀 불러오기
# ============================================================================

@eel.expose
def import_excel_data(base64_data: str, username: str = 'admin') -> Dict[str, Any]:
    """엑셀 파일 데이터를 DB로 일괄 업로드"""
    try:
        import base64
        import io
        from openpyxl import load_workbook
        from ..database.models import WorkRecord
        from ..business.calculations import calculate_record_manpower

        logger.info(f"엑셀 불러오기 시작 by {username}")

        # 셀 값 추출 헬퍼 (기울임체 감지 포함)
        def get_cell_value(cell, wrap_italic=False):
            """셀 값을 문자열로 추출. wrap_italic=True이면 기울임체 셀에 *이름* 형태로 반환"""
            if cell is None or cell.value is None:
                return ''
            val = str(cell.value).strip()
            if not val:
                return ''
            if wrap_italic and cell.font and cell.font.italic:
                return f'*{val}*'
            return val

        # base64 디코딩
        file_bytes = base64.b64decode(base64_data)

        # 파일 형식 감지: OLE2 magic bytes = .xls, ZIP header = .xlsx
        XLS_MAGIC = b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'
        is_xls = file_bytes[:8] == XLS_MAGIC

        if is_xls:
            # .xls (Excel 97-2003) → xlrd 사용 (이탤릭 감지 불가)
            try:
                import xlrd
            except ImportError:
                return {'success': False, 'message': 'xlrd 패키지가 필요합니다. 터미널에서 [pip install xlrd] 를 실행해 주세요.'}
            wb_xls = xlrd.open_workbook(file_contents=file_bytes)
            ws_xls = wb_xls.sheet_by_index(0)
            # 날짜 타입 셀(type 3)을 "날짜 : YYYY.MM.DD." 텍스트로 변환하여
            # 날짜 감지 정규식이 float 값을 놓치지 않도록 처리
            rows_raw = []
            for i in range(ws_xls.nrows):
                vals = list(ws_xls.row_values(i))
                types = ws_xls.row_types(i)
                for j, (v, t) in enumerate(zip(vals, types)):
                    if t == 3:  # XL_CELL_DATE
                        try:
                            dt = xlrd.xldate.xldate_as_datetime(v, wb_xls.datemode)
                            vals[j] = f"날짜 : {dt.strftime('%Y.%m.%d')}."
                        except Exception:
                            pass
                rows_raw.append(vals)
            logger.info(f"xls 파일 감지: {ws_xls.nrows}행")
        else:
            # .xlsx (Office Open XML) → openpyxl 사용
            wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
            ws = wb.active
            rows_raw = [[c.value for c in row] for row in ws.iter_rows()]
            wb.close()
            logger.info(f"xlsx 파일 감지: {len(rows_raw)}행")

        # 날짜별 레코드 그룹핑
        date_records = {}  # { 'YYYY-MM-DD': [WorkRecord, ...] }
        skipped = 0
        total_records = 0

        import re
        current_date = None

        for row_values in rows_raw:
            # row_values: list of cell values (str/float/None)
            row_text = ' '.join(str(v or '') for v in row_values)

            # 날짜 헤더 행 감지: "날짜 : YYYY.MM.DD" 형식 (구분자 유연화)
            date_match = re.search(
                r'날짜\s*[:\：]?\s*(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})',
                row_text
            )
            if not date_match:
                # "날짜 :" 없이 YYYY.MM.DD 단독 셀인 경우
                date_match = re.search(
                    r'^\s*(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})\s*$',
                    row_text.strip()
                )
            if date_match:
                y, m, d = date_match.groups()
                current_date = f'{y}-{int(m):02d}-{int(d):02d}'
                continue

            # 컬럼 헤더 행 스킵 (계약번호, 선사/선명 포함 행)
            if '계약번호' in row_text or ('선사' in row_text and '선명' in row_text):
                continue

            # 날짜가 아직 설정되지 않은 경우 건너뜀
            if current_date is None:
                continue

            # B열(인덱스 1) 이상이어야 데이터 행으로 처리
            if len(row_values) < 2:
                skipped += 1
                continue

            def _cell_str(val):
                """셀 값을 문자열로 변환 (None/숫자 안전)"""
                if val is None:
                    return ''
                # xlrd가 숫자형으로 반환하는 경우 정수 변환
                if isinstance(val, float) and val == int(val):
                    return str(int(val))
                return str(val).strip()

            # B~I열 매핑 (실제 엑셀 형식)
            # B(1)=계약번호, C(2)=선사, D(3)=선명, E(4)=엔진/모델
            # F(5)=작업내용/장소, G(6)=작업자, H(7)=인원(무시), I(8)=동반자
            contract_number = _cell_str(row_values[1]) if len(row_values) > 1 else ''
            company         = _cell_str(row_values[2]) if len(row_values) > 2 else ''
            ship_name       = _cell_str(row_values[3]).upper() if len(row_values) > 3 else ''
            engine_model    = _cell_str(row_values[4]).upper() if len(row_values) > 4 else ''
            work_content    = _cell_str(row_values[5]) if len(row_values) > 5 else ''
            leader_raw      = _cell_str(row_values[6]) if len(row_values) > 6 else ''
            leader          = leader_raw  # xls/xlsx 공통: 이탤릭 감지는 xlsx만 가능
            # H열(인덱스 7): 인원 수 — 자동 계산으로 대체, 건너뜀
            teammates       = _cell_str(row_values[8]) if len(row_values) > 8 else ''

            # 유효 데이터 확인 (빈 행 제외)
            if not any([contract_number, company, ship_name, work_content]):
                skipped += 1
                continue

            # 인원 계산
            manpower = calculate_record_manpower(leader, teammates)

            # 날짜별 그룹핑
            if current_date not in date_records:
                date_records[current_date] = []

            record = WorkRecord(
                date=current_date,
                record_number=len(date_records[current_date]) + 1,
                contract_number=contract_number,
                company=company,
                ship_name=ship_name,
                engine_model=engine_model,
                work_content=work_content,
                location='',
                leader=leader,
                teammates=teammates,
                manpower=manpower
            )
            date_records[current_date].append(record)
            total_records += 1

        # DB에 날짜별로 저장
        saved_dates = 0
        for date_str, records in date_records.items():
            success = db.save_work_records(date_str, records, username)
            if success:
                saved_dates += 1
            else:
                logger.error(f"날짜 {date_str} 저장 실패")

        logger.info(f"엑셀 불러오기 완료: {saved_dates}일, {total_records}건")

        return {
            'success': True,
            'total_dates': saved_dates,
            'total_records': total_records,
            'skipped': skipped
        }

    except Exception as e:
        logger.error(f"엑셀 불러오기 오류: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return {
            'success': False,
            'message': f'엑셀 불러오기 실패: {str(e)}'
        }


# ============================================================================
# 업데이트
# ============================================================================

@eel.expose
def get_startup_patch_result() -> Dict[str, Any]:
    """앱 시작 시 자동 적용된 패치 결과 반환 (로그인 후 JS에서 호출)"""
    try:
        from src.utils.patch_system import patch_system
        applied = getattr(patch_system, '_startup_patches_applied', 0)
        version = config.version

        # 자동 재시작 후 첫 실행 시 마커 파일에서 업데이트 정보 읽기
        import json as _json
        from pathlib import Path as _Path
        marker_path = _Path(__file__).parent.parent.parent / "data" / "just_updated.json"
        if marker_path.exists():
            try:
                data = _json.loads(marker_path.read_text(encoding='utf-8'))
                applied = data.get('applied_count', applied)
                version = data.get('version', version)
                marker_path.unlink()  # 1회 읽고 삭제
                logger.info(f"자동 재시작 마커 확인: v{version} 패치 {applied}개")
            except Exception as me:
                logger.error(f"재시작 마커 읽기 오류: {me}")

        return {
            'needs_restart': False,  # 이미 재시작 완료
            'applied_count': applied,
            'current_version': version
        }
    except Exception as e:
        return {'needs_restart': False, 'applied_count': 0, 'current_version': config.version}


@eel.expose
def check_for_updates(force: bool = False) -> Dict[str, Any]:
    """업데이트 확인"""
    try:
        result = update_manager.check_for_updates(force=force)
        return result
    except Exception as e:
        logger.error(f"업데이트 확인 오류: {e}")
        return {
            'update_available': False,
            'error': str(e)
        }


@eel.expose
def download_and_apply_patches() -> Dict[str, Any]:
    """패치 ZIP 다운로드 및 적용"""
    try:
        result = update_manager.download_and_apply_patches()
        return result
    except Exception as e:
        logger.error(f"패치 적용 오류: {e}")
        return {
            'success': False,
            'message': str(e)
        }


@eel.expose
def get_release_notes_for_version(version_tag: str) -> Dict[str, Any]:
    """특정 버전의 릴리즈 노트 조회"""
    try:
        notes = update_manager.get_release_notes(version_tag)
        if notes:
            return {
                'success': True,
                'release_notes': notes
            }
        else:
            return {
                'success': False,
                'message': '릴리즈 노트를 찾을 수 없습니다.'
            }
    except Exception as e:
        logger.error(f"릴리즈 노트 조회 오류: {e}")
        return {
            'success': False,
            'error': str(e)
        }


# ============================================================================
# 공휴일 데이터
# ============================================================================

@eel.expose
def get_holidays() -> Dict[str, str]:
    """공휴일 데이터 로드 (config/holidays.json)"""
    try:
        import json
        from pathlib import Path
        holidays_path = Path(__file__).parent.parent.parent / "config" / "holidays.json"
        with open(holidays_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        logger.error(f"공휴일 데이터 로드 오류: {e}")
        return {}


# ============================================================================
# 텔레그램 알림 설정
# ============================================================================

@eel.expose
def generate_telegram_link_code(user_id: str) -> Dict[str, Any]:
    """텔레그램 연결 코드 생성"""
    try:
        code = auth_manager.generate_link_code(user_id)
        if code:
            bot_username = telegram_notifier.get_bot_username()
            deep_link = f'https://t.me/{bot_username}?start={code}' if bot_username else ''
            return {
                'success': True,
                'code': code,
                'botUsername': bot_username,
                'deepLink': deep_link
            }
        return {'success': False, 'message': '코드 생성 실패'}
    except Exception as e:
        logger.error(f"텔레그램 코드 생성 오류: {e}")
        return {'success': False, 'message': str(e)}


@eel.expose
def unlink_telegram(user_id: str) -> Dict[str, Any]:
    """텔레그램 연결 해제"""
    try:
        success = auth_manager.unlink_telegram(user_id)
        return {'success': success}
    except Exception as e:
        logger.error(f"텔레그램 연결 해제 오류: {e}")
        return {'success': False, 'message': str(e)}


@eel.expose
def get_telegram_status(user_id: str) -> Dict[str, Any]:
    """텔레그램 연결 상태 조회"""
    try:
        return auth_manager.get_telegram_status(user_id)
    except Exception as e:
        logger.error(f"텔레그램 상태 조회 오류: {e}")
        return {'linked': False}


@eel.expose
def get_telegram_bot_enabled() -> Dict[str, Any]:
    """텔레그램 봇 활성화 상태 조회"""
    try:
        # config 우선, 없으면 DB 폴백 (PC2처럼 settings.json에 토큰 없는 환경 대응)
        bot_token = config.get('telegram.bot_token', '')
        if not bot_token:
            bot_token = db.get_setting('telegram.bot_token', '') or ''
            if bot_token:
                # DB에서 로드한 값을 config에도 반영 (이후 호출 빠르게)
                config.set('telegram.bot_token', bot_token)
                logger.info("get_telegram_bot_enabled: DB app_settings에서 토큰 로드")
        enabled = config.get('telegram.enabled', False)
        return {
            'success': True,
            'enabled': enabled,
            'hasToken': bool(bot_token),
            'botUsername': telegram_notifier.get_bot_username() if enabled else ''
        }
    except Exception as e:
        logger.error(f"텔레그램 봇 상태 조회 오류: {e}")
        return {'success': False, 'enabled': False, 'hasToken': False}


@eel.expose
def admin_save_telegram_settings(bot_token: str, enabled: bool, admin_id: str) -> Dict[str, Any]:
    """텔레그램 봇 설정 저장 (관리자)"""
    try:
        import requests as _requests
        # 토큰 유효성 검사 (비어있지 않을 경우)
        if bot_token and bot_token.strip():
            try:
                r = _requests.get(
                    f'https://api.telegram.org/bot{bot_token.strip()}/getMe',
                    timeout=5
                )
                data = r.json()
                if not r.ok or not data.get('ok'):
                    err_desc = data.get('description', '알 수 없는 오류')
                    logger.warning(f"텔레그램 토큰 유효성 검사 실패: {err_desc}")
                    return {
                        'success': False,
                        'message': f'유효하지 않은 봇 토큰입니다: {err_desc}\nBotFather에서 토큰을 확인하세요.'
                    }
                bot_name = data.get('result', {}).get('username', '')
                logger.info(f"텔레그램 토큰 유효: @{bot_name}")
            except _requests.exceptions.Timeout:
                logger.warning("텔레그램 토큰 검사 타임아웃 (네트워크 불안정) - 저장 계속 진행")
            except Exception as ve:
                logger.warning(f"텔레그램 토큰 검사 중 예외 (저장 계속): {ve}")

        config.set('telegram.bot_token', bot_token)
        config.set('telegram.enabled', enabled)
        config.save()
        # DB app_settings에도 동기화 (다른 PC에서 자동 로드용)
        db.set_setting('telegram.bot_token', bot_token)
        db.set_setting('telegram.enabled', 'true' if enabled else 'false')
        telegram_notifier.reconfigure(bot_token, enabled)
        logger.info(f"텔레그램 설정 변경: enabled={enabled}, by={admin_id}")
        return {'success': True, 'message': '텔레그램 설정이 저장되었습니다.'}
    except Exception as e:
        logger.error(f"텔레그램 설정 저장 오류: {e}")
        return {'success': False, 'message': str(e)}
